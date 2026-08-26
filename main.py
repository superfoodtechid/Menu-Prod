import os
import json
import sys
import uuid
import logging

# Force urllib3 to use IPv4 only because IPv6 is broken/blocked on some hosts and causes connection hangs
try:
    import urllib3.util.connection
    import socket
    urllib3.util.connection.allowed_gai_family = lambda: socket.AF_INET
except Exception:
    pass

import threading
import time
import builtins
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import List, Optional, Union, Dict, Any

# Override standard print to include timestamp
_original_print = builtins.print
def timestamped_print(*args, **kwargs):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _original_print(f"[{now_str}]", *args, **kwargs)
builtins.print = timestamped_print

from fastapi import FastAPI, Depends, HTTPException, BackgroundTasks, Query, status, Request, File, UploadFile, Form
from fastapi.responses import FileResponse, JSONResponse
from fastapi.security import APIKeyHeader
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session, joinedload

from dotenv import load_dotenv
load_dotenv()

# Setup Dynamic Paths to ensure reliable server deployment
BASE_DIR = Path(__file__).resolve().parent
sys.path.append(str(BASE_DIR))
sys.path.append(str(BASE_DIR / "menu_core"))

from menu_core.database import get_db, init_db, Account, Outlet, Job, AuditTrail

# Setup Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger("FoodMasterAPI")

# Initialize database tables on startup
app = FastAPI(
    title="FoodMaster Menu Portal API",
    description="Backend API for managing multi-platform menus (Shopee, Grab, GoFood)",
    version="1.0.0"
)

# Parse restricted CORS origins from environment variable
raw_origins = os.getenv(
    "ALLOWED_ORIGINS",
    "http://localhost:3000,http://localhost:5173,http://127.0.0.1:3000,http://127.0.0.1:5173,http://168.144.143.203:3000"
)
allowed_origins = [origin.strip() for origin in raw_origins.split(",") if origin.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# API Key Middleware enforcing authentication on all /api/ routes
@app.middleware("http")
async def api_key_middleware(request: Request, call_next):
    # Allow OPTIONS (CORS preflight requests), non-/api/ paths, and public download endpoint
    if request.method == "OPTIONS" or not request.url.path.startswith("/api/") or request.url.path == "/api/jobs/download-file":
        return await call_next(request)

    expected_key = os.getenv("API_SECRET_KEY", "foodmaster-secret-api-key-2026")
    if expected_key:
        api_key = request.headers.get("X-API-Key") or request.query_params.get("api_key")
        if not api_key or api_key != expected_key:
            return JSONResponse(
                status_code=status.HTTP_401_UNAUTHORIZED,
                content={"detail": "Invalid or missing X-API-Key header"}
            )

    return await call_next(request)

@app.on_event("startup")
def startup_event():
    from src.core.browser_factory import cleanup_zombie_chromium
    cleanup_zombie_chromium()
    logger.info("🚀 Initializing database tables...")
    init_db()
    # Ensure persistent shopee profile & session sync from /app/data
    try:
        import shutil
        persistent_data_dir = BASE_DIR / "data"
        shopee_auto_dir = BASE_DIR / "src" / "shopee-omzet-automation" / "data"
        shopee_core_dir = BASE_DIR / "shopee" / "data"
        shopee_auto_dir.mkdir(parents=True, exist_ok=True)
        shopee_core_dir.mkdir(parents=True, exist_ok=True)

        for src_f in persistent_data_dir.glob("session*.json"):
            for dst_dir in [shopee_auto_dir, shopee_core_dir]:
                dst_f = dst_dir / src_f.name
                if not dst_f.exists() or src_f.stat().st_mtime > dst_f.stat().st_mtime:
                    shutil.copy2(src_f, dst_f)
            if src_f.name == "session_allvbadmin.json":
                for dst_dir in [shopee_auto_dir, shopee_core_dir, persistent_data_dir]:
                    dst_sess = dst_dir / "session.json"
                    if not dst_sess.exists() or src_f.stat().st_mtime > dst_sess.stat().st_mtime:
                        shutil.copy2(src_f, dst_sess)

        for src_d in persistent_data_dir.glob("chrome_profile*"):
            if src_d.is_dir():
                for dst_dir in [shopee_auto_dir, shopee_core_dir]:
                    dst_prof = dst_dir / (src_d.name if "allvbadmin" in src_d.name else "chrome_profile")
                    if not dst_prof.exists():
                        shutil.copytree(src_d, dst_prof, dirs_exist_ok=True)
                # also mirror to standard chrome_profile name in shopee_core
                dst_shopee_prof = shopee_core_dir / "chrome_profile"
                if not dst_shopee_prof.exists():
                    shutil.copytree(src_d, dst_shopee_prof, dirs_exist_ok=True)
        logger.info("✅ Persistent Shopee sessions & Chrome profiles synced from /app/data")
    except Exception as err:
        logger.warning(f"⚠️ Shopee session persistence sync warning: {err}")

    # Ensure export directories exist dynamically
    exports_dir = BASE_DIR / "data" / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    logger.info(f"📂 Exports directory verified at: {exports_dir}")

    # Launch background auto-sync loop (syncs Google Sheets once every 24 hours)
    def _bg_auto_sync_loop():
        import time
        time.sleep(600)  # Grace period: wait 10 minutes after startup before initial background sync
        logger.info("🔄 Background GSheets auto-sync worker started (interval: 24 hours)")
        while True:
            try:
                from menu_core.database import SessionLocal
                db = SessionLocal()
                try:
                    sync_sheets(db)
                finally:
                    db.close()
            except Exception as ex:
                logger.warning(f"⚠️ Background auto-sync iteration error: {ex}")
            time.sleep(86400)  # Sync once every 24 hours



    import threading
    t = threading.Thread(target=_bg_auto_sync_loop, daemon=True)
    t.start()


# ─── PYDANTIC SCHEMAS ─────────────────────────────────────────────────────────

class AccountCreate(BaseModel):
    platform: str = Field(..., description="shopee | grab | gofood")
    username: str
    password: str
    portal: Optional[str] = None

class AccountResponse(BaseModel):
    id: uuid.UUID
    platform: str
    username: str
    password: Optional[str] = None
    portal: Optional[str] = None
    created_at: datetime

    class Config:
        from_attributes = True

class OutletCreate(BaseModel):
    account_id: uuid.UUID
    store_id: Optional[str] = None
    merchant_name: str = Field(..., description="Nama merchant / portal selector di web")
    owner: Optional[str] = None
    nama_outlet: Optional[str] = None
    cabang: Optional[str] = None
    nama_resto_final: Optional[str] = None
    brand: Optional[str] = None
    is_active: bool = True

class OutletResponse(BaseModel):
    id: uuid.UUID
    account_id: uuid.UUID
    store_id: Optional[str]
    merchant_name: str
    owner: Optional[str] = None
    nama_outlet: Optional[str]
    cabang: Optional[str]
    nama_resto_final: Optional[str]
    brand: Optional[str]
    is_active: bool
    last_sync_at: Optional[datetime]
    created_at: datetime
    platform: Optional[str] = None
    account: Optional[AccountResponse] = None

    class Config:
        from_attributes = True

class JobResponse(BaseModel):
    id: uuid.UUID
    outlet_id: Optional[uuid.UUID]
    job_type: str
    platform: str
    status: str
    progress_pct: int
    current_step: Optional[str]
    payload: Optional[dict]
    result_metadata: Optional[dict]
    error_message: Optional[str]
    created_by: str
    created_at: datetime
    started_at: Optional[datetime]
    completed_at: Optional[datetime]

    class Config:
        from_attributes = True

class AuditTrailResponse(BaseModel):
    id: uuid.UUID
    job_id: uuid.UUID
    outlet_id: uuid.UUID
    item_id: str
    item_name: str
    change_type: str
    field_changed: str
    old_value: Optional[str]
    new_value: str
    status: str
    error_message: Optional[str]
    created_at: datetime

    class Config:
        from_attributes = True


class PriceUpdateItem(BaseModel):
    item_id: str
    category_id: Optional[str] = ""
    item_name: Optional[str] = ""
    new_price: float

class PriceUpdateRequest(BaseModel):
    outlet_id: uuid.UUID
    updates: List[PriceUpdateItem]

class CombineC5Request(BaseModel):
    job_ids: Optional[List[str]] = None
    outlet_name: Optional[str] = None

class C5PushItemUpdate(BaseModel):
    sid: Optional[str] = ""
    outlet_name: Optional[str] = ""
    item_id: str
    category_id: Optional[str] = ""
    category: Optional[str] = ""
    item_name: Optional[str] = ""
    item_name_new: Optional[str] = ""
    photo_link: Optional[str] = ""
    current_fake_price: Optional[float] = None
    new_fake_price: Optional[float] = None
    current_real_price: Optional[float] = None
    changes: Optional[List[str]] = []

class C5PushRequest(BaseModel):
    platform: str = Field("gofood", description="Target platform (default: gofood)")
    selected_sids: List[str]
    updates: List[C5PushItemUpdate]




# ─── GSHEETS SYNC ENDPOINT ───────────────────────────────────────────────────

import io
import requests
import pandas as pd

@app.post("/api/sync-sheets", status_code=status.HTTP_200_OK)
def sync_sheets(db: Session = Depends(get_db)):
    from menu_core.sheets import GSHEETS_URL, GSHEETS_HEADERS
    try:
        logger.info("⏳ Downloading latest merchant sheet for database sync...")
        resp = requests.get(GSHEETS_URL, headers=GSHEETS_HEADERS, timeout=30)
        resp.raise_for_status()
        df = pd.read_csv(io.StringIO(resp.text))
        try:
            from menu_core.sheets import CACHE_PATH
            df.to_csv(CACHE_PATH, index=False)
        except Exception as ce:
            logger.warning(f"⚠️ Failed to update master_merchants_cache.csv: {ce}")
    except Exception as e:
        logger.error(f"❌ Failed to fetch Google Sheet online: {e}")
        from menu_core.sheets import CACHE_PATH
        if os.path.exists(CACHE_PATH):
            logger.info("⚠️ Using local master_merchants_cache.csv fallback.")
            try:
                df = pd.read_csv(CACHE_PATH)
            except Exception as fe:
                raise HTTPException(status_code=500, detail=f"Failed to load cached merchant sheet: {str(fe)}")
        else:
            raise HTTPException(status_code=500, detail=f"Failed to fetch Google Sheet: {str(e)}")

    added_accounts = 0
    added_outlets = 0
    updated_outlets = 0

    # Forward-fill merged/header columns in Google Sheet dataframe (e.g. Owner, Status, Nama Outlet, Merchant Name)
    # Use block_id (derived from non-null No column) to prevent global ffill leak across empty rows/different outlets
    if "No" in df.columns:
        df["block_id"] = df["No"].notna().cumsum()
        for col in ["Owner", "Status", "Nama Outlet", "Outlet", "Merchant Name", "Cabang", "Nama Resto Final", "Brand"]:
            if col in df.columns:
                df[col] = df.groupby("block_id")[col].ffill()
    else:
        for col in ["Owner", "Status", "Nama Outlet", "Outlet", "Merchant Name", "Cabang", "Nama Resto Final", "Brand"]:
            if col in df.columns:
                df[col] = df[col].ffill()

    # Filter Live and Pending status merchants
    df_live = df[df["Status"].astype(str).str.lower().str.contains("live|pending", na=False)]

    # Pre-fetch all accounts and outlets to prevent N+1 queries in the loop
    all_accounts = db.query(Account).all()
    accounts_by_key = {(a.username, a.platform): a for a in all_accounts}

    all_outlets = db.query(Outlet).all()
    outlets_by_store_id = {o.store_id: o for o in all_outlets if o.store_id}
    outlets_by_fallback = {
        (o.account_id, o.merchant_name, o.nama_outlet, o.cabang): o for o in all_outlets
    }

    for _, row in df_live.iterrows():
        app_val = str(row.get("Aplikasi", "")).strip().lower()
        if app_val == "shopeefood":
            platform = "shopee"
        elif app_val == "grabfood":
            platform = "grab"
        elif app_val == "gofood":
            platform = "gofood"
        else:
            continue

        # Extract Username & Password based on platform logic
        username = None
        password = None

        if platform == "shopee":
            user_col_q = "Nama Pengguna"
            user_col_z = "Nama Pengguna.1"
            pwd_col_s = "Kata Sandi"
            pwd_col_ab = "Kata Sandi.1"

            user_q_val = row.get(user_col_q)
            if pd.notna(user_q_val) and str(user_q_val).strip() not in ("-", "", "nan", "None"):
                username = str(user_q_val).strip()
                pwd_s_val = row.get(pwd_col_s)
                if pd.notna(pwd_s_val) and str(pwd_s_val).strip() not in ("-", "", "nan", "None"):
                    password = str(pwd_s_val).strip()
                else:
                    password = "" # No password used, login directly via OTP
            else:
                user_z_val = row.get(user_col_z)
                if pd.notna(user_z_val) and str(user_z_val).strip() not in ("-", "", "nan", "None"):
                    username = str(user_z_val).strip()
                else:
                    username = "superfoodapp"
                
                pwd_ab_val = row.get(pwd_col_ab)
                if pd.notna(pwd_ab_val) and str(pwd_ab_val).strip() not in ("-", "", "nan", "None"):
                    password = str(pwd_ab_val).strip()
                else:
                    password = "Master@00@"
        elif platform == "grab":
            user_col_sf = "Nama Pengguna.1"
            user_col_mt = "Nama Pengguna"
            pwd_col_sf = "Kata Sandi.1"
            pwd_col_mt = "Kata Sandi"

            user_val = row.get(user_col_sf) if pd.notna(row.get(user_col_sf)) and str(row.get(user_col_sf)).strip() != "-" else row.get(user_col_mt)
            pwd_val = row.get(pwd_col_sf) if pd.notna(row.get(pwd_col_sf)) and str(row.get(pwd_col_sf)).strip() != "-" else row.get(pwd_col_mt)

            if pd.notna(user_val) and str(user_val).strip() != "":
                username = str(user_val).strip()
            if pd.notna(pwd_val) and str(pwd_val).strip() != "":
                password = str(pwd_val).strip()
        elif platform == "gofood":
            # GoFood strictly uses Email Login Go 1 (Kolom Y), Email Login Go 2 (Kolom Z), or Email (Kolom P)
            email_1 = row.get("Email Login Go 1")
            email_2 = row.get("Email Login Go 2")
            email_p = row.get("Email")
            email_o = row.get("Nama Akses Mitra")

            for candidate in [email_1, email_2, email_p, email_o]:
                if pd.notna(candidate) and "@" in str(candidate) and str(candidate).strip() not in ("-", "", "nan", "None"):
                    username = str(candidate).strip()
                    break

            # Never fallback to Shopee staff 'allvbadmin' (Nama Pengguna.1) for GoFood!
            pwd_val = row.get("Kata Sandi") if pd.notna(row.get("Kata Sandi")) else row.get("Kata Sandi.1")
            if pd.notna(pwd_val) and str(pwd_val).strip() not in ("-", "", "nan", "None"):
                password = str(pwd_val).strip()

        if not username:
            continue
        if password is None:
            password = ""

        # 1. Upsert Account
        db_account = accounts_by_key.get((username, platform))
        if not db_account:
            db_account = Account(
                platform=platform,
                username=username,
                password=password,
                portal="shopee_partner" if platform == "shopee" else "merchant_portal"
            )
            db.add(db_account)
            db.flush()  # Generate primary key ID without committing
            accounts_by_key[(username, platform)] = db_account
            added_accounts += 1
        else:
            if db_account.password != password:
                db_account.password = password
                db.flush()

        # 2. Extract Outlet Info
        store_id_raw = row.get("Store ID")
        store_id = str(store_id_raw).strip().split(".")[0] if pd.notna(store_id_raw) and str(store_id_raw).strip() != "-" else None
        
        m_name_raw = row.get("Merchant Name")
        merchant_name = str(m_name_raw).strip() if pd.notna(m_name_raw) and str(m_name_raw).strip() != "-" else str(row.get("Nama Outlet", "")).strip()

        owner_raw = row.get("Owner")
        owner = str(owner_raw).strip() if pd.notna(owner_raw) and str(owner_raw).strip() not in ("-", "") else None

        n_out_raw = row.get("Nama Outlet") if (pd.notna(row.get("Nama Outlet")) and str(row.get("Nama Outlet")).strip() not in ("-", "")) else (row.get("Outlet") if (pd.notna(row.get("Outlet")) and str(row.get("Outlet")).strip() not in ("-", "")) else row.get("Nama Resto Final"))
        cabang = str(row.get("Cabang", "")).strip() if pd.notna(row.get("Cabang")) else str(row.get("Brand", "")).strip()
        nama_resto_final = str(row.get("Nama Resto Final", "")).strip() if pd.notna(row.get("Nama Resto Final")) else None
        brand = str(row.get("Brand", "")).strip() if pd.notna(row.get("Brand")) else None
        nama_outlet = str(n_out_raw).strip() if pd.notna(n_out_raw) and str(n_out_raw).strip() not in ("-", "") else (nama_resto_final or merchant_name or None)

        # 3. Upsert Outlet
        db_outlet = None
        if store_id:
            db_outlet = outlets_by_store_id.get(store_id)

        if not db_outlet:
            # Fallback query if store_id was not provided
            db_outlet = outlets_by_fallback.get((db_account.id, merchant_name, nama_outlet, cabang))

        if not db_outlet:
            db_outlet = Outlet(
                account_id=db_account.id,
                store_id=store_id,
                merchant_name=merchant_name,
                owner=owner,
                nama_outlet=nama_outlet,
                cabang=cabang,
                nama_resto_final=nama_resto_final,
                brand=brand,
                is_active=True
            )
            db.add(db_outlet)
            db.flush()
            if store_id:
                outlets_by_store_id[store_id] = db_outlet
            outlets_by_fallback[(db_account.id, merchant_name, nama_outlet, cabang)] = db_outlet
            added_outlets += 1
        else:
            db_outlet.account_id = db_account.id
            db_outlet.store_id = store_id
            db_outlet.owner = owner
            if merchant_name and merchant_name != "-":
                db_outlet.merchant_name = merchant_name
            db_outlet.nama_outlet = nama_outlet or db_outlet.nama_resto_final or db_outlet.merchant_name
            if cabang:
                db_outlet.cabang = cabang
            db_outlet.nama_resto_final = nama_resto_final
            db_outlet.brand = brand
            db_outlet.is_active = True
            db.flush()
            if store_id and store_id not in outlets_by_store_id:
                outlets_by_store_id[store_id] = db_outlet
            updated_outlets += 1

    db.commit()
    logger.info(f"📊 Sync Sheet Complete. Added Accounts: {added_accounts}, Added Outlets: {added_outlets}, Updated Outlets: {updated_outlets}")
    return {
        "status": "success",
        "added_accounts": added_accounts,
        "added_outlets": added_outlets,
        "updated_outlets": updated_outlets
    }


# ─── ACCOUNTS ENDPOINTS ───────────────────────────────────────────────────────

@app.post("/api/accounts", response_model=AccountResponse, status_code=status.HTTP_201_CREATED)
def create_account(account: AccountCreate, db: Session = Depends(get_db)):
    db_account = db.query(Account).filter(
        Account.username == account.username, 
        Account.platform == account.platform
    ).first()
    if db_account:
        raise HTTPException(status_code=400, detail="Account already exists for this platform")
    
    new_account = Account(
        platform=account.platform,
        username=account.username,
        password=account.password,
        portal=account.portal
    )
    db.add(new_account)
    db.commit()
    db.refresh(new_account)
# ─── REALTIME SHEETS AUTO-SYNC FOR DROPDOWN ────────────────────────────────────
LAST_SYNC_TIMESTAMP = 0.0
SYNC_LOCK = threading.Lock()

def check_and_auto_sync_sheets(db: Session, force: bool = False, min_interval_seconds: float = 3600.0):
    global LAST_SYNC_TIMESTAMP
    now = time.time()
    if force or (LAST_SYNC_TIMESTAMP > 0 and now - LAST_SYNC_TIMESTAMP > min_interval_seconds):
        if SYNC_LOCK.acquire(blocking=False):
            try:
                LAST_SYNC_TIMESTAMP = time.time()
                logger.info("🔄 Auto-sync: Fetching latest merchants from Google Sheets...")
                sync_sheets(db)
            except Exception as e:
                logger.warning(f"⚠️ Auto-sync Google Sheets warning: {e}")
            finally:
                SYNC_LOCK.release()

@app.get("/api/accounts", response_model=List[AccountResponse])
def list_accounts(
    refresh: bool = Query(default=False, description="Paksa sync ulang dari Google Sheets"),
    db: Session = Depends(get_db),
):
    if refresh:
        check_and_auto_sync_sheets(db, force=True)
    return db.query(Account).all()


# ─── OUTLETS ENDPOINTS ────────────────────────────────────────────────────────

SUPPORTED_PLATFORMS = {"shopee", "grab", "gofood"}


def normalize_platform_filters(platforms: Optional[List[str]]) -> List[str]:
    normalized = list(dict.fromkeys(
        value.strip().lower()
        for value in (platforms or [])
        if value and value.strip()
    ))
    unsupported = sorted(set(normalized) - SUPPORTED_PLATFORMS)
    if unsupported:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Unsupported platform filter",
                "unsupported": unsupported,
                "supported": sorted(SUPPORTED_PLATFORMS),
            },
        )
    return normalized

@app.post("/api/outlets", response_model=OutletResponse, status_code=status.HTTP_201_CREATED)
def create_outlet(outlet: OutletCreate, db: Session = Depends(get_db)):
    # Verify account exists
    db_account = db.query(Account).filter(Account.id == outlet.account_id).first()
    if not db_account:
        raise HTTPException(status_code=404, detail="Parent account not found")
    
    if outlet.store_id:
        db_outlet = db.query(Outlet).filter(Outlet.store_id == outlet.store_id).first()
        if db_outlet:
            raise HTTPException(status_code=400, detail="Outlet with this store_id already exists")

    new_outlet = Outlet(
        account_id=outlet.account_id,
        store_id=outlet.store_id,
        merchant_name=outlet.merchant_name,
        nama_outlet=outlet.nama_outlet,
        cabang=outlet.cabang,
        nama_resto_final=outlet.nama_resto_final,
        brand=outlet.brand,
        is_active=outlet.is_active
    )
    db.add(new_outlet)
    db.commit()
    db.refresh(new_outlet)
    return new_outlet

@app.get("/api/outlets", response_model=List[OutletResponse])
def list_outlets(
    platform: Optional[List[str]] = Query(
        default=None,
        description="Filter platform berulang, contoh: ?platform=grab&platform=gofood",
    ),
    refresh: bool = Query(default=False, description="Paksa sync ulang dari Google Sheets"),
    db: Session = Depends(get_db),
):
    if refresh:
        check_and_auto_sync_sheets(db, force=True)

    platforms = normalize_platform_filters(platform)
    query = db.query(Outlet).options(joinedload(Outlet.account))
    if platforms:
        query = query.join(Outlet.account).filter(Account.platform.in_(platforms))
    return query.all()


# ─── BACKGROUND JOBS WORKER ───────────────────────────────────────────────────

# Platform-specific locks to allow parallel execution between different platforms,
# but enforce sequential execution within each platform (especially Shopee).
PLATFORM_LOCKS = {
    "shopee": threading.Lock(),
    "grab": threading.Lock(),
    "gofood": threading.Lock()
}

def run_pull_job(job_id: uuid.UUID, outlet_id: uuid.UUID):
    # Setup job-specific session context
    from menu_core.database import SessionLocal
    db = SessionLocal()
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        db.close()
        return

    platform = (job.platform or "").lower()
    lock = PLATFORM_LOCKS.get(platform)
    if lock:
        logger.info(f"🔒 Job {job_id} ({platform}) waiting for lock...")
        lock.acquire()
        logger.info(f"🔓 Job {job_id} ({platform}) acquired lock. Starting execution.")

    try:
        # Re-fetch job under lock to ensure we have the latest database state
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            return

        job.status = "RUNNING"
        job.started_at = datetime.utcnow()
        job.progress_pct = 10
        job.current_step = "Memuat kredensial dan inisialisasi browser..."
        db.commit()

        outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
        account = db.query(Account).filter(Account.id == outlet.account_id).first()
        
        # Determine paths dynamically
        import re
        raw_outlet = outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown'
        clean_outlet = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
        clean_outlet = re.sub(r'\s+', ' ', clean_outlet).lower()
        
        exports_dir = BASE_DIR / "data" / "exports" / job.platform / clean_outlet
        exports_dir.mkdir(parents=True, exist_ok=True)

        # Trigger Applicator specific script
        if job.platform == "shopee":
            job.progress_pct = 30
            job.current_step = "Membuka browser dan login portal Shopee Partner..."
            db.commit()
            
            # Add project root to sys.path to resolve shopee.* absolute imports correctly
            if str(BASE_DIR) not in sys.path:
                sys.path.insert(0, str(BASE_DIR))
                
            from shopee.core.pull import extract_shopee_menu, get_shopee_master_credentials
            master_user, master_pass = get_shopee_master_credentials()

            # Setup store_metadata payload for shopee.core.pull
            store_metadata = {
                "store_id": outlet.store_id,
                "merchant_name": outlet.merchant_name,
                "nama_outlet": outlet.nama_outlet,
                "cabang": outlet.cabang,
                "nama_resto_final": outlet.nama_resto_final,
                "brand": outlet.brand,
                "username": master_user,
                "password": master_pass,
                "portal": account.portal
            }
            
            # Run shopee extraction (headless=True for stability on headless servers/Raspberry Pi)
            is_headless = True
            success, result = extract_shopee_menu(store_metadata, str(exports_dir), headless=is_headless)
            
            if not success:
                raise Exception(f"Shopee extraction failed: {result}")
                
            # If store_id was dynamically resolved and wasn't set in DB, update it!
            resolved_store_id = store_metadata.get("store_id")
            if resolved_store_id and not outlet.store_id:
                # Check for uniqueness before updating to prevent constraints failure
                existing_outlet = db.query(Outlet).filter(Outlet.store_id == resolved_store_id).first()
                if not existing_outlet:
                    outlet.store_id = resolved_store_id
                    logger.info(f"💾 Dynamically updated store_id to {resolved_store_id} for outlet {outlet.merchant_name}")
            
            job.status = "SUCCESS"
            job.progress_pct = 100
            job.current_step = "Penarikan menu selesai!"
            job.result_metadata = {
                "excel_path": result.get("excel"),
                "items_csv_path": result.get("items_csv"),
                "mods_csv_path": result.get("mods_csv"),
                "items_count": result.get("items_count", 0),
                "mods_count": result.get("mods_count", 0),
                "completed_at": datetime.utcnow().isoformat()
            }
            job.completed_at = datetime.utcnow()
            outlet.last_sync_at = datetime.utcnow()
            db.commit()
            
        elif job.platform == "gofood":
            job.progress_pct = 30
            job.current_step = "Menyiapkan parameter penarikan GoFood..."
            db.commit()
            
            store_metadata = {
                "store_id": outlet.store_id,
                "merchant_name": outlet.merchant_name,
                "nama_outlet": outlet.nama_outlet,
                "cabang": outlet.cabang,
                "nama_resto_final": outlet.nama_resto_final,
                "brand": outlet.brand,
                "username": account.username,
                "password": account.password
            }
            
            job.progress_pct = 50
            job.current_step = "Meluncurkan browser GoFood & memproses penarikan..."
            db.commit()
            
            from menu_core.gofood import extract_gofood_menu
            success, result = extract_gofood_menu(store_metadata, str(exports_dir))
            
            if not success:
                raise Exception(f"GoFood extraction failed: {result}")
                
            job.status = "SUCCESS"
            job.progress_pct = 100
            job.current_step = "Penarikan menu GoFood selesai!"
            job.result_metadata = {
                "excel_path": result.get("excel"),
                "items_count": result.get("items_count", 0),
                "mods_count": result.get("mods_count", 0),
                "completed_at": datetime.utcnow().isoformat()
            }
            job.completed_at = datetime.utcnow()
            outlet.last_sync_at = datetime.utcnow()
            db.commit()
            
        elif job.platform == "grab":
            job.progress_pct = 30
            job.current_step = "Menyiapkan parameter penarikan GrabFood..."
            db.commit()
            
            store_metadata = {
                "store_id": outlet.store_id,
                "merchant_name": outlet.merchant_name,
                "nama_outlet": outlet.nama_outlet,
                "cabang": outlet.cabang,
                "nama_resto_final": outlet.nama_resto_final,
                "brand": outlet.brand,
                "username": account.username,
                "password": account.password
            }
            
            job.progress_pct = 50
            job.current_step = "Meluncurkan browser GrabFood & memproses penarikan..."
            db.commit()
            
            from menu_core.grab import extract_grab_menu
            success, result = extract_grab_menu(store_metadata, str(exports_dir))
            
            if not success:
                raise Exception(f"Grab extraction failed: {result}")
                
            job.status = "SUCCESS"
            job.progress_pct = 100
            job.current_step = "Penarikan menu GrabFood selesai!"
            job.result_metadata = {
                "excel_path": result.get("excel"),
                "items_count": result.get("items_count", 0),
                "mods_count": result.get("mods_count", 0),
                "completed_at": datetime.utcnow().isoformat()
            }
            job.completed_at = datetime.utcnow()
            outlet.last_sync_at = datetime.utcnow()
            db.commit()

        logger.info(f"✅ Job {job_id} completed successfully.")

    except Exception as e:
        logger.error(f"❌ Job {job_id} failed: {e}")
        job.status = "FAILED"
        if "user membatalkan otp" in str(e).lower():
            job.error_message = "user membatalkan otp"
            job.current_step = "Gagal: user membatalkan otp"
        else:
            job.error_message = str(e)
            err_msg = f"Terjadi kesalahan: {str(e)}"
            job.current_step = err_msg if len(err_msg) <= 255 else err_msg[:252] + "..."
        job.completed_at = datetime.utcnow()
        db.commit()
    finally:
        if lock:
            try:
                lock.release()
                logger.info(f"🔓 Job {job_id} ({platform}) lock released.")
            except Exception as le:
                logger.warning(f"⚠️ Failed to release lock: {le}")
        db.close()


def run_push_price_job(job_id: uuid.UUID, outlet_id: uuid.UUID, updates_list: list):
    """Background task to push price changes to GoFood, GrabFood, or ShopeeFood."""
    import asyncio
    from menu_core.database import SessionLocal
    db = SessionLocal()
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        db.close()
        return

    platform = (job.platform or "").lower()
    lock = PLATFORM_LOCKS.get(platform)
    if lock:
        logger.info(f"🔒 Job {job_id} ({platform}) waiting for lock...")
        lock.acquire()
        logger.info(f"🔓 Job {job_id} ({platform}) acquired lock. Starting execution.")

    try:
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            return

        platform = (job.platform or "").lower()
        logger.info(f"🚀 run_push_price_job starting for job {job_id}, platform {platform}. updates_list: {updates_list}")

        job.status = "RUNNING"
        job.started_at = datetime.utcnow()
        job.progress_pct = 10
        job.current_step = "Menginisialisasi kredensial..."
        db.commit()

        outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
        account = db.query(Account).filter(Account.id == outlet.account_id).first()

        total_updates = len(updates_list)
        success_count = 0
        fail_count = 0

        logger.info(f"📋 Total updates to process: {total_updates}")

        if platform == "shopee":
            # Add project root to sys.path
            if str(BASE_DIR) not in sys.path:
                sys.path.insert(0, str(BASE_DIR))
            from shopee.core.push import push_price_update_batch

            store_metadata = {
                "store_id": outlet.store_id,
                "username": account.username,
                "password": account.password,
                "merchant_name": outlet.merchant_name,
                "nama_resto_final": outlet.nama_resto_final,
                "nama_outlet": outlet.nama_outlet
            }

            job.progress_pct = 30
            job.current_step = "Membuka browser Shopee & mengautentikasi sesi..."
            db.commit()

            def on_progress(idx, total, name, price):
                job.current_step = f"Memproses update harga '{name}' ({idx + 1}/{total}) ke Rp {price:,.0f}..."
                job.progress_pct = int(30 + ((idx + 1) / total) * 60)
                db.commit()

            # Headless=True for stability on headless servers/Raspberry Pi
            is_headless = True
            results = push_price_update_batch(store_metadata, updates_list, headless=is_headless, on_item_progress=on_progress)

            for res in results:
                item_id = res["item_id"]
                item_name = res["item_name"]
                new_price = res["new_price"]
                
                if res["success"]:
                    success_count += 1
                    status_str = "SUCCESS"
                    err_msg = None
                elif res.get("status") == "SKIPPED_ACTIVE_PROMO":
                    status_str = "SKIPPED_ACTIVE_PROMO"
                    err_msg = res.get("error_message") or "Item sedang dalam promo/slash price aktif di ShopeeFood."
                else:
                    fail_count += 1
                    status_str = "FAILED"
                    err_msg = res["error_message"]

                trail = AuditTrail(
                    job_id=job.id,
                    outlet_id=outlet.id,
                    item_id=item_id,
                    item_name=item_name,
                    change_type="PRICE_UPDATE",
                    field_changed="price",
                    old_value=None,
                    new_value=str(new_price),
                    status=status_str,
                    error_message=err_msg
                )
                db.add(trail)
                db.commit()

        elif platform == "grab":
            from playwright.async_api import async_playwright
            from grab.core.grab_api_scraper import GrabAPI, perform_login, SESSION_DIR

            username = account.username
            password = account.password
            store_id = outlet.store_id

            job.progress_pct = 20
            job.current_step = "Meluncurkan browser Grab & login portal..."
            db.commit()

            async def grab_async_flow():
                nonlocal success_count, fail_count
                async with async_playwright() as p:
                    session_path = os.path.join(SESSION_DIR, f"{username}.json")
                    storage_state = session_path if os.path.exists(session_path) else None
                    
                    use_system_chromium = os.path.exists("/usr/lib/chromium/chromium") or os.path.exists("/usr/bin/chromium")
                    if use_system_chromium:
                        import socket, subprocess
                        def get_free_port():
                            s = socket.socket()
                            s.bind(('', 0))
                            port = s.getsockname()[1]
                            s.close()
                            return port
                        cdp_port = get_free_port()
                        chromium_bin = "/usr/lib/chromium/chromium" if os.path.exists("/usr/lib/chromium/chromium") else "/usr/bin/chromium"
                        proc = subprocess.Popen([
                            chromium_bin, "--no-sandbox", "--no-zygote", "--in-process-gpu", "--disable-gpu",
                            "--disable-dev-shm-usage", "--disable-gpu-sandbox", "--dbus-stub", f"--remote-debugging-port={cdp_port}", "--headless=new"
                        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                        await asyncio.sleep(3.0)
                        browser = await p.chromium.connect_over_cdp(f"http://127.0.0.1:{cdp_port}")
                    else:
                        browser = await p.chromium.launch(headless=True)
                    context = await browser.new_context(
                        storage_state=storage_state,
                        user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
                    )
                    page = await context.new_page()

                    try:
                        await page.goto("https://merchant.grab.com/dashboard", wait_until="domcontentloaded", timeout=30000)
                    except Exception as e:
                        logger.warning(f"Grab dashboard navigate warning: {e}")

                    api = GrabAPI(page, username, password)
                    mgid = await api.get_merchant_group_id()
                    if not mgid:
                        logger.info("Session state invalid or expired, running perform_login...")
                        if await perform_login(page, username, password):
                            mgid = await api.get_merchant_group_id()
                            if mgid:
                                await context.storage_state(path=session_path)
                            else:
                                await browser.close()
                                return "Failed to retrieve Grab merchant group ID after login."
                        else:
                            await browser.close()
                            return "Grab login failed."

                    # Navigasi ke store-specific menu page untuk mengaktifkan konteks sesi store di Grab
                    try:
                        menu_tab_url = f"https://merchant.grab.com/food/menu/{store_id}"
                        logger.info(f"Navigating to Grab store menu page: {menu_tab_url}")
                        await page.goto(menu_tab_url, wait_until="domcontentloaded", timeout=30000)
                        await page.wait_for_timeout(3000)
                    except Exception as e:
                        logger.warning(f"Grab store menu page navigate warning: {e}")

                    menu_data, err = await api.fetch_menu(mgid, store_id, is_menu_group=False)
                    if err or not menu_data:
                        await browser.close()
                        return f"Failed to fetch Grab menu: {err}"

                    # Detect if menu_data is from a Menu Group V2
                    is_menu_group_detected = bool(menu_data.get("is_menu_group"))
                    detected_menu_group_id = menu_data.get("menuGroupID")
                    if not is_menu_group_detected:
                        for cat in menu_data.get("categories", []):
                            if cat.get("menuGroupID"):
                                is_menu_group_detected = True
                                detected_menu_group_id = cat.get("menuGroupID")
                                break
                            for item in cat.get("items") or []:
                                if item.get("menuGroupID"):
                                    is_menu_group_detected = True
                                    detected_menu_group_id = item.get("menuGroupID")
                                    break
                            if is_menu_group_detected:
                                break

                    grab_items_by_id = {}
                    grab_items_by_name = {}
                    for cat in menu_data.get("categories", []):
                        items_list = cat.get("items") or cat.get("menuItems") or []
                        selling_time_id = cat.get("sellingTimeID")
                        for item in items_list:
                            info = {
                                "item": item,
                                "category_id": cat.get("categoryID"),
                                "sellingTimeID": selling_time_id
                            }
                            grab_items_by_id[str(item.get("itemID"))] = info
                            item_name_key = (item.get("itemName") or "").strip().lower()
                            if item_name_key:
                                grab_items_by_name[item_name_key] = info

                    thread_db = SessionLocal()
                    try:
                        items_breakdown = []
                        # FASE 1: Push Perubahan Harga
                        for idx, update in enumerate(updates_list):
                            item_id = str(update["item_id"])
                            new_price = float(update["new_price"])
                            item_name_req = (update.get("item_name") or "").strip().lower()

                            item_info = grab_items_by_id.get(item_id)
                            if not item_info and item_name_req:
                                item_info = grab_items_by_name.get(item_name_req)

                            if not item_info:
                                fail_count += 1
                                item_label = update.get("item_name") or item_id
                                trail = AuditTrail(
                                    job_id=job.id,
                                    outlet_id=outlet.id,
                                    item_id=item_id,
                                    item_name=item_label,
                                    change_type="PRICE_UPDATE",
                                    field_changed="price",
                                    old_value=None,
                                    new_value=str(new_price),
                                    status="FAILED",
                                    error_message=f"Item '{item_label}' (ID: {item_id}) tidak ditemukan di menu Grab aktif."
                                )
                                thread_db.add(trail)
                                thread_db.commit()
                                items_breakdown.append({
                                    "item_id": item_id,
                                    "item_name": item_label,
                                    "old_price": None,
                                    "requested_price": new_price,
                                    "verified_price": None,
                                    "status": "FAILED",
                                    "error_message": f"Item '{item_label}' tidak ditemukan di menu Grab aktif."
                                })
                                continue

                            orig_item = item_info["item"]
                            real_item_id = str(orig_item.get("itemID"))
                            category_id = item_info["category_id"]
                            selling_time_id = item_info["sellingTimeID"]
                            old_p = float(orig_item.get("priceInMin", 0)) / 100.0

                            # 1. Cek Active Promo / Slash Price Grab (itemCampaignInfo)
                            campaign = orig_item.get("itemCampaignInfo")
                            if campaign and isinstance(campaign, dict):
                                item_label = orig_item.get("itemName", real_item_id)
                                err_promo = f"Item '{item_label}' sedang dalam promo/slash price aktif di GrabFood. Perubahan harga dasar dikunci."
                                logger.warning(f"🔒 Promo Grab Terdeteksi: {err_promo}")
                                trail = AuditTrail(
                                    job_id=job.id,
                                    outlet_id=outlet.id,
                                    item_id=real_item_id,
                                    item_name=item_label,
                                    change_type="PRICE_UPDATE",
                                    field_changed="price",
                                    old_value=str(old_p),
                                    new_value=str(new_price),
                                    status="SKIPPED_ACTIVE_PROMO",
                                    error_message=err_promo
                                )
                                thread_db.add(trail)
                                thread_db.commit()
                                items_breakdown.append({
                                    "item_id": real_item_id,
                                    "item_name": item_label,
                                    "old_price": old_p,
                                    "requested_price": new_price,
                                    "verified_price": old_p,
                                    "status": "SKIPPED_ACTIVE_PROMO",
                                    "error_message": err_promo
                                })
                                continue

                            # 2. Cek Kuota Bulanan Grab (Maksimal 15x Perubahan Harga per Item per 30 Hari)
                            thirty_days_ago = datetime.utcnow() - timedelta(days=30)
                            monthly_count = thread_db.query(AuditTrail).filter(
                                AuditTrail.outlet_id == outlet.id,
                                AuditTrail.item_id == real_item_id,
                                AuditTrail.field_changed.ilike("price"),
                                AuditTrail.status.ilike("SUCCESS"),
                                AuditTrail.created_at >= thirty_days_ago
                            ).count()

                            if monthly_count >= 15:
                                fail_count += 1
                                item_label = orig_item.get("itemName", real_item_id)
                                err_quota = f"Item '{item_label}' telah mencapai batas maksimal 15x perubahan harga per bulan di Grab ({monthly_count}/15 terpakai)."
                                logger.warning(f"⚠️ Kuota Grab Terlampaui: {err_quota}")
                                trail = AuditTrail(
                                    job_id=job.id,
                                    outlet_id=outlet.id,
                                    item_id=real_item_id,
                                    item_name=item_label,
                                    change_type="PRICE_UPDATE",
                                    field_changed="price",
                                    old_value=str(old_p),
                                    new_value=str(new_price),
                                    status="FAILED",
                                    error_message=err_quota
                                )
                                thread_db.add(trail)
                                thread_db.commit()
                                items_breakdown.append({
                                    "item_id": real_item_id,
                                    "item_name": item_label,
                                    "old_price": old_p,
                                    "requested_price": new_price,
                                    "verified_price": None,
                                    "status": "FAILED",
                                    "error_message": err_quota
                                })
                                continue

                            # 3. Hitung Tahapan Kenaikan / Penurunan (>15% Push Bertahap)
                            steps = calculate_price_steps(old_p, new_price, max_step_pct=0.15)
                            if len(steps) > 1:
                                logger.info(f"📊 Grab Push Bertahap (>15%) untuk {orig_item.get('itemName')}: Rp {old_p:,.0f} -> Rp {new_price:,.0f} via {len(steps)} tahapan: {steps}")

                            item_data = dict(orig_item)
                            if category_id and "categoryID" not in item_data:
                                item_data["categoryID"] = category_id
                            if selling_time_id and "sellingTimeID" not in item_data:
                                item_data["sellingTimeID"] = selling_time_id

                            status_str = "SUCCESS"
                            err_msg = None

                            for step_idx, step_p in enumerate(steps):
                                item_data["priceInMin"] = int(step_p * 100)

                                val_ok, val_err = await api.validate_item(
                                    mgid, store_id, category_id, item_data,
                                    is_menu_group=is_menu_group_detected,
                                    menu_group_id=detected_menu_group_id
                                )
                                if val_err:
                                    logger.warning(f"Grab validation warning for item {real_item_id}: {val_err}")

                                upsert_res, upsert_err = await api.upsert_item(
                                    mgid, store_id, category_id, item_data,
                                    is_menu_group=is_menu_group_detected,
                                    menu_group_id=detected_menu_group_id
                                )

                                if not (upsert_res and not upsert_err):
                                    status_str = "FAILED"
                                    err_msg = upsert_err or "Unknown Grab API error."
                                    logger.error(f"❌ Grab PUSH tahap harga Rp {step_p:,.0f} gagal: {err_msg}")
                                    break

                                if len(steps) > 1:
                                    logger.info(f"   [Grab Tahap {step_idx+1}/{len(steps)}] Berhasil PUSH harga intermediate: Rp {step_p:,.0f}")
                                    if step_idx < len(steps) - 1:
                                        await asyncio.sleep(1.5)

                            if status_str == "SUCCESS":
                                success_count += 1
                            else:
                                fail_count += 1

                            trail = AuditTrail(
                                job_id=job.id,
                                outlet_id=outlet.id,
                                item_id=real_item_id,
                                item_name=orig_item.get("itemName", real_item_id),
                                change_type="PRICE_UPDATE",
                                field_changed="price",
                                old_value=str(old_p),
                                new_value=str(new_price),
                                status=status_str,
                                error_message=err_msg
                            )
                            thread_db.add(trail)
                            thread_db.commit()

                            items_breakdown.append({
                                "item_id": real_item_id,
                                "item_name": orig_item.get("itemName", real_item_id),
                                "old_price": old_p,
                                "requested_price": new_price,
                                "verified_price": new_price if status_str == "SUCCESS" else None,
                                "status": status_str,
                                "error_message": err_msg
                            })

                        # FASE 2: Memverifikasi perubahan harga dengan penarikan menu real-time
                        if success_count > 0:
                            logger.info("Executing Phase 2: Live re-fetch menu verification from Grab...")
                            await page.wait_for_timeout(3000)
                            re_menu, _ = await api.fetch_menu(mgid, store_id, is_menu_group=False)
                            if re_menu:
                                re_items = {}
                                for cat in re_menu.get("categories", []):
                                    for item in (cat.get("items") or cat.get("menuItems") or []):
                                        re_items[str(item.get("itemID"))] = item
                                        n_key = (item.get("itemName") or "").strip().lower()
                                        if n_key:
                                            re_items[n_key] = item

                                for b_item in items_breakdown:
                                    if b_item["status"] == "SUCCESS":
                                        bid = b_item["item_id"]
                                        bname = (b_item["item_name"] or "").strip().lower()
                                        live_item = re_items.get(bid) or re_items.get(bname)
                                        if live_item:
                                            live_p = float(live_item.get("priceInMin", 0)) / 100.0
                                            b_item["verified_price"] = live_p
                                            if abs(live_p - b_item["requested_price"]) > 0.01:
                                                b_item["status"] = "UNVERIFIED"
                                                b_item["error_message"] = f"Verification mismatch: requested Rp {b_item['requested_price']}, live Rp {live_p}"

                        # Store items_breakdown in job.result_metadata
                        job_record = thread_db.query(Job).filter(Job.id == job.id).first()
                        if job_record:
                            job_record.result_metadata = {
                                "total_updates": total_updates,
                                "success_count": success_count,
                                "fail_count": fail_count,
                                "items_breakdown": items_breakdown,
                                "completed_at": datetime.utcnow().isoformat()
                            }
                            thread_db.commit()
                    finally:
                        thread_db.close()

                    await browser.close()
                    return None

            import threading
            err_msg = None
            def run_in_thread():
                nonlocal err_msg
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    err_msg = loop.run_until_complete(grab_async_flow())
                except Exception as e:
                    import traceback
                    err_msg = str(e)
                    traceback.print_exc()
                finally:
                    try:
                        loop.close()
                    except:
                        pass

            t = threading.Thread(target=run_in_thread)
            t.start()
            t.join()

            if err_msg:
                raise Exception(err_msg)

        elif platform == "gofood":
            from playwright.sync_api import sync_playwright
            from Gofood.GO.actions import _menu_api as go_api
            from Gofood.GO.updater_gofood import SESSION_DIR as GO_SESSION_DIR

            email = account.username
            password = account.password
            merchant_id = outlet.store_id

            if not merchant_id:
                raise Exception("Merchant ID (store_id) is missing for GoFood outlet.")
            merchant_id = str(merchant_id).strip()
            if merchant_id.startswith("GM"):
                merchant_id = merchant_id[1:]
            elif merchant_id.isdigit():
                merchant_id = "G" + merchant_id

            job.progress_pct = 30
            job.current_step = "Meluncurkan browser GoFood..."
            db.commit()

            with sync_playwright() as p:
                import re
                from login_gofood import load_gofood_session
                sanitized_email = re.sub(r'[^a-zA-Z0-9_.-]', '_', email.strip().lower())
                session_path = os.path.join(BASE_DIR, "Gofood", f"session_gofood_{sanitized_email}.json")
                cached_data = load_gofood_session(email)
                if not cached_data and account.portal:
                    cached_data = load_gofood_session(account.portal)

                def _get_token_from_session_dict(sdata):
                    if not sdata: return None
                    t = sdata.get("access_token") or sdata.get("token") or sdata.get("authorization")
                    if t and len(t) > 20:
                        return t if t.startswith("Bearer ") else f"Bearer {t}"
                    for c in sdata.get("cookies", []):
                        if c.get("name") in ('access_token', 'token', 'gobiz_token', 'authorization'):
                            val = c.get("value", "")
                            if val and len(val) > 20:
                                return val if val.startswith("Bearer ") else f"Bearer {val}"
                    return None

                token = _get_token_from_session_dict(cached_data)
                if not token and session_path and os.path.exists(session_path):
                    try:
                        with open(session_path, "r", encoding="utf-8") as sf:
                            token = _get_token_from_session_dict(json.load(sf))
                    except Exception: pass

                def _find_gofood_cache_file(m_id):
                    cands = [m_id, m_id.replace("GM", "M"), m_id.lstrip("G"), m_id.strip()]
                    for cid in cands:
                        cp = os.path.join(BASE_DIR, "Gofood", "API", f"menu-response-{cid}.json")
                        if os.path.exists(cp):
                            return cp
                    return None

                rest_uuid = None
                cache_path = _find_gofood_cache_file(merchant_id)
                if cache_path and os.path.exists(cache_path):
                    try:
                        with open(cache_path, "r", encoding="utf-8") as f:
                            raw_cdata = f.read()
                            match = re.search(r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}', raw_cdata, re.I)
                            if match:
                                rest_uuid = match.group(0)
                    except Exception: pass

                headless_env = os.getenv("HEADLESS") or os.getenv("HEADLESS_GOFOOD")
                is_headless = headless_env.lower() in ("true", "1", "yes") if headless_env else True
                from src.core.browser_factory import launch_universal_playwright_browser
                browser, proc = launch_universal_playwright_browser(p, headless=is_headless)
                context = browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                )
                if cached_data and cached_data.get("cookies"):
                    try:
                        context.add_cookies(cached_data["cookies"])
                    except Exception as e:
                        logger.warning(f"Error adding cookies: {e}")

                page = context.new_page()
                api_headers = {}
                if token:
                    api_headers['authorization'] = token

                def capture_headers(request):
                    url_lower = request.url.lower()
                    if "api.gojekapi.com" in url_lower or "api.gobiz.co.id" in url_lower or "portal.gofoodmerchant.co.id" in url_lower:
                        h = request.headers
                        if 'authorization' in h:
                            api_headers['authorization'] = h['authorization']
                        if 'x-passkey' in h:
                            api_headers['x-passkey'] = h['x-passkey']
                    if "restaurants/" in url_lower:
                        parts = request.url.split("/")
                        for i, part in enumerate(parts):
                            if part.lower() == "restaurants" and i + 1 < len(parts):
                                candidate = parts[i + 1].split("?")[0]
                                if len(candidate) == 36 and "-" in candidate:
                                    api_headers['restaurant_uuid'] = candidate
                    if "menu_groups/" in url_lower:
                        parts = request.url.split("/")
                        for i, part in enumerate(parts):
                            if part.lower() == "menu_groups" and i + 1 < len(parts):
                                candidate = parts[i + 1].split("?")[0]
                                if len(candidate) == 36 and "-" in candidate:
                                    api_headers['menu_group_id'] = candidate

                page.on("request", capture_headers)

                def perform_fresh_login():
                    logger.info(f"🔄 Token GoFood expired/tidak ditemukan. Menjalankan login_outlet untuk {email} (store: {merchant_id})...")
                    try:
                        import threading, asyncio
                        res_box = [None]
                        err_box = [None]
                        
                        outlet_meta = {
                            'store_id': merchant_id,
                            'email': email,
                            'emails': [email] if email else [],
                        }

                        def _worker():
                            try:
                                asyncio.set_event_loop(asyncio.new_event_loop())
                                from login_gofood import login_outlet
                                res_box[0] = login_outlet(outlet_meta)
                            except Exception as ex:
                                err_box[0] = ex

                        t = threading.Thread(target=_worker)
                        t.start()
                        t.join()

                        if err_box[0]:
                            raise err_box[0]

                        res = res_box[0]
                        if res and res.get('access_token'):
                            tok = res['access_token']
                            api_headers['authorization'] = tok if tok.startswith("Bearer ") else f"Bearer {tok}"
                            if res.get('restaurant_uuid'):
                                api_headers['restaurant_uuid'] = res['restaurant_uuid']
                            if res.get('cookies'):
                                try: context.add_cookies(res['cookies'])
                                except Exception: pass
                            return True
                    except Exception as e:
                        logger.warning(f"Terjadi kesalahan saat login_outlet: {e}")
                    return False

                if "/auth" in page.url or "login" in page.url:
                    if not api_headers.get('authorization'):
                        perform_fresh_login()
                    page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}/menu-items", wait_until="domcontentloaded")
                    time.sleep(2)

                page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}/menu-items", wait_until="domcontentloaded")
                time.sleep(2)
                page.reload(wait_until="domcontentloaded")
                time.sleep(2)

                if "/auth" in page.url or "login" in page.url:
                    perform_fresh_login()
                    page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}/menu-items", wait_until="domcontentloaded")
                    time.sleep(3)

                def tutup_semua_popup(p):
                    cookie_sels = ['button:has-text("Terima Semua Cookie")', 'button:has-text("Accept All Cookies")', 'button:has-text("Terima")', 'button:has-text("Accept")']
                    for sel in cookie_sels:
                        try:
                            loc = p.locator(sel)
                            if loc.count() > 0 and loc.first.is_visible():
                                loc.first.click(timeout=1500)
                                time.sleep(0.5)
                        except Exception: pass

                    dismiss_sels = ['button:has-text("Lewati")', 'button:has-text("Lewati Tutorial")', 'button:has-text("Selesai")', 'button:has-text("Tutup")', 'button:has-text("Nanti Saja")', '[aria-label="close"]', '[aria-label="Close"]', 'button.close', '.dismiss-button', 'button[class*="close"]', 'button:has-text("×")', 'button:has-text("✕")']
                    for sel in dismiss_sels:
                        try:
                            loc = p.locator(sel)
                            for i in range(loc.count()):
                                cand = loc.nth(i)
                                if cand.is_visible():
                                    cand.click(timeout=1500)
                                    time.sleep(0.5)
                        except Exception: pass

                for _ in range(2):
                    tutup_semua_popup(page)
                    time.sleep(1)

                # Wait dynamically (up to 15s) for restaurant_uuid, authorization, and x-passkey
                start_wait = time.time()
                while (time.time() - start_wait) < 15:
                    if api_headers.get('restaurant_uuid') and api_headers.get('authorization') and api_headers.get('x-passkey'):
                        break
                    page.wait_for_timeout(500)

                token = api_headers.get('authorization')
                if not token:
                    cookies = context.cookies()
                    for c in cookies:
                        if c['name'] in ('access_token', 'token', 'gobiz_token'):
                            token = f"Bearer {c['value']}"
                            break

                if not token:
                    try:
                        token_eval = page.evaluate("""() => {
                            const keys = ['token', 'access_token', 'accessToken', 'auth_token', 'authorization', 'gobiz-token', 'go-id-token'];
                            for (const k of keys) {
                                let val = localStorage.getItem(k) || sessionStorage.getItem(k);
                                if (val) {
                                    if (val.startsWith('{')) {
                                        try {
                                            const parsed = JSON.parse(val);
                                            val = parsed.token || parsed.access_token || parsed.accessToken || val;
                                        } catch(e){}
                                    }
                                    if (val && val.length > 20) return val;
                                }
                            }
                            const tokenRegex = /[A-Za-z0-9-_=]+\\.[A-Za-z0-9-_=]+\\.?[A-Za-z0-9-_.+/=]*/;
                            for (let i = 0; i < localStorage.length; i++) {
                                const val = localStorage.getItem(localStorage.key(i));
                                if (val && val.length > 20) {
                                    if (val.includes('eyJ')) return val;
                                    const match = val.match(tokenRegex);
                                    if (match) return match[0];
                                }
                            }
                            for (let i = 0; i < sessionStorage.length; i++) {
                                const val = sessionStorage.getItem(sessionStorage.key(i));
                                if (val && val.length > 20) {
                                    if (val.includes('eyJ')) return val;
                                    const match = val.match(tokenRegex);
                                    if (match) return match[0];
                                }
                            }
                            return null;
                        }""")
                        if token_eval:
                            token = token_eval if token_eval.startswith("Bearer ") else f"Bearer {token_eval}"
                    except Exception as e:
                        logger.warning(f"Gagal mengekstrak token dari web storage: {e}")

                def _get_token_from_session_dict(sdata):
                    t = sdata.get("access_token") or sdata.get("token") or sdata.get("authorization")
                    if t and len(t) > 20:
                        return t if t.startswith("Bearer ") else f"Bearer {t}"
                    for c in sdata.get("cookies", []):
                        if c.get("name") in ('access_token', 'token', 'gobiz_token', 'authorization'):
                            val = c.get("value", "")
                            if val and len(val) > 20:
                                return val if val.startswith("Bearer ") else f"Bearer {val}"
                    return None

                if not token and session_path and os.path.exists(session_path):
                    try:
                        with open(session_path, "r", encoding="utf-8") as sf:
                            sdata = json.load(sf)
                            extracted = _get_token_from_session_dict(sdata)
                            if extracted:
                                token = extracted
                                logger.info(f"🔑 Berhasil mengekstrak token dari session file {os.path.basename(session_path)}")
                    except Exception as se:
                        logger.warning(f"Gagal mengekstrak token dari session_path file: {se}")

                if not token:
                    gofood_dir = os.path.join(BASE_DIR, "Gofood")
                    if os.path.exists(gofood_dir):
                        for fname in os.listdir(gofood_dir):
                            if fname.startswith("session_gofood_") and fname.endswith(".json"):
                                sp = os.path.join(gofood_dir, fname)
                                try:
                                    with open(sp, "r", encoding="utf-8") as sf:
                                        sdata = json.load(sf)
                                        extracted = _get_token_from_session_dict(sdata)
                                        if extracted:
                                            token = extracted
                                            logger.info(f"🔑 Berhasil mengekstrak token dari {fname}")
                                            break
                                except Exception: pass

                rest_uuid = api_headers.get('restaurant_uuid')
                if not rest_uuid or len(rest_uuid) != 36:
                    try:
                        uuid_eval = page.evaluate("""() => {
                            try {
                                if (window.__NEXT_DATA__ && window.__NEXT_DATA__.props) {
                                    const strData = JSON.stringify(window.__NEXT_DATA__.props);
                                    const match = /[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}/i.exec(strData);
                                    if (match) return match[0];
                                }
                            } catch(e){}
                            const uuidRegex = /[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}/i;
                            for (let i = 0; i < localStorage.length; i++) {
                                const val = localStorage.getItem(localStorage.key(i));
                                const match = uuidRegex.exec(val);
                                if (match) return match[0];
                            }
                            for (let i = 0; i < sessionStorage.length; i++) {
                                const val = sessionStorage.getItem(sessionStorage.key(i));
                                const match = uuidRegex.exec(val);
                                if (match) return match[0];
                            }
                            const urlMatch = uuidRegex.exec(window.location.href);
                            if (urlMatch) return urlMatch[0];
                            return null;
                        }""")
                        if uuid_eval:
                            rest_uuid = uuid_eval
                    except Exception as e:
                        logger.warning(f"Gagal mengekstrak rest_uuid dari web storage: {e}")

                def _find_gofood_cache_file(m_id):
                    cands = [m_id, m_id.replace("GM", "M"), m_id.lstrip("G"), m_id.strip()]
                    for cid in cands:
                        cp = os.path.join(BASE_DIR, "Gofood", "API", f"menu-response-{cid}.json")
                        if os.path.exists(cp):
                            return cp
                    return None

                if not rest_uuid or len(rest_uuid) != 36:
                    # Coba baca dari cached menu response hasil Pull sebelumnya
                    cache_path = _find_gofood_cache_file(merchant_id)
                    if cache_path and os.path.exists(cache_path):
                        try:
                            with open(cache_path, "r", encoding="utf-8") as f:
                                raw_cdata = f.read()
                                match = re.search(r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}', raw_cdata, re.I)
                                if match:
                                    rest_uuid = match.group(0)
                        except Exception as e:
                            logger.error(f"Gagal membaca cached restaurant_id: {e}")

                if not rest_uuid or len(rest_uuid) != 36:
                    rest_uuid = None

                menu_data = None
                if token and rest_uuid:
                    try:
                        menu_data = go_api.fetch_menus(page, token, rest_uuid)
                    except Exception as fe:
                        logger.warning(f"⚠️ Fetch menus via API 36-char UUID gagal ({fe}).")

                # Jika token missing atau menu_data 401 / None, picu fresh login
                if not token or not menu_data:
                    logger.warning("⚠️ GoFood session expired / 401 / 422. Memicu fresh login & reload...")
                    perform_fresh_login()
                    page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}/menu-items", wait_until="domcontentloaded")
                    time.sleep(3)

                    start_wait = time.time()
                    while (time.time() - start_wait) < 15:
                        if api_headers.get('authorization'):
                            break
                        page.wait_for_timeout(500)

                    token = api_headers.get('authorization')
                    if not token:
                        cookies = context.cookies()
                        for c in cookies:
                            if c['name'] == 'access_token':
                                token = f"Bearer {c['value']}"
                                break

                    cand_uuid = api_headers.get('restaurant_uuid')
                    if cand_uuid and len(cand_uuid) == 36:
                        rest_uuid = cand_uuid

                    if token and rest_uuid and len(rest_uuid) == 36:
                        try:
                            menu_data = go_api.fetch_menus(page, token, rest_uuid)
                        except Exception as fe2:
                            logger.warning(f"⚠️ Retry fetch menus gagal ({fe2}).")

                if not menu_data:
                    # Emergency fallback: Try reading offline cache file if available
                    cache_path = _find_gofood_cache_file(merchant_id)
                    if cache_path and os.path.exists(cache_path):
                        try:
                            with open(cache_path, "r", encoding="utf-8") as f:
                                menu_data = json.load(f)
                                logger.info(f"📦 Emergency Fallback: Berhasil membaca menu GoFood dari offline cache {os.path.basename(cache_path)}")
                        except Exception as e:
                            logger.error(f"Gagal membaca offline cache: {e}")

                if not token and not menu_data:
                    raise Exception("Gagal menangkap Authorization Token untuk GoFood setelah percobaan fresh login.")

                if not menu_data:
                    raise Exception(f"Gagal menarik menu GoFood ({merchant_id}) untuk perbandingan harga.")

                group_id = api_headers.get('menu_group_id')
                if not group_id and rest_uuid and token and len(rest_uuid) == 36:
                    try:
                        mg_data = go_api.fetch_menu_groups(page, token, rest_uuid)
                        if isinstance(mg_data, str):
                            group_id = mg_data
                        elif isinstance(mg_data, list) and len(mg_data) > 0:
                            group_id = mg_data[0].get('id') or mg_data[0].get('common_id')
                        elif isinstance(mg_data, dict):
                            group_id = mg_data.get('menu_group_id') or mg_data.get('v2_menus_group_id') or mg_data.get('id')
                            if not group_id:
                                mgs = mg_data.get('menu_groups') or mg_data.get('data') or []
                                if mgs and len(mgs) > 0:
                                    group_id = mgs[0].get('id') or mgs[0].get('common_id')
                        logger.info(f"🔑 Retrived menu_group_id via API fallback: {group_id}")
                    except Exception as e:
                        logger.warning(f"Could not fetch menu_groups fallback: {e}")

                if not token:
                    raise Exception("Gagal menangkap Authorization Token untuk GoFood setelah percobaan fresh login.")

                if not menu_data:
                    raise Exception("Gagal menarik menu GoFood untuk perbandingan harga.")

                categories = go_api.parse_menus(menu_data)
                go_items_by_id = {}
                for cat in categories:
                    for item in cat.get("menu_items") or []:
                        iid = item.get("common_id") or item.get("id")
                        go_items_by_id[str(iid)] = {
                            "item": item,
                            "category_id": cat.get("id"),
                            "category_common_id": cat.get("common_id")
                        }

                # Fetch active MPP promotions to prevent updating promo items
                mpp_promos_map = {}
                try:
                    from menu_core.gofood import fetch_gofood_mpp_promotions
                    clean_tok = token.replace("Bearer ", "").strip() if token else ""
                    mpp_promos_map = fetch_gofood_mpp_promotions(clean_tok, rest_uuid)
                except Exception as mpp_err:
                    logger.warning(f"⚠️ Gagal fetch MPP promo map di run_push_price_job: {mpp_err}")

                # ──────────────────────────────────────────────────────────────
                # BARRIER: Tunggu x-passkey DAN menu_group_id dari SPA sebelum
                # memulai PATCH. Kedua nilai ini dikirim SPA via background
                # request setelah halaman dimuat — tanpa ini item pertama 403.
                # ──────────────────────────────────────────────────────────────
                _need_passkey = not api_headers.get('x-passkey')
                _need_groupid = not api_headers.get('menu_group_id') and not group_id
                if _need_passkey or _need_groupid:
                    _missing = []
                    if _need_passkey: _missing.append("x-passkey")
                    if _need_groupid: _missing.append("menu_group_id")
                    logger.info(f"⏳ [BARRIER] Menunggu dari SPA: {', '.join(_missing)} (maks 30 detik)...")
                    _barrier_start = time.time()
                    while (time.time() - _barrier_start) < 30:
                        _has_pk = bool(api_headers.get('x-passkey'))
                        _has_gid = bool(api_headers.get('menu_group_id') or group_id)
                        if _has_pk and _has_gid:
                            logger.info("✅ [BARRIER] x-passkey + menu_group_id tertangkap! Memulai PATCH.")
                            break
                        page.wait_for_timeout(500)
                    else:
                        _still_missing = []
                        if not api_headers.get('x-passkey'): _still_missing.append("x-passkey")
                        if not (api_headers.get('menu_group_id') or group_id): _still_missing.append("menu_group_id")
                        if _still_missing:
                            logger.warning(f"⚠️ [BARRIER] Timeout 30 detik. Masih belum ada: {', '.join(_still_missing)}. Melanjutkan...")

                # Update group_id dari api_headers jika baru ditangkap BARRIER
                if not group_id and api_headers.get('menu_group_id'):
                    group_id = api_headers['menu_group_id']
                    logger.info(f"🔑 [BARRIER] group_id diupdate dari SPA listener: {group_id}")

                for idx, update in enumerate(updates_list):
                    item_id = update["item_id"]
                    new_price = update["new_price"]

                    item_info = go_items_by_id.get(item_id)
                    if not item_info:
                        fail_count += 1
                        trail = AuditTrail(
                            job_id=job.id,
                            outlet_id=outlet.id,
                            item_id=item_id,
                            item_name=item_id,
                            change_type="PRICE_UPDATE",
                            field_changed="price",
                            old_value=None,
                            new_value=str(new_price),
                            status="FAILED",
                            error_message="Item ID tidak ditemukan di menu GoFood."
                        )
                        db.add(trail)
                        db.commit()
                        continue

                    orig_item = item_info["item"]
                    cat_common_id = item_info["category_common_id"] or item_info["category_id"]

                    # Cek active promo pada item GoFood
                    item_name = (orig_item.get("name") or "").strip().lower()
                    mpp_info = mpp_promos_map.get(str(item_id).strip()) or mpp_promos_map.get(item_name)
                    promo_info = orig_item.get("promo_info") or orig_item.get("discount") or orig_item.get("campaign") or mpp_info
                    original_p = float(orig_item.get("original_price") or orig_item.get("list_price") or 0)
                    cur_p = float(orig_item.get("price") or 0)
                    is_go_promo = bool(promo_info) or (original_p > cur_p > 0)

                    # Klasifikasi tipe promo: Nominal vs Percentage
                    is_nominal_promo = False
                    promo_desc = ""
                    if is_go_promo:
                        if isinstance(promo_info, dict):
                            pct = promo_info.get("discount_percentage") or promo_info.get("percentage")
                            val = promo_info.get("discount_value") or promo_info.get("value") or promo_info.get("amount")
                            if pct and float(pct) > 0:
                                is_nominal_promo = False
                                promo_desc = f"Persentase ({int(float(pct))}%)"
                            elif val and float(val) > 0:
                                is_nominal_promo = True
                                promo_desc = f"Nominal (Rp {int(float(val)):,})"
                            else:
                                is_nominal_promo = False
                        else:
                            is_nominal_promo = False

                    # Hanya LOCK jika promo bertipe NOMINAL (Fixed Amount)
                    if is_go_promo and is_nominal_promo:
                        item_label = orig_item.get("name", item_id)
                        err_promo = f"Item '{item_label}' sedang dalam promo nominal tetap GoFood ({promo_desc}). Perubahan harga dasar dikunci untuk mencegah kerugian margin."
                        logger.warning(f"🔒 Promo Nominal GoFood Terdeteksi: {err_promo}")
                        trail = AuditTrail(
                            job_id=job.id,
                            outlet_id=outlet.id,
                            item_id=item_id,
                            item_name=item_label,
                            change_type="PRICE_UPDATE",
                            field_changed="price",
                            old_value=str(cur_p),
                            new_value=str(new_price),
                            status="SKIPPED_ACTIVE_PROMO",
                            error_message=err_promo
                        )
                        db.add(trail)
                        db.commit()
                        continue
                    elif is_go_promo and not is_nominal_promo:
                        logger.info(f"⚡ Item '{orig_item.get('name')}' sedang promo persentase ({promo_desc or 'Dynamic %'}). Push harga dasar tetap diizinkan.")

                    old_price = int(float(orig_item.get('price') or 0))
                    target_price = float(new_price)
                    steps = calculate_price_steps(old_price, target_price, max_step_pct=0.15)

                    if len(steps) > 1:
                        logger.info(f"📊 Single/Batch PUSH bertahap (>15%) untuk {orig_item.get('name')}: Rp {old_price:,.0f} -> Rp {target_price:,.0f} via {len(steps)} tahapan: {steps}")

                    v2_payload = {
                        "menu_common_id": orig_item.get('menu_common_id') or cat_common_id,
                        "image_url": orig_item.get('image_url', orig_item.get('image', '')),
                        "name": orig_item.get('name'),
                        "description": orig_item.get('description', ''),
                        "price": int(steps[0]),
                        "active": orig_item.get('is_active', orig_item.get('active', True)),
                        "signature": orig_item.get('signature', False)
                    }

                    patch_group_id = group_id or api_headers.get('menu_group_id') or orig_item.get('menu_common_id') or cat_common_id
                    passkey = api_headers.get('x-passkey') or "1729b182-c60e-4568-849d-5eb7d794fd09"
                    
                    headers_direct = {
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'id',
                        'Authentication-Type': 'go-id',
                        'Authorization': token,
                        'Content-Type': 'application/json',
                        'Gojek-Country-Code': 'ID',
                        'x-passkey': passkey,
                        'Origin': 'https://portal.gofoodmerchant.co.id',
                        'Referer': 'https://portal.gofoodmerchant.co.id/'
                    }

                    # V2 PATCH via context.request (bypass CORS) — Opsi Utama tanpa variant_category_common_ids
                    v2_url = f'https://api.gojekapi.com/gofood/merchant/v2/menu_groups/{patch_group_id}/menu_items/{item_id}'
                    
                    # Function helper dengan auto-retry jika terkena Rate Limit (HTTP 429/403/503)
                    def send_patch_request(payload_data, max_retries=2):
                        for attempt_idx in range(max_retries + 1):
                            try:
                                cr_res = context.request.fetch(
                                    v2_url,
                                    method='PATCH',
                                    headers=headers_direct,
                                    data=json.dumps(payload_data)
                                )
                                status_code = cr_res.status
                                if status_code in (429, 403, 503, 504) and attempt_idx < max_retries:
                                    backoff_sec = 12.0 * (attempt_idx + 1)
                                    logger.warning(f"⚠️ Terdeteksi Rate Limit (HTTP {status_code}) pada item {item_id}. Menunggu {int(backoff_sec)} detik untuk cooldown (attempt {attempt_idx+1}/{max_retries})...")
                                    time.sleep(backoff_sec)
                                    continue
                                return {'ok': cr_res.ok, 'status': status_code, 'body': cr_res.text()}
                            except Exception as ex:
                                if attempt_idx < max_retries:
                                    time.sleep(3.0)
                                    continue
                                return {'ok': False, 'error': str(ex)}

                    res = None
                    for step_idx, step_p in enumerate(steps):
                        v2_payload["price"] = int(step_p)
                        res = send_patch_request(v2_payload)

                        # Jika terkena HTTP 429, berikan cooldown dan JANGAN langsung pemboman request fallback
                        if res and res.get('status') == 429:
                            logger.warning(f"⚠️ GoFood API Rate Limited (HTTP 429). Mengistirahatkan proses 10 detik agar server pulih...")
                            time.sleep(10.0)

                        # Fallback 1: Jika gagal (bukan 429) dan ada variant_category_common_ids, coba sertakan
                        if (not res or not res.get('ok')) and res.get('status') != 429:
                            time.sleep(0.6)  # Jeda jeda sebelum fallback
                            vars_ids = orig_item.get('variant_category_common_ids') or orig_item.get('variant_category_ids')
                            if vars_ids and isinstance(vars_ids, list) and len(vars_ids) > 0:
                                v2_payload_with_vars = dict(v2_payload)
                                v2_payload_with_vars["variant_category_common_ids"] = vars_ids
                                res_var = send_patch_request(v2_payload_with_vars, max_retries=1)
                                if res_var and res_var.get('ok'):
                                    res = res_var

                        # Fallback 2: Jika masih gagal (dan bukan 429)
                        if (not res or not res.get('ok')) and res.get('status') != 429:
                            status_code = res.get('status', '?') if res else '?'
                            body_err = (res.get('body') or '')[:500] if res else ''
                            logger.warning(f"GoFood V2 PATCH gagal (HTTP {status_code}), Body: {body_err}, Error: {res.get('error')}. Fallback ke V1 PUT...")
                            time.sleep(0.6)  # Jeda jeda sebelum fallback V1

                            v1_payload = {
                                "name": orig_item.get('name'),
                                "price": int(step_p),
                                "active": orig_item.get('active', True),
                                "description": orig_item.get('description', ''),
                                "image": orig_item.get('image_url', orig_item.get('image', ''))
                            }
                            v1_item_id = orig_item.get('id') or orig_item.get('common_id') or item_id
                            
                            # V1 PUT via context.request (bypass CORS)
                            if v1_item_id:
                                v1_url = f'https://api.gojekapi.com/gofood/merchant/v1/restaurants/{rest_uuid}/menu_items/{v1_item_id}'
                                try:
                                    cr_v1 = context.request.fetch(
                                        v1_url,
                                        method='PUT',
                                        headers=headers_direct,
                                        data=json.dumps(v1_payload)
                                    )
                                    res = {'ok': cr_v1.ok, 'status': cr_v1.status, 'body': cr_v1.text()}
                                except Exception as e:
                                    res = {'ok': False, 'error': str(e)}
                            else:
                                res = {'ok': False, 'error': 'No V1 item ID available for fallback'}

                        if not (res and res.get('ok')):
                            logger.error(f"❌ Tahap harga Rp {step_p:,.0f} gagal: {res.get('error') or res.get('body')}")
                            break
                        
                        if len(steps) > 1:
                            logger.info(f"   [Tahap {step_idx+1}/{len(steps)}] Berhasil PUSH harga intermediate: Rp {step_p:,.0f}")
                            if step_idx < len(steps) - 1:
                                time.sleep(1.5)

                    if res and res.get('ok'):
                        success_count += 1
                        status_str = "SUCCESS"
                        err_msg = None
                    else:
                        fail_count += 1
                        status_str = "FAILED"
                        err_msg = res.get('body') or "GoFood API error."

                    # Pacing delay bervariatif (random jitter 1.2s - 2.5s) untuk memberikan jeda aman antar item
                    import random
                    time.sleep(random.uniform(1.2, 2.5))

                    # Jeda istirahat (batch breather) setiap 10 item agar token bucket rate-limit GoFood pulih
                    if (idx + 1) % 20 == 0 and (idx + 1) < total_updates:
                        logger.info(f"☕ Batch pause (item {idx+1}/{total_updates}): istirahat 3 detik...")
                        time.sleep(3.0)

                    # Update progress setiap 5 item
                    if (idx + 1) % 5 == 0 or (idx + 1) == total_updates:
                        job.progress_pct = int(40 + ((idx + 1) / total_updates) * 55)
                        job.current_step = f"Memproses update harga GoFood ({idx + 1}/{total_updates})..."
                        db.commit()

                    trail = AuditTrail(
                        job_id=job.id,
                        outlet_id=outlet.id,
                        item_id=item_id,
                        item_name=orig_item.get("name", item_id),
                        change_type="PRICE_UPDATE",
                        field_changed="price",
                        old_value=str(orig_item.get("price", 0)),
                        new_value=str(new_price),
                        status=status_str,
                        error_message=err_msg
                    )
                    db.add(trail)
                    db.commit()

                browser.close()

        if total_updates == 0:
            job.status = "FAILED"
            job.error_message = "Tidak ada item yang dipilih untuk dipublikasikan."
            job.current_step = "Gagal: tidak ada item dipublikasikan."
        elif success_count == total_updates and success_count > 0:
            job.status = "SUCCESS"
            job.error_message = None
            job.current_step = f"Pembaruan harga sukses & terverifikasi! ({success_count}/{total_updates} item berhasil)."
        elif success_count > 0 and fail_count > 0:
            job.status = "PARTIAL_SUCCESS"
            job.error_message = f"Sebagian item gagal: {success_count} sukses, {fail_count} gagal dari {total_updates} item."
            job.current_step = f"Sebagian item gagal diperbarui ({success_count} sukses, {fail_count} gagal)."
        else:
            job.status = "FAILED"
            trails_fails = db.query(AuditTrail).filter(AuditTrail.job_id == job.id, AuditTrail.status == "FAILED").all()
            if trails_fails and any("user membatalkan otp" in str(tf.error_message).lower() for tf in trails_fails if tf.error_message):
                job.error_message = "user membatalkan otp"
                job.current_step = "Gagal: user membatalkan otp"
            else:
                job.error_message = f"Pembaruan harga gagal! 0 dari {total_updates} item berhasil diperbarui."
                job.current_step = f"Gagal memperbarui harga! (0/{total_updates} berhasil)."

        trails = db.query(AuditTrail).filter(AuditTrail.job_id == job.id).all()
        breakdown_list = []
        for t in trails:
            old_p = None
            try:
                if t.old_value: old_p = float(t.old_value)
            except: pass
            
            new_p = None
            try:
                if t.new_value: new_p = float(t.new_value)
            except: pass

            breakdown_list.append({
                "item_id": t.item_id,
                "item_name": t.item_name or t.item_id,
                "old_price": old_p,
                "requested_price": new_p,
                "verified_price": new_p if t.status == "SUCCESS" else None,
                "status": t.status,
                "error_message": t.error_message
            })

        # Update local Excel catalog file with the new prices so that get_outlet_menu_items returns updated prices
        excel_path = None
        if success_count > 0:
            try:
                import openpyxl
                latest_pull = db.query(Job).filter(
                    Job.outlet_id == outlet_id,
                    Job.job_type == "PULL",
                    Job.status.in_(["SUCCESS", "PARTIAL_SUCCESS"])
                ).order_by(Job.completed_at.desc()).first()
                if latest_pull and latest_pull.result_metadata:
                    excel_path = latest_pull.result_metadata.get("excel_path")
                
                if not excel_path or not os.path.exists(excel_path):
                    import re
                    raw_outlet = outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown'
                    clean_outlet = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
                    clean_outlet = re.sub(r'\s+', ' ', clean_outlet).lower()
                    exports_dir = BASE_DIR / "data" / "exports" / outlet.platform / clean_outlet
                    excel_files = list(exports_dir.glob("*.xlsx")) if exports_dir.exists() else []
                    if excel_files:
                        excel_path = str(excel_files[0])

                if excel_path and os.path.exists(excel_path):
                    wb = openpyxl.load_workbook(excel_path)
                    if 'Item' in wb.sheetnames:
                        sheet = wb['Item']
                        header_row = [cell.value for cell in sheet[1]]
                        try:
                            item_id_idx = header_row.index('Item ID') + 1
                            price_idx = header_row.index('Current Real Price (Rp)') + 1
                            fake_price_idx = (header_row.index('Current Fake Price (Rp)') + 1) if 'Current Fake Price (Rp)' in header_row else None
                            
                            successful_updates = {}
                            for t in trails:
                                if t.status == "SUCCESS":
                                    try:
                                        successful_updates[str(t.item_id).strip()] = int(float(t.new_value))
                                    except:
                                        pass
                                        
                            for row_idx in range(2, sheet.max_row + 1):
                                cell_item_id = str(sheet.cell(row=row_idx, column=item_id_idx).value).strip()
                                if cell_item_id in successful_updates:
                                    new_val = successful_updates[cell_item_id]
                                    if fake_price_idx:
                                        curr_fake = sheet.cell(row=row_idx, column=fake_price_idx).value
                                        if curr_fake is not None and str(curr_fake).strip() not in ("", "0"):
                                            sheet.cell(row=row_idx, column=fake_price_idx).value = new_val
                                        else:
                                            sheet.cell(row=row_idx, column=price_idx).value = new_val
                                    else:
                                        sheet.cell(row=row_idx, column=price_idx).value = new_val
                                    
                            wb.save(excel_path)
                            logger.info(f"💾 Updated Excel catalog with new prices at: {excel_path}")
                        except ValueError as ve:
                            logger.error(f"Excel column missing during update: {ve}")
                    wb.close()
            except Exception as e:
                logger.error(f"Gagal memperbarui file Excel lokal dengan harga baru: {e}")

        job.progress_pct = 100
        job.result_metadata = {
            "total_updates": total_updates,
            "success_count": success_count,
            "fail_count": fail_count,
            "items_breakdown": breakdown_list,
            "completed_at": datetime.utcnow().isoformat(),
            "excel_path": excel_path
        }
        job.completed_at = datetime.utcnow()
        db.commit()

    except Exception as e:
        logger.error(f"❌ Job {job_id} failed: {e}")
        job.status = "FAILED"
        if "user membatalkan otp" in str(e).lower():
            job.error_message = "user membatalkan otp"
            job.current_step = "Gagal: user membatalkan otp"
        else:
            job.error_message = str(e)
            err_msg = f"Terjadi kesalahan: {str(e)}"
            job.current_step = err_msg if len(err_msg) <= 255 else err_msg[:252] + "..."
        job.completed_at = datetime.utcnow()
        db.commit()

    finally:
        if lock:
            try:
                lock.release()
                logger.info(f"🔓 Job {job_id} ({platform}) lock released.")
            except Exception as le:
                logger.warning(f"⚠️ Failed to release lock: {le}")
        db.close()


@app.post("/api/jobs/push-price", response_model=JobResponse, status_code=status.HTTP_202_ACCEPTED)
def trigger_push_price_job(request: PriceUpdateRequest, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """Triggers a background job to push price changes to the applicator merchant portal."""
    outlet = db.query(Outlet).filter(Outlet.id == request.outlet_id).first()
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
        
    updates_payload = []
    for item in request.updates:
        updates_payload.append({
            "item_id": item.item_id,
            "category_id": item.category_id or "",
            "item_name": item.item_name or "",
            "new_price": item.new_price
        })

    new_job = Job(
        outlet_id=outlet.id,
        job_type="PUSH_UPDATE",
        platform=outlet.account.platform,
        status="PENDING",
        progress_pct=0,
        current_step="Mengantrekan pembaruan harga...",
        payload={"store_id": outlet.store_id, "merchant_name": outlet.merchant_name, "updates_count": len(updates_payload)}
    )
    db.add(new_job)
    db.commit()
    db.refresh(new_job)

    background_tasks.add_task(run_push_price_job, new_job.id, outlet.id, updates_payload)
    return new_job


# ─── JOBS ENDPOINTS ───────────────────────────────────────────────────────────

@app.post("/api/jobs/pull", response_model=JobResponse, status_code=status.HTTP_202_ACCEPTED)
def trigger_pull_job(outlet_id: uuid.UUID, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
    
    # Create Job log entry
    new_job = Job(
        outlet_id=outlet.id,
        job_type="PULL",
        platform=outlet.account.platform,
        status="PENDING",
        progress_pct=0,
        current_step="Mengantrekan tugas penarikan...",
        payload={"store_id": outlet.store_id, "merchant_name": outlet.merchant_name}
    )
    db.add(new_job)
    db.commit()
    db.refresh(new_job)

    # Dispatch to background executor thread
    background_tasks.add_task(run_pull_job, new_job.id, outlet.id)
    return new_job

@app.get("/api/jobs/download-file")
def download_file_by_path(path: str):
    abs_path = os.path.abspath(path)
    base_exports = os.path.abspath(str(BASE_DIR / "data" / "exports"))
    if not abs_path.startswith(base_exports):
        raise HTTPException(status_code=403, detail="Akses ditolak")
    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="File tidak ditemukan di server")
    filename = os.path.basename(abs_path)
    return FileResponse(
        path=abs_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/jobs/{job_id}", response_model=JobResponse)
def get_job_status(job_id: uuid.UUID, db: Session = Depends(get_db)):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job

@app.get("/api/jobs", response_model=List[JobResponse])
def list_jobs(db: Session = Depends(get_db)):
    return db.query(Job).order_by(Job.created_at.desc()).limit(50).all()

@app.get("/api/jobs/download/{job_id}")
def download_job_file(job_id: uuid.UUID, db: Session = Depends(get_db)):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status != "SUCCESS":
        raise HTTPException(status_code=400, detail="Job is not completed successfully")
    
    excel_path = job.result_metadata.get("excel_path") if job.result_metadata else None
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail="Excel file not found on server")
        
    filename = os.path.basename(excel_path)
    return FileResponse(
        path=excel_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/api/jobs/combine-c5")
def combine_c5_endpoint(request: CombineC5Request, db: Session = Depends(get_db)):
    from menu_core.c5_combiner import combine_c5
    from upload_drive import upload_combined_to_drive
    from datetime import datetime

    excel_paths = []
    outlet_name = (request.outlet_name or "").strip()
    owner_name = None

    if request.job_ids:
        for jid in request.job_ids:
            try:
                j_uuid = uuid.UUID(str(jid))
                job = db.query(Job).filter(Job.id == j_uuid).first()
                if job and job.result_metadata and job.result_metadata.get("excel_path"):
                    epath = job.result_metadata["excel_path"]
                    if os.path.exists(epath) and epath not in excel_paths:
                        excel_paths.append(epath)
                if not owner_name and job and job.outlet:
                    owner_name = job.outlet.owner or job.outlet.merchant_name or job.outlet.nama_outlet
            except Exception as ex:
                logger.warning(f"Error resolving job ID {jid}: {ex}")

    if not owner_name:
        owner_name = outlet_name if outlet_name else "Combined Outlet"

    import re
    clean_owner_name = "".join(c for c in owner_name if c.isalnum() or c in (' ', '_', '-')).strip()
    clean_outlet_filename = "".join(c for c in (outlet_name or owner_name) if c.isalnum() or c in (' ', '_', '-')).strip()
    clean_folder_name = re.sub(r'\s+', ' ', clean_outlet_filename).lower()

    if not excel_paths and outlet_name:
        exports_root = BASE_DIR / "data" / "exports"
        for p in exports_root.glob(f"**/{clean_folder_name}/**/*.xlsx"):
            if "combined" not in str(p) and str(p) not in excel_paths:
                excel_paths.append(str(p))

    if not excel_paths:
        raise HTTPException(status_code=400, detail="Tidak ada file Excel C5 yang ditemukan untuk digabungkan.")

    combined_dir = BASE_DIR / "data" / "exports" / "combined" / clean_folder_name
    combined_dir.mkdir(parents=True, exist_ok=True)

    excel_filename = f"O.C5 {clean_outlet_filename}.xlsx"
    combined_path = str(combined_dir / excel_filename)

    ok = combine_c5(excel_paths, combined_path)
    if not ok:
        raise HTTPException(status_code=500, detail="Gagal menggabungkan file C5.")

    # Versioning format for Google Drive filename: C5. YYYY-MM-DD HH:MM <Nama Owner>.xlsx
    timestamp_version = datetime.now().strftime("%Y-%m-%d %H:%M")
    drive_filename = f"C5. {timestamp_version} {clean_owner_name}.xlsx"

    # Upload to Google Drive using folderName: Owner Name, fileName: C5. YYYY-MM-DD HH:MM <Nama Owner>.xlsx
    gspread_url = upload_combined_to_drive(combined_path, clean_owner_name, custom_filename=drive_filename)

    from urllib.parse import quote
    return {
        "ok": True,
        "excel_filename": excel_filename,
        "excel_path": combined_path,
        "gspread_url": gspread_url,
        "download_url": f"/api/jobs/download-file?path={quote(combined_path)}",
        "combined_count": len(excel_paths),
        "outlet_name": owner_name
    }

def calculate_price_steps(current_price: float, target_price: float, max_step_pct: float = 0.15) -> list[int]:
    """Calculates intermediate prices to reach target_price without any single step exceeding max_step_pct change."""
    import math
    curr = float(current_price)
    target = float(target_price)
    if curr <= 0 or curr == target:
        return [int(round(target))]

    steps = []
    if target > curr:
        while curr < target:
            next_p = curr * (1.0 + max_step_pct)
            if next_p >= target:
                steps.append(int(round(target)))
                break
            else:
                next_p_rounded = int(math.floor(next_p / 100.0) * 100)
                if next_p_rounded <= curr:
                    next_p_rounded = int(curr) + 100
                if next_p_rounded >= target:
                    steps.append(int(round(target)))
                    break
                steps.append(next_p_rounded)
                curr = float(next_p_rounded)
    else:  # target < curr
        while curr > target:
            next_p = curr * (1.0 - max_step_pct)
            if next_p <= target:
                steps.append(int(round(target)))
                break
            else:
                next_p_rounded = int(math.ceil(next_p / 100.0) * 100)
                if next_p_rounded >= curr:
                    next_p_rounded = int(curr) - 100
                if next_p_rounded <= target:
                    steps.append(int(round(target)))
                    break
                steps.append(next_p_rounded)
                curr = float(next_p_rounded)
    return steps


# ─── C5 MENU PUSH & PARSER ENDPOINTS ──────────────────────────────────────────

def _push_c5_gofood_for_merchant(email: str, password: str, merchant_id: str, updates: list, progress_cb=None, item_result_cb=None):
    """Logs into GoFood for a single merchant and applies C5 name/price changes to the real store.

    Each entry in `updates` is a C5PushItemUpdate dict. Returns a list of per-item result dicts:
    {item_id, item_name, new_name, new_price, status: SUCCESS|FAILED, error}.
    Modelled on the working PATCH flow in run_push_price_job.
    """
    from playwright.sync_api import sync_playwright
    from Gofood.GO.actions import _menu_api as go_api

    if not merchant_id:
        raise Exception("Merchant ID (store_id) tidak tersedia untuk outlet GoFood.")
    merchant_id = str(merchant_id).strip()
    if merchant_id.startswith("GM"):
        merchant_id = merchant_id[1:]
    elif merchant_id.isdigit():
        merchant_id = "G" + merchant_id

    results = []

    with sync_playwright() as p:
        import re
        from login_gofood import load_gofood_session
        from src.core.browser_factory import launch_universal_playwright_browser

        sanitized_email = re.sub(r'[^a-zA-Z0-9_.-]', '_', (email or "").strip().lower())
        session_path = os.path.join(BASE_DIR, "Gofood", f"session_gofood_{sanitized_email}.json")
        cached_data = load_gofood_session(email) if email else None

        headless_env = os.getenv("HEADLESS") or os.getenv("HEADLESS_GOFOOD")
        is_headless = headless_env.lower() in ("true", "1", "yes") if headless_env else True

        browser, proc = launch_universal_playwright_browser(p, headless=is_headless)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        if cached_data and cached_data.get("cookies"):
            try:
                context.add_cookies(cached_data["cookies"])
            except Exception as e:
                logger.warning(f"Error adding cookies: {e}")

        page = context.new_page()
        api_headers = {}

        if cached_data and cached_data.get("access_token"):
            tok = cached_data["access_token"]
            api_headers['authorization'] = tok if tok.startswith("Bearer ") else f"Bearer {tok}"

        def capture_headers(request):
            url_lower = request.url.lower()
            if "api.gojekapi.com" in url_lower or "api.gobiz.co.id" in url_lower or "portal.gofoodmerchant.co.id" in url_lower:
                h = request.headers
                if 'authorization' in h:
                    api_headers['authorization'] = h['authorization']
                if 'x-passkey' in h:
                    api_headers['x-passkey'] = h['x-passkey']
            if "restaurants/" in url_lower:
                parts = request.url.split("/")
                for i, part in enumerate(parts):
                    if part.lower() == "restaurants" and i + 1 < len(parts):
                        candidate = parts[i + 1].split("?")[0]
                        if len(candidate) == 36 and "-" in candidate:
                            api_headers['restaurant_uuid'] = candidate
            if "menu_groups/" in url_lower:
                parts = request.url.split("/")
                for i, part in enumerate(parts):
                    if part.lower() == "menu_groups" and i + 1 < len(parts):
                        candidate = parts[i + 1].split("?")[0]
                        if len(candidate) == 36 and "-" in candidate:
                            api_headers['menu_group_id'] = candidate

        page.on("request", capture_headers)

        def perform_fresh_login():
            logger.info(f"🔄 Token GoFood expired/tidak ditemukan. Menjalankan login_outlet untuk {email} (store: {merchant_id})...")
            try:
                import threading, asyncio
                res_box = [None]
                err_box = [None]
                
                outlet_meta = {
                    'store_id': merchant_id,
                    'email': email,
                    'emails': [email] if email else [],
                }

                def _worker():
                    try:
                        asyncio.set_event_loop(asyncio.new_event_loop())
                        from login_gofood import login_outlet
                        res_box[0] = login_outlet(outlet_meta)
                    except Exception as ex:
                        err_box[0] = ex

                t = threading.Thread(target=_worker)
                t.start()
                t.join()

                if err_box[0]:
                    raise err_box[0]

                res = res_box[0]
                if res and res.get('access_token'):
                    tok = res['access_token']
                    api_headers['authorization'] = tok if tok.startswith("Bearer ") else f"Bearer {tok}"
                    if res.get('restaurant_uuid'):
                        api_headers['restaurant_uuid'] = res['restaurant_uuid']
                    if res.get('cookies'):
                        try: context.add_cookies(res['cookies'])
                        except Exception: pass
                    return True
            except Exception as e:
                logger.warning(f"Terjadi kesalahan saat login_outlet: {e}")
            return False

        try:
            page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}/menu-items", wait_until="domcontentloaded")
            time.sleep(2)

            if "/auth" in page.url or "login" in page.url or not api_headers.get('authorization'):
                perform_fresh_login()
                page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}/menu-items", wait_until="domcontentloaded")
                time.sleep(3)
                if "/auth" in page.url or "login" in page.url:
                    raise Exception(f"Gagal login ke GoFood portal untuk {email}. Sesi telah kadaluarsa dan membutuhkan login/OTP ulang via 'python login_gofood.py'.")

            # Wait dynamically (up to 20s) for the ids the app emits: restaurant_uuid,
            # authorization, x-passkey and — needed for the V2 PATCH — menu_group_id.
            start_wait = time.time()
            while (time.time() - start_wait) < 20:
                if (api_headers.get('restaurant_uuid') and api_headers.get('authorization')
                        and api_headers.get('x-passkey') and api_headers.get('menu_group_id')):
                    break
                page.wait_for_timeout(500)

            token = api_headers.get('authorization')
            if not token:
                cookies = context.cookies()
                for c in cookies:
                    if c['name'] in ('access_token', 'token', 'gobiz_token'):
                        token = f"Bearer {c['value']}"
                        break

            # Resolve restaurant UUID (36-char). Header capture may miss it, so fall
            # back to localStorage/sessionStorage and finally to the cached PULL file.
            rest_uuid = api_headers.get('restaurant_uuid')
            if not rest_uuid or len(rest_uuid) != 36:
                try:
                    uuid_eval = page.evaluate("""() => {
                        const re = /[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}/i;
                        for (let i = 0; i < localStorage.length; i++) {
                            const m = re.exec(localStorage.getItem(localStorage.key(i)) || '');
                            if (m) return m[0];
                        }
                        for (let i = 0; i < sessionStorage.length; i++) {
                            const m = re.exec(sessionStorage.getItem(sessionStorage.key(i)) || '');
                            if (m) return m[0];
                        }
                        const u = re.exec(window.location.href);
                        return u ? u[0] : null;
                    }""")
                    if uuid_eval:
                        rest_uuid = uuid_eval
                except Exception as e:
                    logger.warning(f"Gagal ekstrak rest_uuid dari web storage: {e}")

            cache_menu_path = os.path.join(BASE_DIR, "Gofood", "API", f"menu-response-{merchant_id}.json")
            if (not rest_uuid or len(rest_uuid) != 36) and os.path.exists(cache_menu_path):
                try:
                    with open(cache_menu_path, "r") as f:
                        cdata = json.load(f)
                    for m in (cdata.get("menus") or cdata.get("categories") or []):
                        cand = m.get("restaurant_id") or m.get("restaurant_uuid")
                        if cand and len(cand) == 36:
                            rest_uuid = cand
                            break
                except Exception as e:
                    logger.warning(f"Gagal baca cached restaurant_id: {e}")

            menu_data = None
            if token and rest_uuid and len(rest_uuid) == 36:
                menu_data = go_api.fetch_menus(page, token, rest_uuid)

            if not token or not menu_data:
                logger.warning("⚠️ GoFood session/menu tidak valid. Memicu fresh login ulang...")
                perform_fresh_login()
                page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}/menu-items", wait_until="domcontentloaded")
                time.sleep(3)
                start_wait = time.time()
                while (time.time() - start_wait) < 15:
                    if api_headers.get('authorization'):
                        break
                    page.wait_for_timeout(500)
                token = api_headers.get('authorization') or token
                if api_headers.get('restaurant_uuid') and len(api_headers['restaurant_uuid']) == 36:
                    rest_uuid = api_headers['restaurant_uuid']
                if token and rest_uuid and len(rest_uuid) == 36:
                    menu_data = go_api.fetch_menus(page, token, rest_uuid)

            if not token:
                raise Exception("Gagal menangkap Authorization Token GoFood setelah fresh login.")

            # Resolve menu_group_id (needed for V2 PATCH path)
            group_id = api_headers.get('menu_group_id')
            if not group_id and rest_uuid and token and len(rest_uuid) == 36:
                try:
                    mg_data = go_api.fetch_menu_groups(page, token, rest_uuid)
                    if isinstance(mg_data, str):
                        group_id = mg_data
                    elif isinstance(mg_data, list) and len(mg_data) > 0:
                        group_id = mg_data[0].get('id') or mg_data[0].get('common_id')
                    elif isinstance(mg_data, dict):
                        group_id = mg_data.get('menu_group_id') or mg_data.get('v2_menus_group_id') or mg_data.get('id')
                        if not group_id:
                            mgs = mg_data.get('menu_groups') or mg_data.get('data') or []
                            if mgs and len(mgs) > 0:
                                group_id = mgs[0].get('id') or mgs[0].get('common_id')
                except Exception as e:
                    logger.warning(f"Could not fetch menu_groups fallback: {e}")

            # If the live V1 menu fetch failed but we have a group_id, use V2 menus
            # (only needs group_id) — else fall back to the cached PULL file on disk.
            if not menu_data and group_id and token:
                menu_data = go_api.fetch_menus_v2(page, token, group_id)
            if not menu_data and os.path.exists(cache_menu_path):
                logger.warning("⚠️ Memakai cache menu PULL terakhir untuk indeks item.")
                try:
                    with open(cache_menu_path, "r") as f:
                        menu_data = json.load(f)
                except Exception as e:
                    logger.warning(f"Gagal baca cache menu: {e}")

            if not menu_data:
                raise Exception("Gagal menarik menu GoFood untuk perbandingan.")

            # Index live menu items by id/common_id
            categories = go_api.parse_menus(menu_data)
            go_items_by_id = {}
            for cat in categories:
                cat_group = cat.get("menu_common_id") or cat.get("common_id") or cat.get("id")
                for it in cat.get("menu_items") or []:
                    iid = it.get("common_id") or it.get("id")
                    go_items_by_id[str(iid)] = {
                        "item": it,
                        "category_id": cat.get("id"),
                        "category_common_id": cat.get("common_id") or cat_group,
                    }

            # Fetch active MPP promotions to prevent updating promo items
            mpp_promos_map = {}
            try:
                from menu_core.gofood import fetch_gofood_mpp_promotions
                clean_tok = token.replace("Bearer ", "").strip() if token else ""
                mpp_promos_map = fetch_gofood_mpp_promotions(clean_tok, rest_uuid)
            except Exception as mpp_err:
                logger.warning(f"⚠️ Gagal fetch MPP promo map di run_push_c5_job: {mpp_err}")

            passkey = api_headers.get('x-passkey') or "1729b182-c60e-4568-849d-5eb7d794fd09"
            passkey = api_headers.get('x-passkey') or "1729b182-c60e-4568-849d-5eb7d794fd09"
            headers_direct = {
                'Accept': 'application/json, text/plain, */*',
                'Accept-Language': 'id',
                'Authentication-Type': 'go-id',
                'Authorization': token,
                'Content-Type': 'application/json',
                'Gojek-Country-Code': 'ID',
                'x-passkey': passkey,
                'Origin': 'https://portal.gofoodmerchant.co.id',
                'Referer': 'https://portal.gofoodmerchant.co.id/',
            }

            patch_group_id = group_id or api_headers.get('menu_group_id')

            # ── 1. Execute New Categories Creation & Category Renames (GoFood V2 API) ──
            created_cat_ids = {}
            category_renames = {}
            for upd in updates:
                cat_id = upd.get("category_id")
                cat_name = upd.get("category")
                change_types = upd.get("changes") or upd.get("change_types") or []
                is_dict_changes = isinstance(upd.get("changes"), dict)
                is_new_cat = ("NEW_CATEGORY" in change_types) or (is_dict_changes and upd["changes"].get("is_new_category"))

                if cat_name and is_new_cat and norm_str(cat_name) not in created_cat_ids and patch_group_id:
                    try:
                        logger.info(f"📂 Creating New Category '{cat_name}' on GoFood merchant...")
                        cat_res = go_api.create_category(page, token, patch_group_id, {"name": cat_name, "active": True}, passkey=passkey)
                        if cat_res and cat_res.get("ok"):
                            res_data = cat_res.get("data") or cat_res
                            new_c_id = res_data.get("id") or res_data.get("common_id") or res_data.get("menu_common_id")
                            if new_c_id:
                                created_cat_ids[norm_str(cat_name)] = new_c_id
                            logger.info(f"✅ Category '{cat_name}' created successfully (ID: {new_c_id})")
                        else:
                            logger.warning(f"⚠️ Category creation warning for '{cat_name}': {cat_res}")
                    except Exception as cat_ex:
                        logger.warning(f"⚠️ Exception creating category '{cat_name}': {cat_ex}")

                if cat_id and cat_name and not is_new_cat:
                    is_cat_changed = ("CATEGORY_CHANGE" in change_types) or (is_dict_changes and upd["changes"].get("category_changed"))
                    if is_cat_changed:
                        category_renames[cat_id] = cat_name

            if category_renames and patch_group_id:
                for cat_id, new_cat_name in category_renames.items():
                    try:
                        logger.info(f"🏷️ Pushing Category Rename for Category ID '{cat_id}' -> '{new_cat_name}'...")
                        cat_res = go_api.update_menu_item(page, token, patch_group_id, cat_id, {"name": new_cat_name, "active": True}, passkey=passkey)
                        if cat_res and cat_res.get("ok"):
                            logger.info(f"✅ Category '{cat_id}' successfully renamed to '{new_cat_name}'")
                        else:
                            logger.warning(f"⚠️ Category rename warning for '{cat_id}': {cat_res}")
                    except Exception as cat_ex:
                        logger.warning(f"⚠️ Exception renaming category '{cat_id}': {cat_ex}")

            # ── 2. Execute Item Updates & New Item Creation (tambah_item) ──
            import random
            total = len(updates)
            for idx, upd in enumerate(updates):
                item_id = str(upd.get("item_id") or "")
                new_name = (upd.get("item_name_new") or "").strip()
                raw_price = upd.get("new_fake_price")
                new_photo = (upd.get("photo_link") or "").strip()
                new_desc = (upd.get("description") or "").strip()
                new_cat = (upd.get("category") or "").strip()

                change_types = upd.get("changes") or upd.get("change_types") or []
                is_dict_changes = isinstance(upd.get("changes"), dict)

                is_new_item = ("NEW_ITEM" in change_types) or (is_dict_changes and upd["changes"].get("is_new_item")) or (not item_id)
                want_name = ("NAME_CHANGE" in change_types) or (is_dict_changes and upd["changes"].get("name_changed")) or bool(new_name)
                want_price = ("PRICE_CHANGE" in change_types) or (is_dict_changes and upd["changes"].get("price_changed")) or (raw_price is not None)
                want_photo = ("PHOTO_CHANGE" in change_types) or (is_dict_changes and upd["changes"].get("photo_changed")) or bool(new_photo and new_photo.startswith("http"))
                want_desc = ("DESCRIPTION_CHANGE" in change_types) or (is_dict_changes and upd["changes"].get("description_changed")) or bool(new_desc)

                if progress_cb:
                    progress_cb(idx, total, upd)

                item_info = go_items_by_id.get(item_id)
                orig_item = item_info["item"] if item_info else {}
                cat_common_id = item_info["category_common_id"] if item_info else (created_cat_ids.get(norm_str(new_cat)) or patch_group_id)

                final_name = new_name if (want_name and new_name) else (orig_item.get('name') or upd.get('item_name') or "Item Baru")
                final_photo = new_photo if (want_photo and new_photo) else orig_item.get('image_url', orig_item.get('image', ''))
                final_desc = new_desc if (want_desc and new_desc) else orig_item.get('description', '')

                try:
                    final_price = int(float(raw_price)) if (want_price and raw_price is not None) else int(float(orig_item.get('price') or 0))
                except (ValueError, TypeError):
                    final_price = int(float(orig_item.get('price') or 0))

                is_deleted_item = ("DELETE_ITEM" in change_types) or (is_dict_changes and upd["changes"].get("is_deleted_item"))
                if is_deleted_item and item_id and not item_id.startswith("NEW_"):
                    logger.info(f"🗑️ Deleting item on GoFood merchant: '{item_id}' ({upd.get('item_name')})...")
                    del_res = go_api.delete_v2_menu_item(page, token, patch_group_id, item_id, passkey=passkey)
                    if not (del_res and del_res.get("ok")):
                        del_res = go_api.delete_menu_item(page, token, patch_group_id, item_id, passkey=passkey)
                    if del_res and del_res.get("ok"):
                        logger.info(f"✅ Item '{item_id}' deleted successfully!")
                        results.append({
                            "item_id": item_id, "item_name": upd.get("item_name"),
                            "status": "SUCCESS", "error": None, "action": "DELETED"
                        })
                        continue
                    else:
                        logger.warning(f"⚠️ Failed deleting item '{item_id}': {del_res}")

                # ── Handle New Item Creation (tambah_item) ──
                if is_new_item:
                    logger.info(f"✨ Adding New Item (tambah_item) to GoFood: '{final_name}'...")
                    target_cat_id = cat_common_id or patch_group_id
                    create_payload = {
                        "menu_common_id": target_cat_id,
                        "name": final_name,
                        "price": final_price,
                        "description": final_desc,
                        "image_url": final_photo,
                        "active": True,
                        "signature": False
                    }
                    create_res = go_api.create_menu_item(page, token, patch_group_id, create_payload, passkey=passkey)
                    if create_res and create_res.get("ok"):
                        logger.info(f"✅ New item '{final_name}' created successfully!")
                        results.append({
                            "item_id": "NEW_ITEM", "item_name": final_name,
                            "new_name": final_name, "new_price": final_price,
                            "new_photo": final_photo, "new_desc": final_desc,
                            "new_category": new_cat, "status": "SUCCESS", "error": None,
                        })
                    else:
                        logger.warning(f"⚠️ Failed to create new item '{final_name}': {create_res}")
                        results.append({
                            "item_id": "NEW_ITEM", "item_name": final_name,
                            "new_name": final_name, "new_price": final_price,
                            "status": "FAILED",
                            "error": (create_res.get('body') or create_res.get('error') or "Gagal membuat item baru.") if create_res else "Gagal membuat item baru.",
                        })
                    time.sleep(random.uniform(1.2, 2.5))
                    continue

                if not item_info:
                    results.append({
                        "item_id": item_id, "item_name": upd.get("item_name", item_id),
                        "new_name": new_name or None, "new_price": raw_price,
                        "status": "FAILED", "error": "Item ID tidak ditemukan di menu GoFood.",
                    })
                    continue

                try:
                    final_price = int(float(raw_price)) if (want_price and raw_price is not None) else int(float(orig_item.get('price') or 0))
                except (ValueError, TypeError):
                    final_price = int(float(orig_item.get('price') or 0))

                # ── Active Promo Protection for GoFood C5 Push ──
                item_name_key = (orig_item.get("name") or "").strip().lower()
                mpp_info = mpp_promos_map.get(str(item_id).strip()) or mpp_promos_map.get(item_name_key)
                promo_info = orig_item.get("promo_info") or orig_item.get("discount") or orig_item.get("campaign") or mpp_info
                original_p = float(orig_item.get("original_price") or orig_item.get("list_price") or 0)
                cur_p = float(orig_item.get("price") or 0)
                is_go_promo = bool(promo_info) or (original_p > cur_p > 0)

                # Klasifikasi promo: hanya lock harga jika diskon NOMINAL
                is_nominal_promo = False
                promo_desc = ""
                if is_go_promo:
                    if isinstance(promo_info, dict):
                        pct = promo_info.get("discount_percentage") or promo_info.get("percentage")
                        val = promo_info.get("discount_value") or promo_info.get("value") or promo_info.get("amount")
                        if pct and float(pct) > 0:
                            is_nominal_promo = False
                            promo_desc = f"Persentase ({int(float(pct))}%)"
                        elif val and float(val) > 0:
                            is_nominal_promo = True
                            promo_desc = f"Nominal (Rp {int(float(val)):,})"
                        else:
                            is_nominal_promo = False
                    else:
                        is_nominal_promo = False

                if is_go_promo and is_nominal_promo and want_price and raw_price is not None and int(float(raw_price)) != int(cur_p):
                    logger.info(f"🔒 [GoFood C5 Push] Item '{final_name}' ({item_id}) sedang promo nominal ({promo_desc}). Mempertahankan harga asli Rp{cur_p:,.0f} dan tetap mengupdate nama/foto/kategori.")
                    final_price = int(cur_p)
                    price_steps = [final_price]
                    promo_note = f" (Harga tidak diubah karena promo nominal aktif: {promo_desc})"
                elif is_go_promo and not is_nominal_promo and want_price:
                    logger.info(f"⚡ [GoFood C5 Push] Item '{final_name}' ({item_id}) sedang promo persentase ({promo_desc or 'Dynamic %'}). Push harga baru diizinkan.")
                    orig_price = float(orig_item.get('price') or 0)
                    price_steps = calculate_price_steps(orig_price, final_price) if want_price and orig_price > 0 else [final_price]
                    promo_note = ""
                else:
                    orig_price = float(orig_item.get('price') or 0)
                    price_steps = calculate_price_steps(orig_price, final_price) if want_price and orig_price > 0 else [final_price]
                    promo_note = ""

                res = None
                if len(price_steps) > 1:
                    logger.info(f"🔄 Step-push harga untuk '{final_name}' ({item_id}): Rp{orig_price:,.0f} -> Rp{final_price:,.0f} ({len(price_steps)} tahap: {price_steps})")

                for step_idx, step_price in enumerate(price_steps):
                    v2_payload["price"] = step_price
                    res = send_patch_request(v2_payload)

                    if res and res.get('status') == 429:
                        logger.warning("⚠️ GoFood API Rate Limited (HTTP 429). Cooldown 10s...")
                        time.sleep(10.0)

                    # Fallback 1: include variant_category_common_ids if present
                    if (not res or not res.get('ok')) and res.get('status') != 429:
                        time.sleep(0.6)
                        vars_ids = orig_item.get('variant_category_common_ids') or orig_item.get('variant_category_ids')
                        if vars_ids and isinstance(vars_ids, list) and len(vars_ids) > 0:
                            v2_payload_with_vars = dict(v2_payload)
                            v2_payload_with_vars["variant_category_common_ids"] = vars_ids
                            res_var = send_patch_request(v2_payload_with_vars, max_retries=1)
                            if res_var and res_var.get('ok'):
                                res = res_var

                    # Fallback 2: V1 PUT
                    if (not res or not res.get('ok')) and res.get('status') != 429:
                        status_code = res.get('status', '?') if res else '?'
                        body_err = (res.get('body') or '')[:300] if res else ''
                        logger.warning(f"GoFood V2 PATCH gagal (HTTP {status_code}): {body_err}. Fallback V1 PUT...")
                        time.sleep(0.6)
                        v1_payload = {
                            "name": final_name,
                            "price": step_price,
                            "active": orig_item.get('is_active', orig_item.get('active', True)),
                            "description": final_desc,
                            "image": final_photo,
                        }
                        v1_item_id = orig_item.get('id') or orig_item.get('common_id') or item_id
                        if v1_item_id:
                            v1_url = f'https://api.gojekapi.com/gofood/merchant/v1/restaurants/{rest_uuid}/menu_items/{v1_item_id}'
                            try:
                                cr_v1 = context.request.fetch(
                                    v1_url, method='PUT', headers=headers_direct, data=json.dumps(v1_payload)
                                )
                                res = {'ok': cr_v1.ok, 'status': cr_v1.status, 'body': cr_v1.text()}
                            except Exception as e:
                                res = {'ok': False, 'error': str(e)}

                    if res and res.get('ok'):
                        if len(price_steps) > 1:
                            logger.info(f"  ✅ Step {step_idx+1}/{len(price_steps)}: Rp{step_price:,} berhasil dipush.")
                            if step_idx < len(price_steps) - 1:
                                time.sleep(1.5)
                    else:
                        if len(price_steps) > 1:
                            logger.warning(f"  ❌ Step {step_idx+1}/{len(price_steps)}: Rp{step_price:,} gagal dipush.")
                        break

                if res and res.get('ok'):
                    res_entry = {
                        "item_id": item_id, "item_name": orig_item.get('name', item_id),
                        "new_name": final_name if want_name else None,
                        "new_price": final_price if want_price else None,
                        "new_photo": final_photo if want_photo else None,
                        "new_desc": final_desc if want_desc else None,
                        "new_category": new_cat if upd.get("category_id") in category_renames else None,
                        "status": "SUCCESS", "error": None,
                    }
                    results.append(res_entry)
                    if item_result_cb:
                        try:
                            item_result_cb(upd, "SUCCESS", None, applied=res_entry)
                        except Exception:
                            pass
                else:
                    err_msg = (res.get('body') or res.get('error') or "GoFood API error.") if res else "GoFood API error."
                    res_entry = {
                        "item_id": item_id, "item_name": orig_item.get('name', item_id),
                        "new_name": new_name or None, "new_price": raw_price,
                        "status": "FAILED",
                        "error": err_msg,
                    }
                    results.append(res_entry)
                    if item_result_cb:
                        try:
                            item_result_cb(upd, "FAILED", err_msg, applied=res_entry)
                        except Exception:
                            pass

                # Pacing + batch breather to respect GoFood rate limits
                time.sleep(random.uniform(1.2, 2.5))
                if (idx + 1) % 20 == 0 and (idx + 1) < total:
                    logger.info(f"☕ Batch pause item {idx+1}/{total}: istirahat 3s...")
                    time.sleep(3.0)
        finally:
            proc_killed = False
            try:
                if proc:
                    from src.core.browser_factory import kill_process_tree
                    kill_process_tree(proc)
                    proc_killed = True
            except Exception:
                pass
            if not proc_killed and proc:
                try:
                    proc.kill()
                    proc.wait(timeout=2)
                except Exception:
                    pass
            try:
                browser.close()
            except Exception:
                pass

    return results


def run_push_c5_job(job_id: uuid.UUID, selected_sids: list, updates_list: list):
    """Background task to push C5 menu changes (name + price + photos + categories) to real store (GoFood or GrabFood)."""
    from menu_core.database import SessionLocal
    db = SessionLocal()
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        db.close()
        return

    platform = (job.platform or "gofood").lower()
    lock = PLATFORM_LOCKS.get(platform)
    if lock:
        logger.info(f"🔒 C5 job {job_id} ({platform}) waiting for lock...")
        lock.acquire()
        logger.info(f"🔓 C5 job {job_id} ({platform}) acquired lock.")

    try:
        job.status = "RUNNING"
        job.started_at = datetime.utcnow()
        job.progress_pct = 10
        job.current_step = f"Menginisialisasi kredensial {platform.title()}..."
        db.commit()

        total_updates = len(updates_list)
        success_count = 0
        fail_count = 0

        logger.info(f"🚀 run_push_c5_job starting for job {job_id}, platform {platform}. Selected SIDs: {selected_sids}, total updates: {total_updates}")

        # Group updates by Store ID — each SID is a separate merchant/login.
        updates_by_sid = {}
        for upd in updates_list:
            sid = (upd.get("sid") or "").strip() or (selected_sids[0] if selected_sids else "")
            updates_by_sid.setdefault(sid, []).append(upd)

        processed = 0

        def record_trail(upd, status_str, err_msg, applied=None):
            change_str = ", ".join(upd.get("changes") or upd.get("change_types") or ["C5 Update"])
            new_val = []
            if applied and applied.get("new_name"):
                new_val.append(f"Nama Item: {applied['new_name']}")
            if applied and applied.get("new_price") is not None:
                new_val.append(f"Harga Baru: Rp {applied['new_price']}")
            if applied and applied.get("new_photo"):
                new_val.append(f"Foto Link: {applied['new_photo']}")
            if applied and applied.get("new_desc"):
                new_val.append(f"Deskripsi: {applied['new_desc']}")
            if applied and applied.get("new_category"):
                new_val.append(f"Nama Kategori: {applied['new_category']}")
            trail = AuditTrail(
                job_id=job.id,
                outlet_id=job.outlet_id or uuid.uuid4(),
                item_id=str(upd.get("item_id", "")),
                item_name=str(upd.get("item_name", "")),
                change_type=f"C5_PUSH_{platform.upper()}",
                field_changed=change_str,
                old_value=str(upd.get("item_name") or upd.get("current_fake_price") or ""),
                new_value=" | ".join(new_val) if new_val else ("Updated" if status_str == "SUCCESS" else ""),
                status=status_str,
                error_message=err_msg,
            )
            db.add(trail)
            db.commit()

        for sid, sid_updates in updates_by_sid.items():
            # Resolve the outlet + account credentials for this Store ID & platform.
            outlet = db.query(Outlet).filter(Outlet.store_id == sid, Outlet.platform == platform).first()
            if not outlet:
                outlet = db.query(Outlet).filter(Outlet.store_id == sid).first()
            if not outlet and sid:
                cands = [sid.replace("GM", "M"), sid.lstrip("G"), "G" + sid]
                for csid in cands:
                    outlet = db.query(Outlet).filter(Outlet.store_id == csid).first()
                    if outlet:
                        break
            account = db.query(Account).filter(Account.id == outlet.account_id).first() if outlet else None

            if not outlet or not account:
                logger.warning(f"⚠️ Outlet/akun tidak ditemukan untuk SID {sid} ({platform}). Menandai {len(sid_updates)} item gagal.")
                for upd in sid_updates:
                    processed += 1
                    fail_count += 1
                    record_trail(upd, "FAILED", f"Outlet atau akun {platform.title()} tidak ditemukan untuk Store ID {sid}.")
                continue

            job.current_step = f"Login {platform.title()} & memproses Store {sid} ({len(sid_updates)} item)..."
            db.commit()

            def progress_cb(idx, total, upd, _sid=sid):
                job.current_step = f"Store {_sid} ({idx+1}/{total}): {upd.get('item_name')}..."
                job.progress_pct = int(15 + (processed / max(1, total_updates)) * 80)
                db.commit()

            def item_cb(upd, status_str, err_msg=None, applied=None):
                if status_str == "SUCCESS":
                    nonlocal success_count
                    success_count += 1
                else:
                    nonlocal fail_count
                    fail_count += 1
                record_trail(upd, status_str, err_msg, applied=applied)

            try:
                if platform == "grab":
                    from grab.core.push_c5 import push_c5_grab_for_merchant
                    results = push_c5_grab_for_merchant(
                        username=account.username,
                        password=account.password,
                        store_id=outlet.store_id,
                        updates=sid_updates,
                        progress_cb=progress_cb,
                        item_result_cb=item_cb,
                    )
                elif platform == "shopee":
                    from shopee.core.push_c5 import push_c5_shopee_for_merchant
                    store_meta = {
                        "store_id": outlet.store_id,
                        "username": account.username,
                        "password": account.password,
                        "merchant_name": outlet.merchant_name or outlet.outlet_name,
                        "nama_resto_final": outlet.outlet_name,
                        "nama_outlet": outlet.outlet_name,
                    }
                    results = push_c5_shopee_for_merchant(
                        store_metadata=store_meta,
                        updates=sid_updates,
                        progress_cb=progress_cb,
                        item_result_cb=item_cb,
                    )
                else:
                    from concurrent.futures import ThreadPoolExecutor
                    with ThreadPoolExecutor(max_workers=1) as executor:
                        results = executor.submit(
                            _push_c5_gofood_for_merchant,
                            email=account.username,
                            password=account.password,
                            merchant_id=outlet.store_id,
                            updates=sid_updates,
                            progress_cb=progress_cb,
                            item_result_cb=item_cb,
                        ).result()
            except Exception as ex:
                logger.error(f"❌ Gagal push {platform.title()} untuk SID {sid}: {ex}")
                # Record trail only for remaining items not already recorded via item_result_cb
                processed_in_results = {str(r.get("item_id")) for r in (results or [])}
                for upd in sid_updates:
                    if str(upd.get("item_id")) not in processed_in_results:
                        processed += 1
                        fail_count += 1
                        record_trail(upd, "FAILED", str(ex))
                continue

            job.progress_pct = int(15 + (processed / max(1, total_updates)) * 80)
            db.commit()

        job.status = "SUCCESS" if success_count > 0 else "FAILED"
        job.completed_at = datetime.utcnow()
        job.progress_pct = 100
        if fail_count == 0:
            job.current_step = f"Selesai di-push ke {platform.title()}!"
        elif success_count > 0:
            job.current_step = f"Selesai dengan sebagian gagal: {success_count} sukses, {fail_count} gagal."
        else:
            job.current_step = f"Gagal: tidak ada item yang berhasil di-push ({fail_count} gagal)."
        job.result_metadata = {
            "success_count": success_count,
            "fail_count": fail_count,
            "selected_sids": selected_sids,
            "total_updates": total_updates,
            "platform": platform,
        }
        db.commit()
    except Exception as ex:
        logger.error(f"Error in run_push_c5_job: {ex}")
        import traceback
        traceback.print_exc()
        job.status = "FAILED"
        job.completed_at = datetime.utcnow()
        job.current_step = f"Gagal: {str(ex)}"
        db.commit()
    finally:
        db.close()
        if lock:
            lock.release()
            logger.info(f"🔓 C5 job {job_id} (gofood) released lock.")



def _download_gdrive_or_gsheet_bytes(url: str) -> tuple[bytes, str]:
    """Mengunduh file C5 XLSX dari link Google Drive atau Google Sheets."""
    import re
    import urllib.request
    import urllib.error

    clean_url = url.strip()
    export_url = None
    filename = "gdrive_c5.xlsx"

    # 1. Google Sheets (/spreadsheets/d/<id> atau /spreadsheets/u/<num>/d/<id>)
    m_sheet = re.search(r'/spreadsheets(?:/u/\d+)?/d/([a-zA-Z0-9-_]+)', clean_url)
    if m_sheet:
        sheet_id = m_sheet.group(1)
        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        filename = f"gsheet_{sheet_id[:8]}.xlsx"
    else:
        # 2. Google Drive File (/file/d/<id> atau /file/u/<num>/d/<id>)
        m_file = re.search(r'/file(?:/u/\d+)?/d/([a-zA-Z0-9-_]+)', clean_url)
        if m_file:
            file_id = m_file.group(1)
            export_url = f"https://drive.google.com/uc?export=download&id={file_id}"
            filename = f"gdrive_{file_id[:8]}.xlsx"
        else:
            # 3. Parameter id= atau /d/<id>
            m_id = re.search(r'(?:[?&]id=|/d/)([a-zA-Z0-9-_]{20,})', clean_url)
            if m_id:
                file_id = m_id.group(1)
                export_url = f"https://drive.google.com/uc?export=download&id={file_id}"
                filename = f"gdrive_{file_id[:8]}.xlsx"

    if not export_url:
        raise HTTPException(
            status_code=400,
            detail="Format Link Google Drive / Google Sheets tidak valid. Pastikan link berisi ID dokumen (contoh: https://docs.google.com/spreadsheets/d/...)"
        )

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    try:
        req = urllib.request.Request(export_url, headers=headers)
        with urllib.request.urlopen(req, timeout=30) as resp:
            # Cegah DoS / download raksasa (max 20 MB)
            max_bytes = 20 * 1024 * 1024
            data = resp.read(max_bytes + 1)
            if len(data) > max_bytes:
                raise HTTPException(status_code=400, detail="Ukuran file Google Drive melebihi batas maksimal 20 MB.")
            if not data:
                raise HTTPException(status_code=400, detail="Data file kosong dari Google Drive.")
            
            # Cek jika respon HTML login/error google
            if data[:100].strip().startswith(b"<!DOCTYPE") or b"<html" in data[:200].lower():
                raise HTTPException(
                    status_code=403,
                    detail="Akses Google Drive ditolak atau memerlukan izin akses. Pastikan tautan disetel 'Siapa saja yang memiliki link / Anyone with link (Viewer)'."
                )
            return data, filename
    except urllib.error.HTTPError as e:
        if e.code in (401, 403):
            raise HTTPException(
                status_code=403,
                detail="Akses Google Drive ditolak (403). Pastikan share setting disetel 'Anyone with the link can view'."
            )
        elif e.code == 404:
            raise HTTPException(
                status_code=404,
                detail="Dokumen Google Drive tidak ditemukan (404). Periksa kembali link Anda."
            )
        else:
            raise HTTPException(status_code=400, detail=f"Gagal mengunduh file dari Google Drive (HTTP {e.code}): {e.reason}")
    except urllib.error.URLError as e:
        raise HTTPException(status_code=400, detail=f"Gagal menghubungi server Google: {e.reason}")
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Terjadi kesalahan saat mengunduh Google Drive: {str(e)}")


@app.post("/api/jobs/parse-c5")
async def parse_c5_endpoint(
    file: Optional[UploadFile] = File(None),
    drive_url: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Parses an uploaded C5 Excel file (.xlsx) or Google Drive link and detects Store IDs (SIDs) plus menu changes
    for GoFood/Grab by comparing each C5 row against the last PULL baseline.
    """
    import io
    import openpyxl
    import re

    filename = ""
    contents = b""

    if file and file.filename:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File harus berformat Excel (.xlsx / .xls)")
        filename = file.filename
        contents = await file.read()
    elif drive_url and drive_url.strip():
        contents, filename = _download_gdrive_or_gsheet_bytes(drive_url.strip())
    else:
        raise HTTPException(status_code=400, detail="Harap unggah file Excel C5 (.xlsx) atau masukkan Link Google Drive / Sheets.")

    try:
        wb_raw = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=False)
        sheet_raw = wb_raw['Item'] if 'Item' in wb_raw.sheetnames else wb_raw.active
        raw_row1 = [sheet_raw.cell(row=1, column=c).value for c in range(1, sheet_raw.max_column + 1)]
    except Exception:
        raw_row1 = []

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca file Excel C5: {str(e)}")

    if 'Item' not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Sheet 'Item' tidak ditemukan di file Excel C5.")

    sheet = wb['Item']
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) <= 1:
        return {
            "success": True,
            "filename": filename,
            "stores": [],
            "items": [],
            "summary": {"total_stores": 0, "total_items": 0, "total_changes": 0}
        }

    headers = []
    data_row1 = rows[0]
    max_cols = max(len(data_row1), len(raw_row1))
    for c_idx in range(max_cols):
        d_val = data_row1[c_idx] if c_idx < len(data_row1) else None
        r_val = raw_row1[c_idx] if c_idx < len(raw_row1) else None
        if d_val is not None and str(d_val).strip() != "":
            headers.append(str(d_val).strip())
        elif r_val is not None:
            txt = r_val.text if hasattr(r_val, 'text') else str(r_val)
            m = re.search(r'\"([^\"]+)\"', txt)
            if m:
                headers.append(m.group(1).strip())
            else:
                headers.append(str(txt).strip())
        else:
            headers.append("")

    header_map = {h: i for i, h in enumerate(headers) if h}

    def get_val(row, col_name, default=""):
        if col_name in header_map:
            idx = header_map[col_name]
            if idx < len(row) and row[idx] is not None:
                val = str(row[idx]).strip()
                if val and val not in ("#N/A", "nan", "None"):
                    return val
        return default

    # Flexible column resolution for standard C5 headers
    def resolve_col(candidates):
        for c in candidates:
            if c in header_map:
                return c
        return ""

    sid_col = resolve_col(['SID', 'Store ID'])
    outlet_col = resolve_col(['Outlet Name', 'Outlet'])
    cat_id_col = resolve_col(['Category ID'])
    cat_col = resolve_col(['Category', 'Nama Kategori', 'Kategori'])
    item_id_col = resolve_col(['Item ID'])
    item_col = resolve_col(['Item', 'Item Name', 'Nama Item'])
    photo_col = resolve_col(['Photo Link', 'Gambar', 'Link Foto', 'Photo'])
    desc_col = resolve_col(['Description', 'Deskripsi'])
    avail_col = resolve_col(['Availability', 'Ketersediaan'])
    vis_col = resolve_col(['Visibility'])
    item_name_imp_col = resolve_col(['Item Name Improvement'])
    new_fake_col = resolve_col(['New Fake Price (Rp)', 'New Fake Price'])
    curr_fake_col = resolve_col(['Current Fake Price (Rp)', 'Current Fake Price', 'Harga Fake'])
    notes_col = resolve_col(['Notes', 'Catatan'])

    # ── Category ID Consistency Pass ──
    # All rows sharing the same (SID, Category ID) in C5 MUST have the exact same Category Name.
    cat_id_name_map = {}
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        s_id = get_val(row, sid_col) or "STORE-DEFAULT"
        c_id = get_val(row, cat_id_col)
        c_name = get_val(row, cat_col)
        if c_id and c_name:
            key = (s_id, c_id)
            if key not in cat_id_name_map:
                cat_id_name_map[key] = set()
            cat_id_name_map[key].add(c_name)

    inconsistent_cat_ids = {}
    validation_error_messages = []

    for (s_id, c_id), names in cat_id_name_map.items():
        if len(names) > 1:
            names_str = ", ".join([f"'{n}'" for n in sorted(names)])
            msg = f"Nama kategori tidak konsisten untuk Category ID '{c_id}' (Store ID: {s_id}). Ditemukan {len(names)} nama berbeda: {names_str}. Semua item dengan Category ID yang sama harus memiliki nama kategori yang sama."
            inconsistent_cat_ids[(s_id, c_id)] = {
                "names": list(names),
                "error": msg
            }
            validation_error_messages.append(msg)

    stores_dict = {}
    parsed_items = []

    total_changes = 0
    name_changes_count = 0
    price_changes_count = 0
    price_warning_count = 0
    category_changes_count = 0
    photo_changes_count = 0
    description_changes_count = 0
    other_changes_count = 0
    validation_errors_count = 0

    new_items_count = 0
    new_categories_count = 0

    import re

    def parse_price(val_str):
        if val_str is None or str(val_str).strip() == "":
            return None
        try:
            cleaned = re.sub(r'[^\d.]', '', str(val_str))
            if cleaned:
                return float(cleaned)
        except Exception:
            pass
        return None

    def norm_str(s):
        return re.sub(r'\s+', ' ', str(s or '').strip().lower())

    # ── Baseline PULL cache loader (Multi-platform: GoFood / Grab / Shopee) ──
    _baseline_cache = {}

    def _find_baseline_file(sid):
        sid_str = str(sid or "").strip()
        candidates = [
            sid_str,
            sid_str.replace("GM", "M"),
            sid_str.replace("GM", "G"),
            sid_str.lstrip("G"),
            sid_str.lstrip("M"),
            f"G{sid_str.lstrip('G')}",
            f"M{sid_str.lstrip('M')}",
        ]
        unique_cands = []
        for c in candidates:
            if c and c not in unique_cands:
                unique_cands.append(c)

        search_dirs = [
            Path(__file__).parent / "Gofood" / "API",
            Path(__file__).parent / "grab" / "API",
            Path(__file__).parent / "shopee" / "API",
        ]

        for s_dir in search_dirs:
            if not s_dir.exists():
                continue
            for cand in unique_cands:
                p = s_dir / f"menu-response-{cand}.json"
                if p.exists():
                    return ("json", p)

        # Fallback: scan existing exported C5 Excel workbooks in data/exports/
        exports_dir = Path(__file__).parent / "data" / "exports"
        if exports_dir.exists():
            for p in sorted(exports_dir.rglob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
                try:
                    wb_exp = openpyxl.load_workbook(p, data_only=True, read_only=True)
                    if "Item" not in wb_exp.sheetnames:
                        continue
                    ws_exp = wb_exp["Item"]
                    exp_rows = list(ws_exp.iter_rows(values_only=True))
                    if not exp_rows:
                        continue
                    exp_headers = [str(c).strip() if c is not None else "" for c in exp_rows[0]]
                    exp_hmap = {h: i for i, h in enumerate(exp_headers) if h}
                    exp_sid_idx = exp_hmap.get("SID") or exp_hmap.get("Store ID")
                    if exp_sid_idx is None:
                        continue

                    for r in exp_rows[1:]:
                        if r and exp_sid_idx < len(r) and str(r[exp_sid_idx] or "").strip().lower() == sid_str.lower():
                            return ("xlsx", p)
                except Exception:
                    pass

        return (None, None)
    def _parse_json_baseline(path, sid, by_id, by_name, categories):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if "items" in data and isinstance(data.get("items"), list):
                for it in data["items"]:
                    it_sid = str(it.get("Store ID") or "").strip()
                    if it_sid and sid and it_sid.lower() != str(sid).strip().lower():
                        continue
                    cat_name = it.get("Category") or it.get("Nama Kategori") or ""
                    if cat_name:
                        categories.add(norm_str(cat_name))
                    price_val = parse_price(it.get("Current Fake Price (Rp)") or it.get("Harga Fake") or it.get("Current Real Price (Rp)") or it.get("Harga Real") or it.get("price"))
                    rec = {
                        "name": it.get("Item") or it.get("Nama Item") or it.get("name") or "",
                        "price": price_val,
                        "image": it.get("Photo Link") or it.get("Link Foto") or it.get("image") or "",
                        "category": cat_name,
                        "description": it.get("Description") or it.get("Deskripsi") or it.get("description") or "",
                    }
                    iid = str(it.get("Item ID") or it.get("common_id") or it.get("id") or "").strip()
                    if iid:
                        by_id[iid] = rec
                    nn = norm_str(rec["name"])
                    if nn and nn not in by_name:
                        by_name[nn] = rec
            elif "catalogs" in (data.get("data") or data if isinstance(data.get("data") or data, dict) else {}):
                catalogs_list = data.get("data", {}).get("catalogs") or data.get("catalogs") or []
                for cat in catalogs_list:
                    cat_name = cat.get("name") or ""
                    if cat_name:
                        categories.add(norm_str(cat_name))
                    for dish in cat.get("dishes", []):
                        price_raw = dish.get("price", "0")
                        price_val = float(price_raw) / 100000.0 if float(price_raw) > 1000 else float(price_raw)
                        rec = {
                            "name": dish.get("name") or "",
                            "price": price_val,
                            "image": dish.get("picture") or "",
                            "category": cat_name,
                            "description": dish.get("description") or "",
                        }
                        iid = str(dish.get("id") or "").strip()
                        if iid:
                            by_id[iid] = rec
                        nn = norm_str(rec["name"])
                        if nn and nn not in by_name:
                            by_name[nn] = rec
            else:
                for menu in (data.get("menus") or data.get("categories") or []):
                    cat_name = menu.get("name") or menu.get("category_name") or ""
                    if cat_name:
                        categories.add(norm_str(cat_name))
                    for it in (menu.get("menu_items") or menu.get("items") or []):
                        rec = {
                            "name": it.get("name") or it.get("item_name") or "",
                            "price": parse_price(it.get("price")),
                            "image": it.get("image") or it.get("photo_url") or "",
                            "category": cat_name,
                            "description": it.get("description") or "",
                        }
                        iid = str(it.get("common_id") or it.get("id") or it.get("item_id") or "").strip()
                        if iid:
                            by_id[iid] = rec
                        nn = norm_str(rec["name"])
                        if nn and nn not in by_name:
                            by_name[nn] = rec
        except Exception as ex:
            logger.warning(f"Gagal membaca baseline PULL JSON untuk SID {sid}: {ex}")
    def _parse_xlsx_baseline(path, sid, by_id, by_name, categories):
        try:
            wb_exp = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if "Item" in wb_exp.sheetnames:
                ws_exp = wb_exp["Item"]
                exp_rows = list(ws_exp.iter_rows(values_only=True))
                if exp_rows:
                    exp_headers = [str(c).strip() if c is not None else "" for c in exp_rows[0]]
                    exp_hmap = {h: i for i, h in enumerate(exp_headers) if h}
                    exp_sid_idx = exp_hmap.get("SID") or exp_hmap.get("Store ID")
                    cat_idx = exp_hmap.get("Category") or exp_hmap.get("Nama Kategori")
                    iid_idx = exp_hmap.get("Item ID")
                    name_idx = exp_hmap.get("Item") or exp_hmap.get("Item Name")
                    photo_idx = exp_hmap.get("Photo Link") or exp_hmap.get("Gambar")
                    desc_idx = exp_hmap.get("Description") or exp_hmap.get("Deskripsi")
                    price_idx = exp_hmap.get("Current Fake Price (Rp)") or exp_hmap.get("Current Real Price (Rp)") or exp_hmap.get("Current Fake Price")

                    for r in exp_rows[1:]:
                        if not r or exp_sid_idx is None or exp_sid_idx >= len(r):
                            continue
                        if str(r[exp_sid_idx] or "").strip().lower() != str(sid).strip().lower():
                            continue
                        cname = str(r[cat_idx]).strip() if cat_idx is not None and cat_idx < len(r) and r[cat_idx] is not None else ""
                        if cname:
                            categories.add(norm_str(cname))
                        iname = str(r[name_idx]).strip() if name_idx is not None and name_idx < len(r) and r[name_idx] is not None else ""
                        iid = str(r[iid_idx]).strip() if iid_idx is not None and iid_idx < len(r) and r[iid_idx] is not None else ""
                        photo = str(r[photo_idx]).strip() if photo_idx is not None and photo_idx < len(r) and r[photo_idx] is not None else ""
                        desc = str(r[desc_idx]).strip() if desc_idx is not None and desc_idx < len(r) and r[desc_idx] is not None else ""
                        price = parse_price(r[price_idx]) if price_idx is not None and price_idx < len(r) else None

                        rec = {
                            "name": iname,
                            "price": price,
                            "image": photo,
                            "category": cname,
                            "description": desc,
                        }
                        if iid:
                            by_id[iid] = rec
                        nn = norm_str(iname)
                        if nn and nn not in by_name:
                            by_name[nn] = rec
        except Exception as ex:
            logger.warning(f"Gagal membaca baseline PULL XLSX untuk SID {sid}: {ex}")
    def load_baseline(sid):
        """Loads the last-pulled menu snapshot for a SID and indexes items by id/name and categories."""
        if sid in _baseline_cache:
            return _baseline_cache[sid]

        by_id, by_name, categories = {}, {}, set()
        ftype, path = _find_baseline_file(sid)

        if path and path.exists():
            if ftype == "json":
                _parse_json_baseline(path, sid, by_id, by_name, categories)
            elif ftype == "xlsx":
                _parse_xlsx_baseline(path, sid, by_id, by_name, categories)

        result = {"by_id": by_id, "by_name": by_name, "categories": categories, "found": bool(by_id or by_name)}
        _baseline_cache[sid] = result
        return result

    c5_parsed_ids_by_sid = {}
    c5_parsed_names_by_sid = {}

    for r_idx, row in enumerate(rows[1:], start=2):
        if not row or all(v is None for v in row):
            continue

        sid = get_val(row, sid_col) or "STORE-DEFAULT"
        outlet_name = get_val(row, outlet_col) or "Outlet Utama"
        cat_id = get_val(row, cat_id_col)
        cat_name = get_val(row, cat_col)
        item_id = get_val(row, item_id_col)
        item_name = get_val(row, item_col)
        photo_link_raw = get_val(row, photo_col)
        design_imp = get_val(row, resolve_col(['Design Improvement', 'Photo Improvement'])) if resolve_col(['Design Improvement', 'Photo Improvement']) else ""
        
        # Check if Design Improvement contains a new photo link / Google Drive URL
        if design_imp and (design_imp.startswith("http://") or design_imp.startswith("https://") or "drive.google.com" in design_imp):
            photo_link = design_imp
        else:
            photo_link = photo_link_raw

        description = get_val(row, desc_col) if desc_col else ""
        availability = get_val(row, avail_col) if avail_col else ""
        visibility = get_val(row, vis_col) if vis_col else ""
        notes = get_val(row, notes_col) if notes_col else ""

        new_fake_raw = get_val(row, new_fake_col) if new_fake_col else ""
        curr_fake_raw = get_val(row, curr_fake_col) if curr_fake_col else ""
        item_name_imp = get_val(row, item_name_imp_col) if item_name_imp_col else ""

        display_name = item_name_imp.strip() if item_name_imp else item_name
        if not item_id and not display_name:
            continue

        if sid not in c5_parsed_ids_by_sid:
            c5_parsed_ids_by_sid[sid] = set()
            c5_parsed_names_by_sid[sid] = set()

        if item_id:
            c5_parsed_ids_by_sid[sid].add(str(item_id).strip())
        if display_name:
            c5_parsed_names_by_sid[sid].add(norm_str(display_name))
        if item_name:
            c5_parsed_names_by_sid[sid].add(norm_str(item_name))

        # Match this C5 row against the baseline PULL cache (by Item ID, else by name).
        baseline = load_baseline(sid)
        base_rec = None
        if item_id and item_id in baseline["by_id"]:
            base_rec = baseline["by_id"][item_id]
        elif display_name and norm_str(display_name) in baseline["by_name"]:
            base_rec = baseline["by_name"][norm_str(display_name)]
        elif item_name and norm_str(item_name) in baseline["by_name"]:
            base_rec = baseline["by_name"][norm_str(item_name)]

        base_name = base_rec["name"] if base_rec else item_name
        base_price = base_rec["price"] if base_rec else None
        base_cat = base_rec["category"] if base_rec else ""
        base_img = base_rec["image"] if base_rec else ""
        base_desc = base_rec["description"] if base_rec else ""

        new_fake_price = parse_price(new_fake_raw) or parse_price(curr_fake_raw)

        # ── Detect New Item (tambah_item) & New Category (new_categories) ──
        is_new_item = False
        is_new_category = False

        if not item_id or (base_rec is None and norm_str(display_name) not in baseline["by_name"]):
            is_new_item = True

        if cat_name and (not cat_id or norm_str(cat_name) not in baseline["categories"]):
            is_new_category = True

        if sid not in stores_dict:
            stores_dict[sid] = {
                "sid": sid,
                "name": outlet_name,
                "item_count": 0,
                "changed_count": 0,
                "baseline_found": baseline["found"],
            }
        stores_dict[sid]["item_count"] += 1

        # ── Comprehensive All-Column Change Detection & Validation ──
        diff_details = []
        is_valid = True
        validation_error = None

        # Check Category ID consistency error for this row
        if (sid, cat_id) in inconsistent_cat_ids:
            is_valid = False
            validation_errors_count += 1
            validation_error = inconsistent_cat_ids[(sid, cat_id)]["error"]
            diff_details.append({
                "column": "Category ID",
                "old_val": "Konflik Nama Kategori",
                "new_val": f"Gunakan nama yang sama untuk Category ID '{cat_id}'"
            })

        # 1. New Item & New Category Indications
        name_changed = False
        if is_new_item:
            diff_details.append({"column": "Item Status", "old_val": "(Item Baru)", "new_val": f"Tambah Item Baru: {display_name}"})
        elif display_name and norm_str(display_name) != norm_str(base_name):
            name_changed = True
            diff_details.append({"column": "Item Name", "old_val": base_name, "new_val": display_name})
        else:
            name_changed = False

        if is_new_category:
            diff_details.append({"column": "Category Status", "old_val": "(Kategori Baru)", "new_val": f"Kategori Baru: {cat_name}"})
            category_changed = True
        else:
            category_changed = bool(cat_name and base_cat and norm_str(cat_name) != norm_str(base_cat))
            if category_changed:
                diff_details.append({"column": "Category", "old_val": base_cat, "new_val": cat_name})

        # 3. Photo Link Change
        photo_changed = bool(photo_link and photo_link != base_img and not (not base_img and not photo_link))
        if photo_changed:
            diff_details.append({"column": "Photo Link", "old_val": base_img or "(Kosong)", "new_val": photo_link})

        # 4. Description Change
        description_changed = bool(description and norm_str(description) != norm_str(base_desc))
        if description_changed:
            diff_details.append({"column": "Description", "old_val": base_desc or "(Kosong)", "new_val": description})

        # 5. Price Change
        price_changed = False
        price_warning = False
        price_diff_percent = 0.0
        if new_fake_price is not None:
            if base_price is not None and float(base_price) > 0:
                price_changed = float(new_fake_price) != float(base_price)
                price_diff_percent = (abs(float(new_fake_price) - float(base_price)) / float(base_price)) * 100.0
                if price_changed and price_diff_percent > 15.0:
                    price_warning = True
            else:
                price_changed = True

        if price_changed:
            if price_warning:
                diff_details.append({
                    "column": "Price",
                    "old_val": f"Rp {float(base_price):,.0f}",
                    "new_val": f"Rp {float(new_fake_price):,.0f} (⚠️ Perubahan {price_diff_percent:.1f}% >15% - Push Bertahap)"
                })
            else:
                diff_details.append({"column": "Price", "old_val": base_price, "new_val": new_fake_price})

        # 6. Check all other columns in row for any explicit edits
        other_changed = False
        for col_k, col_idx in header_map.items():
            if col_k in (sid_col, outlet_col, cat_id_col, cat_col, item_id_col, item_col, photo_col, desc_col, item_name_imp_col, new_fake_col, curr_fake_col):
                continue
            if col_idx < len(row) and row[col_idx] is not None:
                v_str = str(row[col_idx]).strip()
                if v_str and v_str not in ("#N/A", "nan", "None"):
                    if col_k in ('Design Improvement', 'Keyword', 'Notes', 'Offline Price (Rp)', 'Offline Adjustment (Rp)'):
                        other_changed = True
                        diff_details.append({"column": col_k, "old_val": "", "new_val": v_str})

        is_changed = is_new_item or is_new_category or name_changed or category_changed or photo_changed or description_changed or price_changed or other_changed or (not is_valid)

        change_types = []
        if not is_valid:
            change_types.append("INVALID_CATEGORY_CONSISTENCY")
        if is_new_item:
            change_types.append("NEW_ITEM")
            new_items_count += 1
        if is_new_category:
            change_types.append("NEW_CATEGORY")
            new_categories_count += 1
        if name_changed:
            change_types.append("NAME_CHANGE")
            name_changes_count += 1
        if category_changed:
            change_types.append("CATEGORY_CHANGE")
            category_changes_count += 1
        if photo_changed:
            change_types.append("PHOTO_CHANGE")
            photo_changes_count += 1
        if description_changed:
            change_types.append("DESCRIPTION_CHANGE")
            description_changes_count += 1
        if price_changed:
            change_types.append("PRICE_CHANGE")
            price_changes_count += 1
            if price_warning:
                change_types.append("PRICE_WARNING_STEP_PUSH")
                price_warning_count += 1
        if other_changed:
            change_types.append("OTHER_CHANGE")
            other_changes_count += 1

        if is_changed:
            stores_dict[sid]["changed_count"] += 1
            total_changes += 1

        parsed_items.append({
            "row_number": r_idx,
            "sid": sid,
            "outlet_name": outlet_name,
            "category_id": cat_id,
            "category": cat_name,
            "item_id": item_id,
            "item_name": item_name,
            "item_name_new": display_name if is_changed and display_name else item_name_imp,
            "photo_link": photo_link,
            "description": description,
            "availability": availability,
            "visibility": visibility,
            "notes": notes,
            "baseline_name": base_name,
            "baseline_price": base_price,
            "baseline_category": base_cat,
            "baseline_photo": base_img,
            "baseline_description": base_desc,
            "current_fake_price": base_price,
            "new_fake_price": new_fake_price,
            "baseline_found": base_rec is not None,
            "is_valid": is_valid,
            "validation_error": validation_error,
            "is_changed": is_changed,
            "is_new_item": is_new_item,
            "is_new_category": is_new_category,
            "price_warning": price_warning,
            "price_diff_percent": round(price_diff_percent, 2),
            "change_types": change_types,
            "diff_details": diff_details,
            "changes": {
                "name_changed": name_changed,
                "category_changed": category_changed,
                "photo_changed": photo_changed,
                "description_changed": description_changed,
                "price_changed": price_changed,
                "price_warning": price_warning,
                "price_diff_percent": round(price_diff_percent, 2),
                "other_changed": other_changed,
                "is_new_item": is_new_item,
                "is_new_category": is_new_category,
                "invalid": not is_valid,
            }
        })

    # ── Detect Deleted Items (Items in baseline PULL missing from C5 spreadsheet) ──
    deleted_items_count = 0
    for sid, sinfo in stores_dict.items():
        baseline = load_baseline(sid)
        parsed_ids = c5_parsed_ids_by_sid.get(sid, set())
        parsed_names = c5_parsed_names_by_sid.get(sid, set())

        for base_iid, base_rec in baseline["by_id"].items():
            norm_bname = norm_str(base_rec.get("name"))
            if base_iid not in parsed_ids and norm_bname not in parsed_names:
                deleted_items_count += 1
                total_changes += 1
                sinfo["changed_count"] += 1
                parsed_items.append({
                    "row_number": 0,
                    "sid": sid,
                    "outlet_name": sinfo["name"],
                    "category_id": None,
                    "category": base_rec.get("category"),
                    "item_id": base_iid,
                    "item_name": base_rec.get("name"),
                    "item_name_new": None,
                    "photo_link": base_rec.get("image"),
                    "description": base_rec.get("description"),
                    "availability": "",
                    "visibility": "",
                    "notes": "Item tidak ada di C5 (Hapus Item)",
                    "baseline_name": base_rec.get("name"),
                    "baseline_price": base_rec.get("price"),
                    "baseline_category": base_rec.get("category"),
                    "baseline_photo": base_rec.get("image"),
                    "baseline_description": base_rec.get("description"),
                    "current_fake_price": base_rec.get("price"),
                    "new_fake_price": None,
                    "baseline_found": True,
                    "is_valid": True,
                    "validation_error": None,
                    "is_changed": True,
                    "is_new_item": False,
                    "is_new_category": False,
                    "is_deleted_item": True,
                    "change_types": ["DELETE_ITEM"],
                    "diff_details": [{
                        "column": "Item Status",
                        "old_val": base_rec.get("name"),
                        "new_val": "(Hapus Item)"
                    }],
                    "changes": {
                        "name_changed": False,
                        "category_changed": False,
                        "photo_changed": False,
                        "description_changed": False,
                        "price_changed": False,
                        "other_changed": False,
                        "is_new_item": False,
                        "is_new_category": False,
                        "is_deleted_item": True,
                        "invalid": False,
                    }
                })

    return {
        "success": True,
        "filename": filename,
        "stores": list(stores_dict.values()),
        "items": parsed_items,
        "summary": {
            "total_stores": len(stores_dict),
            "total_items": len(parsed_items),
            "total_changes": total_changes,
            "name_changes": name_changes_count,
            "category_changes": category_changes_count,
            "photo_changes": photo_changes_count,
            "description_changes": description_changes_count,
            "price_changes": price_changes_count,
            "price_warning_count": price_warning_count,
            "other_changes": other_changes_count,
            "new_items_count": new_items_count,
            "new_categories_count": new_categories_count,
            "deleted_items_count": deleted_items_count,
            "validation_errors_count": validation_errors_count,
            "has_validation_errors": len(validation_error_messages) > 0,
            "validation_error_messages": validation_error_messages,
        }
    }


@app.post("/api/jobs/push-c5", response_model=JobResponse, status_code=status.HTTP_202_ACCEPTED)
def trigger_push_c5_job(request: C5PushRequest, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """Triggers background job to push C5 menu updates (GoFood or GrabFood) for selected Store IDs (SIDs)."""
    target_platform = (request.platform or "gofood").lower()
    outlet = None
    if request.selected_sids:
        target_sid = request.selected_sids[0]
        outlet = db.query(Outlet).filter(Outlet.store_id == target_sid, Outlet.platform == target_platform).first()
        if not outlet:
            outlet = db.query(Outlet).filter(Outlet.store_id == target_sid).first()

    if not outlet:
        outlet = db.query(Outlet).join(Account).filter(Account.platform == target_platform).first()

    outlet_id = outlet.id if outlet else uuid.uuid4()
    updates_payload = [item.dict() for item in request.updates]

    new_job = Job(
        outlet_id=outlet_id,
        job_type="PUSH_UPDATE",
        platform=target_platform,
        status="PENDING",
        progress_pct=0,
        current_step=f"Mengantrekan C5 push {target_platform.title()}...",
        payload={
            "selected_sids": request.selected_sids,
            "updates_count": len(updates_payload),
            "platform": target_platform
        }
    )
    db.add(new_job)
    db.commit()
    db.refresh(new_job)

    background_tasks.add_task(run_push_c5_job, new_job.id, request.selected_sids, updates_payload)
    return new_job



# ─── AUDIT TRAILS ENDPOINTS ───────────────────────────────────────────────────

@app.get("/api/audit-trails", response_model=List[AuditTrailResponse])
def get_audit_trails(db: Session = Depends(get_db)):
    return db.query(AuditTrail).order_by(AuditTrail.created_at.desc()).limit(100).all()

@app.get("/api/outlets/{outlet_id}/items/{item_id}/pricing-quota")
def get_pricing_quota(outlet_id: uuid.UUID, item_id: str, db: Session = Depends(get_db)):
    """Retrieve remaining pricing quota for a specific item in a specific outlet."""
    outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
        
    account = db.query(Account).filter(Account.id == outlet.account_id).first()
    platform = (account.platform or "").lower() if account else "grab"

    now = datetime.utcnow()
    one_day_ago = now - timedelta(days=1)
    thirty_days_ago = now - timedelta(days=30)
    
    # Query successful price updates in the last 24 hours
    daily_count = db.query(AuditTrail).filter(
        AuditTrail.outlet_id == outlet_id,
        AuditTrail.item_id == item_id,
        AuditTrail.field_changed.ilike("price"),
        AuditTrail.status.ilike("SUCCESS"),
        AuditTrail.created_at >= one_day_ago
    ).count()
    
    # Query successful price updates in the last 30 days
    monthly_count = db.query(AuditTrail).filter(
        AuditTrail.outlet_id == outlet_id,
        AuditTrail.item_id == item_id,
        AuditTrail.field_changed.ilike("price"),
        AuditTrail.status.ilike("SUCCESS"),
        AuditTrail.created_at >= thirty_days_ago
    ).count()
    
    # Dynamic rules based on platform
    if platform == "shopee":
        daily_limit = 1
        monthly_limit = 99999  # Practically unlimited monthly quota
        max_increase_pct = 25.0
    elif platform == "grab":
        daily_limit = 10
        monthly_limit = 15
        max_increase_pct = 15.0
    else:  # GoFood or fallback
        daily_limit = 99999
        monthly_limit = 99999
        max_increase_pct = 99999.0
    
    daily_remaining = max(0, daily_limit - daily_count)
    monthly_remaining = max(0, monthly_limit - monthly_count)
    
    return {
        "outlet_id": outlet_id,
        "item_id": item_id,
        "platform": platform,
        "daily_limit": daily_limit,
        "daily_count": daily_count,
        "daily_remaining": daily_remaining,
        "monthly_limit": monthly_limit,
        "monthly_count": monthly_count,
        "monthly_remaining": monthly_remaining,
        "max_increase_pct": max_increase_pct
    }

@app.get("/api/outlets/{outlet_id}/menu-cache-status")
def get_menu_cache_status(outlet_id: uuid.UUID, db: Session = Depends(get_db)):
    """Check whether a valid menu pull cache exists for an outlet within 24 hours."""
    outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")

    job = db.query(Job).filter(
        Job.outlet_id == outlet_id,
        Job.job_type == "PULL",
        Job.status.in_(["SUCCESS", "PARTIAL_SUCCESS"])
    ).order_by(Job.completed_at.desc()).first()

    last_pulled_at = None
    excel_exists = False
    
    if job:
        last_pulled_at = job.completed_at or job.created_at
        excel_path = job.result_metadata.get("excel_path") if job.result_metadata else None
        if excel_path and os.path.exists(excel_path):
            excel_exists = True

    if not excel_exists:
        # Fallback check file mtime on exports folder
        import re
        raw_outlet = outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown'
        clean_outlet = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
        clean_outlet = re.sub(r'\s+', ' ', clean_outlet).lower()
        exports_dir = BASE_DIR / "data" / "exports" / outlet.platform / clean_outlet
        excel_files = list(exports_dir.glob("*.xlsx")) if exports_dir.exists() else []
        if excel_files:
            excel_exists = True
            mtime = datetime.fromtimestamp(os.path.getmtime(excel_files[0]))
            mtime_naive = mtime.replace(tzinfo=None) if mtime.tzinfo else mtime
            last_pulled_naive = last_pulled_at.replace(tzinfo=None) if last_pulled_at and last_pulled_at.tzinfo else last_pulled_at
            if not last_pulled_at or mtime_naive > last_pulled_naive:
                last_pulled_at = mtime

    if not last_pulled_at or not excel_exists:
        return {
            "has_cache": False,
            "is_valid_24h": False,
            "last_pulled_at": None,
            "cache_age_hours": None,
            "human_age": "Belum pernah ditarik"
        }

    now = datetime.utcnow()
    if hasattr(last_pulled_at, "tzinfo") and last_pulled_at.tzinfo is not None:
        last_pulled_at_naive = last_pulled_at.replace(tzinfo=None)
    else:
        last_pulled_at_naive = last_pulled_at

    age_seconds = (now - last_pulled_at_naive).total_seconds()
    age_hours = round(age_seconds / 3600.0, 1)
    is_valid_24h = age_hours <= 24.0

    if age_seconds < 60:
        human_age = "Baru saja"
    elif age_seconds < 3600:
        mins = int(age_seconds // 60)
        human_age = f"{mins} menit lalu"
    elif age_seconds < 86400:
        hrs = int(age_seconds // 3600)
        human_age = f"{hrs} jam lalu"
    else:
        days = int(age_seconds // 86400)
        human_age = f"{days} hari lalu"

    return {
        "has_cache": excel_exists,
        "is_valid_24h": is_valid_24h,
        "last_pulled_at": last_pulled_at.isoformat(),
        "cache_age_hours": age_hours,
        "human_age": human_age
    }

def parse_c5_template_headers(wb) -> dict:
    """
    Parses headers from an openpyxl Workbook, prioritizing raw string column headers
    and mapping ArrayFormula column headers only for unique formula calculation columns.
    """
    import re
    if 'Item' not in wb.sheetnames:
        return {}
    ws = wb['Item']
    header_map = {}
    
    # 1. Pass 1: standard string headers (e.g. Category at col 7, Item at col 9, Description at col 11)
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if isinstance(val, str) and val.strip():
            header_map[val.strip()] = col - 1
            
    # 2. Pass 2: formula headers (e.g. Current Slash Price (%) at col 33, Current Slash Price (Rp) at col 34)
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if hasattr(val, 'text'):
            m = re.search(r'\"([^\"]+)\"', val.text)
            if m:
                h_name = m.group(1).strip()
                if h_name not in header_map:
                    header_map[h_name] = col - 1
                    
    return header_map

MENU_ITEMS_CACHE = {}  # excel_path -> (mtime, raw_items)

def _format_epoch_ms_wib(epoch_ms: Union[int, float, str, None]) -> str:
    """Format epoch ms timestamp to human-readable WIB datetime string."""
    if not epoch_ms:
        return ""
    try:
        val = int(epoch_ms)
        # Check if timestamp is in seconds instead of milliseconds
        if val < 10000000000:
            val *= 1000
        utc_dt = datetime.fromtimestamp(val / 1000.0, tz=timezone.utc)
        wib_dt = utc_dt.astimezone(timezone(timedelta(hours=7)))
        return wib_dt.strftime("%d %b %Y, %H:%M WIB")
    except Exception:
        return str(epoch_ms)

def _extract_shopee_dish_promo_info(dish: dict) -> dict:
    """Extract promo details and lock status from Shopee dish API response dictionary."""
    if not isinstance(dish, dict):
        return {"is_flash_sale": False, "is_price_locked": False, "promo_details": None}

    fs = dish.get("flash_sale_dish_discount")
    if fs and isinstance(fs, dict):
        disc = fs.get("discount") or {}
        # Shopee price in nano-units / 100,000
        disc_p = float(disc.get("discount_price", 0)) / 100000.0 if disc.get("discount_price") else 0
        raw_pct = disc.get("discount_percentage", 0)
        pct_val = (float(raw_pct) / 100.0) if raw_pct else 0
        pct_str = f"{int(pct_val)}%" if pct_val.is_integer() else f"{pct_val:.1f}%"

        start_time_str = _format_epoch_ms_wib(fs.get("start_time") or disc.get("create_time"))
        end_time_str = _format_epoch_ms_wib(fs.get("end_time"))

        promo_details = {
            "type": "FLASH_SALE",
            "name": disc.get("flash_sale_dish_name") or disc.get("dish_name") or dish.get("name", ""),
            "discount_price": disc_p,
            "discount_percentage": pct_str,
            "stock": disc.get("stock", 0),
            "sold_num": disc.get("sold_num", 0),
            "limit_per_user": disc.get("limit_per_user", 0),
            "start_time": start_time_str,
            "end_time": end_time_str,
            "campaign_id": str(fs.get("campaign_id") or disc.get("timeslot_id") or "")
        }
        return {
            "is_flash_sale": True,
            "is_price_locked": True,
            "promo_type": "FLASH_SALE",
            "promo_value": pct_str if pct_val > 0 else (f"Rp {disc_p:,.0f}" if disc_p > 0 else "Flash Sale"),
            "promo_details": promo_details
        }

    it = dish.get("item_discount")
    if it and isinstance(it, dict):
        disc_p = float(it.get("discount_price", 0)) / 100000.0 if it.get("discount_price") else 0
        raw_pct = it.get("discount_percentage", 0)
        pct_val = (float(raw_pct) / 100.0) if raw_pct else 0
        pct_str = f"{int(pct_val)}%" if pct_val.is_integer() else f"{pct_val:.1f}%"

        start_time_str = _format_epoch_ms_wib(it.get("campaign_start_time"))
        end_time_str = _format_epoch_ms_wib(it.get("campaign_end_time"))

        promo_details = {
            "type": "ITEM_DISCOUNT",
            "discount_price": disc_p,
            "discount_percentage": pct_str,
            "stock": it.get("stock", 0),
            "sold_num": it.get("sold_num", 0),
            "limit_per_order": it.get("limit_per_order", 0),
            "start_time": start_time_str,
            "end_time": end_time_str,
            "campaign_id": str(it.get("campaign_id") or "")
        }
        return {
            "is_flash_sale": False,
            "is_price_locked": False,
            "promo_type": "PERCENTAGE" if pct_val > 0 else "NOMINAL",
            "promo_value": pct_str if pct_val > 0 else (f"Rp {disc_p:,.0f}" if disc_p > 0 else "Diskon Menu"),
            "promo_details": promo_details
        }

    return {"is_flash_sale": False, "is_price_locked": False, "promo_details": None}

def get_parsed_menu_items(excel_path: str) -> list:
    path = Path(excel_path)
    if not path.exists():
        return []
    try:
        mtime = path.stat().st_mtime
        cached = MENU_ITEMS_CACHE.get(excel_path)
        if cached and cached[0] == mtime:
            return cached[1]
            
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        if 'Item' not in wb.sheetnames:
            return []
        sheet = wb['Item']
        if sheet.max_row <= 1:
            return []
            
        header_map = parse_c5_template_headers(wb)
        
        required_cols = ['Item ID', 'Category', 'Item', 'Current Real Price (Rp)']
        for col in required_cols:
            if col not in header_map:
                logger.warning(f"Missing column '{col}' in Excel sheet mapping: {list(header_map.keys())}")
                return []
                
        # Also load data_only=True workbook to read evaluated/cached values if needed
        raw_items = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue
                
            def get_col_val(col_name, default=""):
                if col_name in header_map:
                    idx = header_map[col_name]
                    if idx < len(row) and row[idx] is not None:
                        return row[idx]
                return default
                
            item_id = str(get_col_val('Item ID', '')).strip()
            category_id = str(get_col_val('Category ID', '')).strip()
            category_name = str(get_col_val('Category', '')).strip()
            item_name = str(get_col_val('Item', '')).strip()
            desc = str(get_col_val('Description', '')).strip()
            
            p_val = get_col_val('Current Real Price (Rp)', 0)
            fake_p_val = get_col_val('Current Fake Price (Rp)', 0)
            try:
                price_val = int(float(p_val)) if p_val is not None and str(p_val).strip() != "" else 0
            except:
                price_val = 0

            try:
                fake_price_val = int(float(fake_p_val)) if fake_p_val is not None and str(fake_p_val).strip() != "" else 0
            except:
                fake_price_val = 0
                
            avail_val = str(get_col_val('Availability', 'Available')).strip() or "Available"
            is_promo_col = str(get_col_val('Sedang promo', '')).strip().lower()
            
            slash_pct_raw = str(get_col_val('Current Slash Price (%)', '')).strip()
            slash_rp_raw = get_col_val('Current Slash Price (Rp)', 0)
            try:
                slash_rp_val = int(float(slash_rp_raw)) if slash_rp_raw is not None and str(slash_rp_raw).strip() != "" else 0
            except:
                slash_rp_val = 0
            
            raw_items.append({
                "id": item_id,
                "category_id": category_id,
                "category": category_name,
                "name": item_name,
                "description": desc,
                "price": price_val,
                "original_price": fake_price_val,
                "availability": avail_val,
                "is_promo_col": is_promo_col,
                "slash_pct": slash_pct_raw,
                "slash_rp": slash_rp_val
            })
            
        MENU_ITEMS_CACHE[excel_path] = (mtime, raw_items)
        return raw_items
    except Exception as e:
        logger.error(f"Error parsing excel menu file at {excel_path}: {e}")
        return []

@app.get("/api/outlets/{outlet_id}/menu-items")
def get_outlet_menu_items(outlet_id: uuid.UUID, db: Session = Depends(get_db)):
    """Retrieve the menu items list of an outlet from the latest pulled Excel sheet catalog."""
    job = db.query(Job).filter(
        Job.outlet_id == outlet_id,
        Job.job_type.in_(["PULL", "PUSH_UPDATE"]),
        Job.status.in_(["SUCCESS", "PARTIAL_SUCCESS"])
    ).order_by(Job.completed_at.desc()).first()
    
    excel_path = None
    if job and job.result_metadata:
        excel_path = job.result_metadata.get("excel_path")
        
    if not excel_path or not os.path.exists(excel_path):
        # Fallback to scanning the exports folder directly
        outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
        if not outlet:
            raise HTTPException(status_code=404, detail="Outlet not found")
        import re
        raw_outlet = outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown'
        clean_outlet = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
        clean_outlet = re.sub(r'\s+', ' ', clean_outlet).lower()
        
        exports_dir = BASE_DIR / "data" / "exports" / outlet.platform / clean_outlet
        if exports_dir.exists():
            excel_files = sorted(exports_dir.glob("*.xlsx"), key=lambda p: os.path.getmtime(str(p)), reverse=True)
            cabang_clean = (outlet.cabang or "").strip().lower()
            store_id_clean = (outlet.store_id or "").strip().lower()
            matched_file = None
            if cabang_clean or store_id_clean:
                for ef in excel_files:
                    ef_name = ef.name.lower()
                    if (store_id_clean and store_id_clean in ef_name) or (cabang_clean and cabang_clean in ef_name):
                        matched_file = str(ef)
                        break
            excel_path = matched_file or (str(excel_files[0]) if excel_files else None)

    if not excel_path or not os.path.exists(excel_path):
        return []

    # Detect items locked by promo constraints from push failures since the last successful PULL job
    latest_pull = db.query(Job).filter(
        Job.outlet_id == outlet_id,
        Job.job_type == "PULL",
        Job.status.in_(["SUCCESS", "PARTIAL_SUCCESS"])
    ).order_by(Job.completed_at.desc()).first()
    latest_pull_time = latest_pull.completed_at if latest_pull and latest_pull.completed_at else datetime.min

    promo_item_ids = set()
    try:
        promo_trails = db.query(AuditTrail).filter(
            AuditTrail.outlet_id == outlet_id,
            AuditTrail.created_at >= latest_pull_time,
            (AuditTrail.error_message.ilike("%promo%") | AuditTrail.error_message.ilike("%campaign%"))
        ).all()
        for pt in promo_trails:
            if pt.item_id:
                promo_item_ids.add(str(pt.item_id))
    except Exception as ex:
        logger.warning(f"Error querying promo audit trails: {ex}")

    # If outlet is Shopee, load latest captured dishes JSON snapshot to enrich promo/flash-sale metadata
    shopee_dish_map = {}
    outlet_obj = db.query(Outlet).filter(Outlet.id == outlet_id).first()
    if outlet_obj and (outlet_obj.platform or "").lower() == "shopee":
        try:
            shopee_api_dir = BASE_DIR / "shopee" / "API"
            store_id_str = str(outlet_obj.store_id or "").strip()
            cands = []
            if store_id_str:
                cands.append(shopee_api_dir / f"menu-response-{store_id_str}.json")
            cands.append(shopee_api_dir / "menu-response.json")
            for cp in cands:
                if cp.exists():
                    with open(cp, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    catalogs = data.get("data", {}).get("catalogs", []) if isinstance(data, dict) else []
                    for cat in catalogs:
                        for dish in cat.get("dishes", []):
                            d_id = str(dish.get("id"))
                            if d_id:
                                shopee_dish_map[d_id] = dish
                    if shopee_dish_map:
                        break
        except Exception as ex:
            logger.warning(f"Failed loading Shopee menu snapshot: {ex}")

    raw_items = get_parsed_menu_items(excel_path)
    
    items = []
    for ri in raw_items:
        slash_pct_str = ri.get("slash_pct", "")
        slash_rp_val = ri.get("slash_rp", 0)
        
        # In catalog Excel (C5):
        # 'Current Fake Price (Rp)' (ri["original_price"]) is base price (e.g. 11.000)
        # 'Current Real Price (Rp)' (ri["price"]) is selling price after discount (e.g. 9.900)
        # If no fake price, base price is price.
        fake_p = ri.get("original_price", 0)
        real_p = ri.get("price", 0)
        
        base_p = fake_p if fake_p > 0 else real_p
        selling_p = real_p
        
        has_pct = bool(slash_pct_str and slash_pct_str not in ("0%", "0", ""))
        has_rp = bool(slash_rp_val and slash_rp_val > 0)
        
        has_promo_col = ri["is_promo_col"] in ("ya", "yes", "true", "1")
        has_price_diff = base_p > selling_p and selling_p > 0
        has_audit_lock = ri["id"] in promo_item_ids
        
        is_in_promo = has_pct or has_rp or has_promo_col or has_price_diff or has_audit_lock
        
        # Determine promo type and price lock status
        promo_type = "NONE"
        promo_value = ""
        is_price_locked = False
        is_flash_sale = False
        promo_details = None

        # Check Shopee live snapshot metadata first if present
        dish_snapshot = shopee_dish_map.get(str(ri["id"]))
        if dish_snapshot:
            shopee_promo = _extract_shopee_dish_promo_info(dish_snapshot)
            if shopee_promo.get("is_flash_sale") or shopee_promo.get("promo_details"):
                is_flash_sale = shopee_promo.get("is_flash_sale", False)
                is_price_locked = shopee_promo.get("is_price_locked", False)
                promo_type = shopee_promo.get("promo_type", "NONE")
                promo_value = shopee_promo.get("promo_value", "")
                promo_details = shopee_promo.get("promo_details")
                is_in_promo = True

        if not promo_details:
            name_lower = ri["name"].lower()
            is_name_nominal = "nominal" in name_lower
            is_name_percentage = "persen" in name_lower or "percentage" in name_lower or "%" in name_lower

            if is_name_nominal:
                promo_type = "NOMINAL"
                diff = (base_p - selling_p) if (base_p > selling_p and base_p > 0) else slash_rp_val
                promo_value = f"Rp {diff:,}" if diff > 0 else (f"Rp {slash_rp_val:,}" if slash_rp_val > 0 else "Nominal")
                is_price_locked = True
            elif is_name_percentage:
                promo_type = "PERCENTAGE"
                if has_pct:
                    promo_value = slash_pct_str
                elif base_p > selling_p and base_p > 0:
                    pct_calc = round(((base_p - selling_p) / base_p) * 100)
                    promo_value = f"{pct_calc}%"
                else:
                    promo_value = "%"
                is_price_locked = False
            elif has_pct:
                promo_type = "PERCENTAGE"
                promo_value = slash_pct_str
                is_price_locked = False  # Percentage promo allows base price changes
            elif has_rp:
                promo_type = "NOMINAL"
                promo_value = f"Rp {slash_rp_val:,}"
                is_price_locked = True   # Nominal promo locks base price
            elif is_in_promo:
                # Fallback deduction if % or Rp not explicitly formatted
                if has_price_diff:
                    diff = base_p - selling_p
                    pct_exact = (diff / base_p) * 100
                    # If exact standard round discount tier (e.g. 10%, 15%, 20%, 25%, 30%, 50%), treat as percentage
                    if abs(pct_exact - round(pct_exact)) < 0.01 and int(round(pct_exact)) % 5 == 0:
                        promo_type = "PERCENTAGE"
                        promo_value = f"{int(round(pct_exact))}%"
                        is_price_locked = False
                    else:
                        promo_type = "NOMINAL"
                        promo_value = f"Rp {diff:,}"
                        is_price_locked = True
                else:
                    promo_type = "NOMINAL"
                    promo_value = "Promo Aktif"
                    is_price_locked = False

        # Discounted selling price if item is in promo:
        disc_price = None
        if promo_details and promo_details.get("discount_price"):
            disc_price = int(promo_details["discount_price"])
        elif is_in_promo and selling_p > 0 and selling_p < base_p:
            disc_price = selling_p
        elif is_in_promo and (has_pct or has_rp):
            disc_price = selling_p

        items.append({
            "id": ri["id"],
            "category_id": ri["category_id"],
            "category": ri["category"],
            "name": ri["name"],
            "description": ri["description"],
            "price": base_p,                       # Base fake price to display & edit (e.g. 11.000)
            "original_price": base_p,              # Base normal price (11.000)
            "discounted_price": disc_price,        # Selling price after promo (e.g. 9.900)
            "slash_pct": ri.get("slash_pct", ""),
            "slash_rp": ri.get("slash_rp", 0),
            "availability": ri["availability"],
            "is_in_promo": is_in_promo,
            "is_flash_sale": is_flash_sale,
            "promo_type": promo_type,
            "promo_value": promo_value,
            "is_price_locked": is_price_locked,
            "promo_details": promo_details
        })
    return items

SESSION_METADATA_CACHE = {}  # filename -> (mtime, last_active_ts)
PHONE_MAP_CACHE = {"mtime": 0.0, "data": {}}

def get_cached_session_last_active(session_path: Path) -> Optional[str]:
    try:
        if not session_path.exists():
            return None
        mtime = session_path.stat().st_mtime
        filename = session_path.name
        
        cached = SESSION_METADATA_CACHE.get(filename)
        if cached and cached[0] == mtime:
            return cached[1]
            
        with open(session_path, "r") as f:
            data = json.load(f)
            ts = data.get("timestamp") or data.get("saved_at")
            
        SESSION_METADATA_CACHE[filename] = (mtime, ts)
        return ts
    except Exception:
        return None

def get_cached_phone_map(cache_path: Path) -> dict:
    if not cache_path.exists():
        return {}
    try:
        mtime = cache_path.stat().st_mtime
        if PHONE_MAP_CACHE["mtime"] == mtime:
            return PHONE_MAP_CACHE["data"]
            
        import pandas as pd
        df = pd.read_csv(cache_path)
        phone_cols = [col for col in df.columns if 'nomor hp' in str(col).lower()]
        col_phone = phone_cols[1] if len(phone_cols) > 1 else (phone_cols[0] if phone_cols else None)
        phone_map = {}
        if col_phone:
            for _, row in df.iterrows():
                sid = str(row.get('Store ID', '')).strip().split('.')[0]
                if not sid or sid == '-' or sid.lower() == 'nan':
                    sid = str(row.get('Merchant ID', '')).strip().split('.')[0]
                p_val = str(row.get(col_phone, '')).strip()
                if sid and p_val and p_val not in ('-', 'nan', ''):
                    phone_map[sid] = p_val
                    
        PHONE_MAP_CACHE["mtime"] = mtime
        PHONE_MAP_CACHE["data"] = phone_map
        return phone_map
    except Exception as e:
        logger.error(f"Error loading phone mapping: {e}")
        return PHONE_MAP_CACHE["data"]

@app.get("/api/sessions")
def get_sessions_status(db: Session = Depends(get_db)):
    """
    Returns the session ingestion status for all outlets (Shopee and GoFood).
    """
    import os
    import json
    import re
    from pathlib import Path
    
    cache_path = BASE_DIR / "master_merchants_cache.csv"
    phone_map = get_cached_phone_map(cache_path)

    # Use eager loading with joinedload to solve N+1 queries
    outlets = db.query(Outlet).options(joinedload(Outlet.account)).all()
    
    result = []
    for o in outlets:
        platform = o.platform
        if platform not in ("shopee", "gofood"):
            continue
            
        # Get phone number from mapping or fallback
        phone = phone_map.get(o.store_id)
        if not phone and o.account and "@" in o.account.username:
            phone = o.account.username
            
        status_info = {
            "id": str(o.id),
            "store_id": o.store_id,
            "merchant_name": o.merchant_name,
            "nama_outlet": o.nama_outlet,
            "nama_resto_final": o.nama_resto_final,
            "brand": o.brand,
            "platform": platform,
            "has_session": False,
            "session_file": None,
            "last_active": None,
            "phone": phone or o.store_id or "-"
        }
        
        if platform == "shopee":
            # Sanitize profile name
            merchant_name = o.merchant_name or o.nama_resto_final or o.nama_outlet or ''
            profile_name = re.sub(r'[^a-zA-Z0-9_]', '_', merchant_name)
            profile_name = re.sub(r'_+', '_', profile_name).strip('_').lower()
            
            session_file = BASE_DIR / "shopee" / "data" / f"session_{profile_name}.json"
            ts = get_cached_session_last_active(session_file)
            if ts is not None:
                status_info["has_session"] = True
                status_info["session_file"] = f"session_{profile_name}.json"
                status_info["last_active"] = ts
                
        elif platform == "gofood" and o.account:
            ident_str = str(o.account.username).strip().lower()
            sanitized = re.sub(r'[^a-zA-Z0-9_.-]', '_', ident_str)
            session_file = BASE_DIR / "Gofood" / f"session_gofood_{sanitized}.json"
            ts = get_cached_session_last_active(session_file)
            if ts is not None:
                status_info["has_session"] = True
                status_info["session_file"] = f"session_gofood_{sanitized}.json"
                status_info["last_active"] = ts
                
        result.append(status_info)
        
    return result


# ─── SHOPEE OTP ENDPOINTS ───────────────────────────────────────────────────

class ShopeeOTPRequest(BaseModel):
    username: str
    code: str
    channel: Optional[str] = "sms"

class ShopeeOTPChannelRequest(BaseModel):
    username: str
    channel: str  # "sms" | "whatsapp"

@app.post("/api/shopee/cancel-otp")
def cancel_shopee_otp(req: ShopeeOTPChannelRequest):
    """Cancels OTP waiting state for the given username and cleans up request files."""
    username = req.username.strip()
    if not username:
        raise HTTPException(status_code=400, detail="Username is required")

    shopee_data_dirs = [
        BASE_DIR / "src" / "shopee-omzet-automation" / "data",
        BASE_DIR / "shopee" / "data"
    ]
    for d in shopee_data_dirs:
        d.mkdir(parents=True, exist_ok=True)
        fpath = d / f"otp_request_{username}.json"
        request_data = {
            "status": "CANCELLED",
            "username": username,
            "error_msg": "user membatalkan otp",
            "cancelled_at": datetime.now().isoformat()
        }
        try:
            fpath.write_text(json.dumps(request_data, indent=2))
            log.info(f"🛑 [OTP] Updated OTP status to CANCELLED for {username} in {fpath}")
        except Exception as e:
            log.error(f"Error writing CANCELLED to OTP file {fpath}: {e}")

    try:
        from menu_core.database import SessionLocal
        db = SessionLocal()
        running_jobs = db.query(Job).filter(
            Job.status.in_(["RUNNING", "PENDING"]),
            Job.platform == "shopee"
        ).all()
        for j in running_jobs:
            if j.outlet and j.outlet.account and (j.outlet.account.username or "").strip().lower() == username.lower():
                j.status = "FAILED"
                j.error_message = "user membatalkan otp"
                j.current_step = "Gagal: user membatalkan otp"
                j.completed_at = datetime.utcnow()
                log.info(f"🛑 [OTP] Direct cancel DB update for Job {j.id}")
        db.commit()
        db.close()
    except Exception as dbe:
        log.error(f"Error updating DB for cancelled OTP: {dbe}")

    return {"status": "SUCCESS", "message": f"OTP request cancelled for {username}"}

@app.post("/api/shopee/select-otp-channel")
def select_shopee_otp_channel(req: ShopeeOTPChannelRequest):
    """Saves user's selected OTP channel (e.g. WhatsApp) to trigger channel switching in browser automation."""
    username = req.username.strip()
    channel = req.channel.strip().lower()
    if not username:
        raise HTTPException(status_code=400, detail="Username is required")

    shopee_data_dirs = [
        BASE_DIR / "src" / "shopee-omzet-automation" / "data",
        BASE_DIR / "shopee" / "data"
    ]
    for d in shopee_data_dirs:
        d.mkdir(parents=True, exist_ok=True)
        fpath = d / f"otp_request_{username}.json"
        if fpath.exists():
            try:
                data = json.loads(fpath.read_text())
                data["requested_channel"] = channel
                data["channel_requested_at"] = datetime.now().isoformat()
                fpath.write_text(json.dumps(data, indent=2))
            except Exception as e:
                log.error(f"Error updating OTP channel in {fpath}: {e}")
        else:
            request_data = {
                "status": "WAITING_OTP",
                "username": username,
                "requested_channel": channel,
                "requested_at": datetime.now().isoformat()
            }
            fpath.write_text(json.dumps(request_data, indent=2))

    return {"status": "SUCCESS", "channel": channel, "message": f"Channel {channel.upper()} selected for {username}"}

@app.get("/api/shopee/otp-status")
def get_shopee_otp_status(username: Optional[str] = None):
    """Checks whether Shopee login engine is currently waiting for an OTP for the given username."""
    shopee_data_dirs = [
        BASE_DIR / "src" / "shopee-omzet-automation" / "data",
        BASE_DIR / "shopee" / "data"
    ]
    
    usernames_to_check = [username] if username else []
    if not usernames_to_check:
        for d in shopee_data_dirs:
            if d.exists():
                for f in d.glob("otp_request_*.json"):
                    u = f.stem.replace("otp_request_", "")
                    if u not in usernames_to_check:
                        usernames_to_check.append(u)
    
    for u in usernames_to_check:
        for d in shopee_data_dirs:
            fpath = d / f"otp_request_{u}.json"
            if fpath.exists():
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        if data.get("status") == "WAITING_OTP":
                            return {
                                "waiting": True,
                                "username": u,
                                "phone": data.get("phone", ""),
                                "requested_at": data.get("requested_at"),
                                "error_msg": data.get("error_msg", "")
                            }
                except Exception: pass
    return {"waiting": False}

@app.post("/api/shopee/submit-otp")
def submit_shopee_otp(req: ShopeeOTPRequest):
    """Submits OTP code to Shopee login engine."""
    username = req.username.strip()
    code = req.code.strip()
    channel = (req.channel or "sms").strip().lower()
    if not username or not code:
        raise HTTPException(status_code=400, detail="Username and OTP code are required")

    shopee_data_dirs = [
        BASE_DIR / "src" / "shopee-omzet-automation" / "data",
        BASE_DIR / "shopee" / "data"
    ]
    
    for d in shopee_data_dirs:
        d.mkdir(parents=True, exist_ok=True)
        fpath = d / f"otp_request_{username}.json"
        request_data = {
            "status": "RECEIVED",
            "code": code,
            "username": username,
            "channel": channel,
            "received_at": datetime.now().isoformat()
        }
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(request_data, f, indent=2)
            
    return {"status": "SUCCESS", "message": f"OTP {code} ({channel.upper()}) berhasil dikirim untuk user {username}."}

