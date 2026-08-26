import os
import json
import asyncio
import re
import pandas as pd
from pathlib import Path
import openpyxl

def run_async(coro):
    """Safely run async coroutines inside a synchronous worker thread context."""
    new_loop = asyncio.new_event_loop()
    asyncio.set_event_loop(new_loop)
    try:
        return new_loop.run_until_complete(coro)
    finally:
        try:
            new_loop.close()
        except Exception:
            pass
        asyncio.set_event_loop(None)

def clean_name_str(s):
    """Helper to clean string for comparison (remove spaces, symbols, lowercase)."""
    return "".join(c for c in str(s).lower() if c.isalnum())

def extract_grab_menu(store_metadata: dict, output_dir: str):
    """
    Extracts GrabFood menu for a specific store.
    Downloads the entire menu under the account and filters for the target store_id.
    """
    username = store_metadata.get('username', '').strip()
    password = store_metadata.get('password', '').strip()
    store_id = store_metadata.get('store_id', '').strip()
    nama_resto = store_metadata.get('nama_resto_final') or store_metadata.get('nama_outlet') or ''
    brand = store_metadata.get('brand') or ''

    print(f"\n[GrabFood Menu Extractor]")
    print(f"[-] Target Outlet: {nama_resto} ({store_id})")

    if not username or not password:
        print("[!] Error: Username or password is empty.")
        return False, "Username/password kosong."

    # 1. Import Grab API Scraper
    try:
        import sys
        from pathlib import Path
        project_root = str(Path(__file__).resolve().parent.parent)
        if project_root not in sys.path:
            sys.path.insert(0, project_root)
        elif sys.path[0] != project_root:
            sys.path.remove(project_root)
            sys.path.insert(0, project_root)
        from grab.core.grab_api_scraper import run_api_download_for_portal
    except ImportError as e:
        print(f"[!] Error importing Grab Scraper: {e}")
        return False, f"Gagal mengimpor Grab Scraper: {e}"

    # 2. Run Playwright Grab download flow
    print(f"[*] Meluncurkan browser untuk menarik menu dari portal Grab Merchant...")
    try:
        json_path, error_msg = run_async(run_api_download_for_portal(username, password, target_store_id=store_id))
        if error_msg or not json_path or not os.path.exists(json_path):
            print(f"[!] Gagal menarik menu GrabFood: {error_msg}")
            return False, f"Gagal menarik menu GrabFood: {error_msg}"
    except Exception as e:
        print(f"[!] Terjadi pengecualian saat menjalankan scraper: {e}")
        return False, f"Terjadi kesalahan saat menjalankan scraper: {e}"

    print(f"   💾 Download JSON berhasil: {json_path}")

    # 3. Load downloaded JSON data
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            scraped_data = json.load(f)
    except Exception as e:
        print(f"[!] Gagal membaca data JSON menu GrabFood: {e}")
        return False, f"Gagal membaca data JSON menu Grab: {e}"

    # 4. Save structured Grab snapshot to grab/API/menu-response-<store_id>.json for Baseline comparison & Clean up temp file
    grab_api_dir = os.path.join(project_root, "grab", "API")
    os.makedirs(grab_api_dir, exist_ok=True)
    if store_id:
        try:
            snapshot_path = os.path.join(grab_api_dir, f"menu-response-{store_id}.json")
            with open(snapshot_path, "w", encoding="utf-8") as f:
                json.dump(scraped_data, f, indent=2)
            print(f"   💾 Snapshot menu Grab berhasil disimpan ke: {snapshot_path}")
        except Exception as e:
            print(f"   ⚠️ Gagal menyimpan snapshot menu Grab ke {grab_api_dir}: {e}")

    try:
        os.unlink(json_path)
    except Exception as e:
        print(f"   ⚠️ Gagal menghapus file unduhan sementara {json_path}: {e}")

    items = scraped_data.get('items', [])
    modifiers = scraped_data.get('modifiers', [])

    # 5. Filter items & modifiers for the target store_id (or fallback matching)
    matched_items = []
    for item in items:
        item_sid = str(item.get("Store ID", "")).strip()
        target_sid = str(store_id).strip()
        if target_sid and item_sid.lower() == target_sid.lower():
            matched_items.append(item)
        elif not target_sid and clean_name_str(item.get("Nama panjang", "")) == clean_name_str(nama_resto):
            matched_items.append(item)

    matched_mods = []
    for mod in modifiers:
        mod_sid = str(mod.get("Store ID", "")).strip()
        target_sid = str(store_id).strip()
        if target_sid and mod_sid.lower() == target_sid.lower():
            matched_mods.append(mod)
        elif not target_sid and clean_name_str(mod.get("Nama panjang", "")) == clean_name_str(nama_resto):
            matched_mods.append(mod)

    print(f"   📊 Berhasil menyaring menu:")
    print(f"      - Item Terkait: {len(matched_items)} dari {len(items)}")
    print(f"      - Modifier Terkait: {len(matched_mods)} dari {len(modifiers)}")

    if not matched_items:
        print(f"[!] Warning: Tidak ada item menu yang cocok untuk Store ID '{store_id}' / Nama '{nama_resto}'.")
        # Kita tetap buat file kosong tapi laporkan warning
        # return False, f"Tidak ada menu yang cocok dengan Store ID {store_id}"

    # 6. Build file name conforming to "O.C5 {nama_outlet} - {brand}.xlsx"
    def clean_filename_part(s):
        return "".join(c for c in s if c.isalnum() or c in (' ', '_', '-')).strip()

    raw_resto_final = store_metadata.get('nama_resto_final') or '' if isinstance(store_metadata, dict) else ''
    clean_outlet = clean_filename_part(nama_resto)
    clean_brand = clean_filename_part(brand)
    clean_resto = clean_filename_part(raw_resto_final)

    if clean_resto and clean_resto.lower() not in (clean_outlet.lower(), clean_brand.lower()):
        excel_filename = f"O.C5 {clean_resto} - {clean_brand}.xlsx"
    elif clean_brand and clean_brand.lower() != clean_outlet.lower():
        excel_filename = f"O.C5 {clean_outlet} - {clean_brand}.xlsx"
    else:
        excel_filename = f"O.C5 {clean_outlet}.xlsx"

    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, excel_filename)

    # 7. Write to O. C5 Template Excel
    BASE_DIR = Path(__file__).resolve().parents[1]
    template_path = BASE_DIR / "O. C5 Template.xlsx"

    # ── Kolom yang bisa diisi user secara manual (akan di-preserve saat re-pull) ──
    USER_EDITABLE_COLS = [
        'Keyword',                   # Kol L
        'Item Name Improvement',     # Kol M
        'Design Improvement',        # Kol N
        'Offline Price (Rp)',        # Kol W
        'Offline Adjustment (Rp)',   # Kol X
        'Current Adjustment (Rp)',   # Kol AC
        'New Markup (%)',            # Kol AI
        'New Adjustment (Rp)',       # Kol AK
        'New Slash Price (%)',       # Kol AM
        'Notes',                     # Kol AP
    ]

    # ── Load data user dari C5 lama (jika ada) ──
    prev_user_data = {}  # { item_id: { col_name: value } }
    if os.path.exists(excel_path):
        try:
            wb_prev = openpyxl.load_workbook(excel_path, data_only=True)
            if 'Item' in wb_prev.sheetnames:
                sheet_prev = wb_prev['Item']
                prev_headers = {}
                for cell in sheet_prev[1]:
                    if cell.value:
                        prev_headers[str(cell.value).strip()] = cell.column - 1  # 0-indexed
                item_id_idx = prev_headers.get('Item ID')
                edit_col_indices = {col: prev_headers[col] for col in USER_EDITABLE_COLS if col in prev_headers}
                if item_id_idx is not None:
                    for row_cells in sheet_prev.iter_rows(min_row=2, values_only=True):
                        raw_iid = row_cells[item_id_idx] if item_id_idx < len(row_cells) else None
                        if raw_iid is None or str(raw_iid).strip() in ('', 'None', 'nan'):
                            continue
                        iid = str(raw_iid).strip()
                        row_data = {}
                        for col_name, col_idx in edit_col_indices.items():
                            if col_idx < len(row_cells):
                                val = row_cells[col_idx]
                                if val is not None and str(val).strip() not in ('', 'None', 'nan'):
                                    row_data[col_name] = val
                        if row_data:
                            prev_user_data[iid] = row_data
            wb_prev.close()
            if prev_user_data:
                print(f"   🔄 Ditemukan C5 sebelumnya — memuat {len(prev_user_data)} baris data user untuk di-preserve.")
        except Exception as prev_err:
            print(f"   ⚠️ Gagal membaca C5 sebelumnya untuk merge: {prev_err}")

    try:
        wb = openpyxl.load_workbook(template_path)
        
        # 1. Fill Item Sheet
        sheet_item = wb['Item']
        # Delete sample data rows (from row 2 onwards)
        if sheet_item.max_row > 1:
            sheet_item.delete_rows(2, sheet_item.max_row - 1)
            
        headers_item = {cell.value: cell.column for cell in sheet_item[1]}
        
        new_row_idx = 2
        for item in matched_items:
            avail_str = str(item.get('Ketersediaan item', '')).lower()
            availability = "Available" if avail_str in ("available", "active", "1", "true") else "Unavailable"
            
            fake_price = item.get('Harga item sebelum promo (harga coret)', 0.0)
            real_price = item.get('Harga item setelah promo (harga coret)', 0.0)
            slash_rp = item.get('Nominal atau persentase promo (harga coret)', 0.0)
            
            if fake_price > real_price and fake_price > 0:
                pct = round(((fake_price - real_price) / fake_price) * 100.0, 2)
                slash_pct = f"{int(pct)}%" if pct.is_integer() else f"{pct}%"
            else:
                slash_pct = "0%"
                slash_rp = 0

            # Lookup data user lama berdasarkan Item ID
            item_id_str = str(item.get('Item ID', '')).strip()
            prev_row = prev_user_data.get(item_id_str, {})
                
            mapping = {
                'OFD': 'GrabFood',
                'Outlet Name': item.get('Nama panjang', nama_resto),
                'Outlet Short Name': brand or item.get('Nama panjang', nama_resto),
                'Outlet Link': item.get('Link outlet', f"https://food.grab.com/id/en/restaurant/{store_id}"),
                'SID': item.get('Store ID', store_id),
                'Category ID': item.get('Category ID', ''),
                'Category': item.get('Nama kategori', ''),
                'Item ID': item.get('Item ID', ''),
                'Item': item.get('Nama item', ''),
                'Photo Link': item.get('Link foto', ''),
                'Description': item.get('Deskripsi item', ''),
                'Keyword': '',
                'Total Sold': item.get('Jumlah terjual', 0),
                'Total Modifier Group': item.get('Jumlah modifier group', 0),
                'Total Modifier': item.get('Jumlah modifier', 0),
                'Availability': availability,
                'Visibility': 'Show',
                'Current Fake Price (Rp)': fake_price,
                'Current Real Price (Rp)': real_price,
                'Current Slash Price (%)': slash_pct,
                'Current Slash Price (Rp)': slash_rp
            }

            # Tambahkan kolom user-editable dari C5 lama jika ada
            for col_name in USER_EDITABLE_COLS:
                if col_name not in mapping and col_name in prev_row:
                    mapping[col_name] = prev_row[col_name]
            
            for key, val in mapping.items():
                if key in headers_item:
                    col_idx = headers_item[key]
                    if pd.isna(val):
                        val = ""
                    elif key in ['SID', 'Category ID', 'Item ID']:
                        if isinstance(val, float):
                            val = str(int(val)) if val.is_integer() else str(val)
                        else:
                            val = str(val)
                    # Jika kolom ini bisa diisi user & nilai baru kosong, restore nilai lama
                    if key in USER_EDITABLE_COLS and (val == "" or val is None):
                        val = prev_row.get(key, val)
                    sheet_item.cell(row=new_row_idx, column=col_idx, value=val)
            new_row_idx += 1
                    
        # 2. Fill Modifier Sheet
        sheet_mod = wb['Modifier']
        # Delete sample data rows (from row 2 onwards)
        if sheet_mod.max_row > 1:
            sheet_mod.delete_rows(2, sheet_mod.max_row - 1)
            
        headers_mod = {cell.value: cell.column for cell in sheet_mod[1]}
        
        new_row_idx = 2
        for mod in matched_mods:
            mod_avail_str = str(mod.get('Ketersediaan modifier', '')).lower()
            mod_availability = "Available" if mod_avail_str in ("available", "active", "1", "true") else "Unavailable"
            
            mapping_mod = {
                'OFD': 'GrabFood',
                'Outlet Name': mod.get('Nama panjang', nama_resto),
                'Outlet Short Name': brand or mod.get('Nama panjang', nama_resto),
                'Outlet Link': mod.get('Link outlet', f"https://food.grab.com/id/en/restaurant/{store_id}"),
                'SID': mod.get('Store ID', store_id),
                'Item': mod.get('Nama item', ''),
                'Modifier Group ID': mod.get('Modifier Group ID', ''),
                'Modifier Group': mod.get('Nama modifier group', ''),
                'Modifier ID': mod.get('Modifier ID', ''),
                'Modifier': mod.get('Nama modifier', ''),
                'Min': mod.get('Minimal', 0),
                'Max': mod.get('Maksimal', 1),
                'Availability': mod_availability,
                'Visibility': 'Show',
                'Current Price (Rp)': mod.get('Harga modifier', 0.0)
            }
            
            for key, val in mapping_mod.items():
                if key in headers_mod:
                    col_idx = headers_mod[key]
                    if pd.isna(val):
                        val = ""
                    elif key in ['SID', 'Modifier Group ID', 'Modifier ID', 'Item']:
                        if isinstance(val, float):
                            val = str(int(val)) if val.is_integer() else str(val)
                        else:
                            val = str(val)
                    sheet_mod.cell(row=new_row_idx, column=col_idx, value=val)
            new_row_idx += 1
                    
        wb.save(excel_path)
        if prev_user_data:
            print(f"   ✅ Berhasil menyimpan file catalog dengan data user yang di-preserve ke: {excel_path}")
        else:
            print(f"   ✅ Berhasil menyimpan file catalog menggunakan template O.C5 ke: {excel_path}")
    except Exception as ex_err:
        print(f"   [-] Gagal menulis ke template O.C5: {ex_err}. Fallback ke Excel biasa.")
        # Fallback to standard excel writer if template fails
        df_items_fallback = pd.DataFrame(matched_items)
        df_mods_fallback = pd.DataFrame(matched_mods)
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_items_fallback.to_excel(writer, sheet_name='Item', index=False)
            df_mods_fallback.to_excel(writer, sheet_name='Modifier', index=False)

    return True, {
        'items_csv': None,
        'mods_csv': None,
        'excel': excel_path,
        'items_count': len(matched_items),
        'mods_count': len(matched_mods)
    }
