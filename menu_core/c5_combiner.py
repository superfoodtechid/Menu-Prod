import os
import re
import openpyxl
import pandas as pd

FILE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(FILE_DIR, "O. C5 Template.xlsx")

def _get_ofd_sort_order(ofd_val):
    v = str(ofd_val or '').strip().lower()
    if 'go' in v:
        return 1
    if 'grab' in v:
        return 2
    if 'shopee' in v or 'shope' in v:
        return 3
    return 4


def combine_c5(excel_paths, output_path):
    """
    Combines multiple C5 Excel files (Item & Modifier sheets) into a single C5 Excel file.
    """
    all_items = []
    all_mods = []
    for f in excel_paths:
        if f and os.path.exists(f):
            try:
                df_item = pd.read_excel(f, sheet_name='Item')
                df_mod = pd.read_excel(f, sheet_name='Modifier')
                all_items.append(df_item)
                all_mods.append(df_mod)
            except Exception as e:
                print(f"[C5 Combiner] Error reading {f}: {e}")
                
    if not all_items:
        return False
        
    df_combined_items = pd.concat(all_items, ignore_index=True)
    df_combined_mods = pd.concat(all_mods, ignore_index=True)

    # Sort deterministically by platform (OFD): GoFood (1) -> GrabFood (2) -> ShopeeFood (3)
    if 'OFD' in df_combined_items.columns:
        df_combined_items['_ofd_sort'] = df_combined_items['OFD'].map(_get_ofd_sort_order)
        sort_cols = [c for c in ['_ofd_sort', 'Outlet Name', 'Category', 'Item'] if c in df_combined_items.columns]
        df_combined_items = df_combined_items.sort_values(by=sort_cols, kind='stable').drop(columns=['_ofd_sort']).reset_index(drop=True)

    if 'OFD' in df_combined_mods.columns:
        df_combined_mods['_ofd_sort'] = df_combined_mods['OFD'].map(_get_ofd_sort_order)
        df_combined_mods = df_combined_mods.sort_values(by=['_ofd_sort'], kind='stable').drop(columns=['_ofd_sort']).reset_index(drop=True)
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    if os.path.exists(TEMPLATE_PATH):
        try:
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
            sheet_item = wb['Item']
            if sheet_item.max_row > 1:
                sheet_item.delete_rows(2, sheet_item.max_row - 1)
                
            headers_item = {}
            for cell in sheet_item[1]:
                if isinstance(cell.value, str):
                    headers_item[cell.value] = cell.column
                    
            for r_idx, row in df_combined_items.iterrows():
                for col_name, val in row.items():
                    if col_name in headers_item:
                        if pd.isna(val):
                            val = ""
                        elif col_name in ['SID', 'Category ID', 'Item ID']:
                            if isinstance(val, float):
                                val = str(int(val)) if val.is_integer() else str(val)
                            else:
                                val = str(val)
                        cell = sheet_item.cell(row=r_idx + 2, column=headers_item[col_name], value=val)
                        
            # Apply percentage formatting to columns with '(%)' in the header
            for cell in sheet_item[1]:
                val_str = cell.value.text if hasattr(cell.value, 'text') else str(cell.value or "")
                if '(%)' in val_str:
                    for r in range(2, sheet_item.max_row + 1):
                        sheet_item.cell(row=r, column=cell.column).number_format = '0%'
                        
            sheet_mod = wb['Modifier']
            if sheet_mod.max_row > 1:
                sheet_mod.delete_rows(2, sheet_mod.max_row - 1)
                
            headers_mod = {}
            for cell in sheet_mod[1]:
                if isinstance(cell.value, str):
                    headers_mod[cell.value] = cell.column
                    
            for r_idx, row in df_combined_mods.iterrows():
                for col_name, val in row.items():
                    if col_name in headers_mod:
                        if pd.isna(val):
                            val = ""
                        elif col_name in ['SID', 'Modifier Group ID', 'Modifier ID', 'Item']:
                            if isinstance(val, float):
                                val = str(int(val)) if val.is_integer() else str(val)
                            else:
                                val = str(val)
                        sheet_mod.cell(row=r_idx + 2, column=headers_mod[col_name], value=val)
                        
            for cell in sheet_mod[1]:
                val_str = cell.value.text if hasattr(cell.value, 'text') else str(cell.value or "")
                if '(%)' in val_str:
                    for r in range(2, sheet_mod.max_row + 1):
                        sheet_mod.cell(row=r, column=cell.column).number_format = '0%'
                        
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"[C5 Combiner] Failed writing to template, fallback to raw pandas export: {e}")
            
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_combined_items.to_excel(writer, sheet_name='Item', index=False)
            df_combined_mods.to_excel(writer, sheet_name='Modifier', index=False)
        return True
    except Exception as ex:
        print(f"[C5 Combiner] Export error: {ex}")
        return False
